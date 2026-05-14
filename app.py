from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO
import csv
import hashlib
import threading
import time
import json
import shutil
from pathlib import Path
import re
from functools import wraps
from zipfile import ZIP_DEFLATED, ZipFile

from flask import Flask, abort, flash, has_request_context, jsonify, redirect, render_template, request, send_file, send_from_directory, session, url_for
from jinja2.runtime import Undefined
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from sqlalchemy import event, func, inspect, or_, text
from sqlalchemy.orm import Session, with_loader_criteria
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from config.enterprise_form_config import (
    COMMON_ENTERPRISE_FIELD_GROUPS,
    ENTRY_SOURCE_OPTIONS,
    ENTRY_STAGE_OPTIONS,
    ENTERPRISE_DEFAULT_STATUS,
    INDUSTRY_EXTRA_FIELD_CONFIG,
)
from config.industry_config import INDUSTRY_MAP, INDUSTRY_OPTIONS
from config.product_form_config import COMMON_PRODUCT_FIELD_GROUPS, INDUSTRY_PRODUCT_EXTRA_FIELD_CONFIG
from models import (
    AuditSoftDeleteMixin,
    BackupRecord,
    Contact,
    Document,
    Enterprise,
    Product,
    ProductSKU,
    AuditLog,
    ExportLog,
    ImportBatch,
    ImportError,
    ImportSupplementItem,
    User,
    db,
)

BASE_DIR = Path(__file__).resolve().parent
PER_PAGE = 10
BLOCKED_EXTENSIONS = {".exe", ".bat", ".js", ".sh", ".cmd", ".com", ".msi", ".ps1"}
MAX_UPLOAD_SIZE = 100 * 1024 * 1024
DEPRECATED_ENTERPRISE_EXTRA_FIELD_KEYS = {"recommendation_level", "recommended_for_pool"}


def 清理企业扩展字段(扩展字段):
    """移除企业库已下线展示字段，保持扩展字段结构与页面一致。"""

    if not 扩展字段:
        return {}
    清理后字段 = dict(扩展字段)
    for 键名 in DEPRECATED_ENTERPRISE_EXTRA_FIELD_KEYS:
        清理后字段.pop(键名, None)
    return 清理后字段


def 行业默认名称(行业代码):
    item = INDUSTRY_MAP.get((行业代码 or "").strip())
    if not item:
        return ""
    return item["name"]


def 行业显示名称(行业代码, 行业名称=None):
    """展示行业名称；当存储值就是行业名称时避免重复显示。"""

    行业代码 = (行业代码 or "").strip()
    行业名称 = (行业名称 or "").strip() or 行业默认名称(行业代码)
    if 行业名称:
        return 行业名称
    return 行业代码


def 估算人员规模(employee_count):
    if not employee_count:
        return ""
    if employee_count <= 20:
        return "1-20人"
    if employee_count <= 50:
        return "21-50人"
    if employee_count <= 100:
        return "51-100人"
    if employee_count <= 300:
        return "101-300人"
    if employee_count <= 500:
        return "301-500人"
    if employee_count <= 1000:
        return "501-1000人"
    return "1000人以上"


def 导出企业资料完整度标签(enterprise):
    ext = 清理企业扩展字段(enterprise.enterprise_extra_fields or {})
    required_values = [
        enterprise.project_owner,
        ext.get("company_full_name") or enterprise.company_name,
        ext.get("registered_name") or enterprise.company_name,
        ext.get("legal_representative"),
        ext.get("unified_social_credit_code") or enterprise.unified_social_credit_code,
        ext.get("registered_capital") or enterprise.registered_capital,
        ext.get("company_type") or enterprise.company_type,
        ext.get("founded_date") or (enterprise.founded_date.strftime("%Y-%m-%d") if enterprise.founded_date else ""),
        ext.get("business_term_start"),
        ext.get("is_listed_or_pre_ipo"),
        ext.get("has_import_export_qualification"),
    ]
    done = sum(1 for value in required_values if value and str(value).strip())
    score = int(round((done / len(required_values)) * 100)) if required_values else 0
    return f"{score}%"


PRODUCT_FORM_SECTIONS = [
    ("A", "所属企业"),
    ("B", "产品基础信息"),
    ("C", "规格参数"),
    ("D", "生产与供货能力"),
    ("E", "价格与贸易条款"),
    ("F", "认证与合规"),
    ("G", "包装与物流"),
    ("H", "市场与卖点"),
    ("I", "售后与备注"),
]

ENTERPRISE_UPLOAD_TYPES = [
    "企业介绍",
    "营业执照",
    "BP/融资材料",
    "工厂照片",
    "资质荣誉",
    "财务/经营资料",
    "其他",
]

PRODUCT_UPLOAD_TYPES = [
    "产品图片",
    "产品目录",
    "报价单",
    "认证证书",
    "检测报告",
    "说明书",
    "包装图",
    "宣传视频",
    "样品资料",
    "其他",
]
PRODUCT_MATERIAL_TYPE_MAPPING = [
    ("产品图片", "产品图片"),
    ("产品目录", "产品目录"),
    ("报价单", "报价单"),
    ("认证证书", "认证证书"),
    ("检测报告", "检测报告"),
    ("说明书", "说明书"),
    ("包装图", "包装图"),
    ("宣传视频", "宣传视频"),
    ("样品资料", "样品资料"),
    ("其他", "其他"),
]
PRODUCT_MATERIAL_STATUS_FIELD_MAPPING = {
    "media_product_images": "产品图片",
    "media_product_video": "宣传视频",
    "media_manual": "说明书",
    "media_brochure": "产品目录",
    "media_english_ppt": "产品目录",
    "media_case_study": "样品资料",
}

ENTERPRISE_DOCUMENT_TYPE_OPTIONS = [(item, item) for item in ENTERPRISE_UPLOAD_TYPES]
PRODUCT_DOCUMENT_TYPE_OPTIONS = [(item, item) for item in PRODUCT_UPLOAD_TYPES]
DOCUMENT_TYPE_OPTIONS = [(item, item) for item in dict.fromkeys(ENTERPRISE_UPLOAD_TYPES + PRODUCT_UPLOAD_TYPES)]
ENTERPRISE_SUB_FOLDERS = [
    "01_企业基础资料",
    "02_企业资质与合规文件",
    "03_企业经营与外贸能力",
    "04_产品目录与产品信息表",
    "05_产品规格与技术资料",
    "06_产品认证与检测报告",
    "07_报价与交易条件",
    "08_包装物流与交付资料",
    "09_宣传展示资料",
    "10_图片视频与样品资料",
    "11_客户案例与项目案例",
    "12_合作需求与撮合记录",
    "13_风险审核与归档确认",
]

DOCUMENT_FOLDER_MAPPING = {
    "企业介绍": "09_宣传展示资料",
    "营业执照": "01_企业基础资料",
    "BP/融资材料": "03_企业经营与外贸能力",
    "工厂照片": "10_图片视频与样品资料",
    "资质荣誉": "02_企业资质与合规文件",
    "财务/经营资料": "03_企业经营与外贸能力",
    "其他": "13_风险审核与归档确认",
    # 历史分类兼容：仅用于已上传存量附件的目录解析，不再出现在上传下拉框。
    "企业宣传册": "09_宣传展示资料",
    "企业介绍PPT": "09_宣传展示资料",
    "资质证书包": "02_企业资质与合规文件",
    "工厂/办公照片": "10_图片视频与样品资料",
    "企业名片": "01_企业基础资料",
    "其他补充资料": "13_风险审核与归档确认",
    "企业简介": "01_企业基础资料",
    "企业资质": "02_企业资质与合规文件",
    "合规文件": "02_企业资质与合规文件",
    "经营资料": "03_企业经营与外贸能力",
    "外贸资料": "03_企业经营与外贸能力",
    "产品目录": "04_产品目录与产品信息表",
    "产品信息表": "04_产品目录与产品信息表",
    "产品规格书": "05_产品规格与技术资料",
    "产品说明书": "05_产品规格与技术资料",
    "说明书": "05_产品规格与技术资料",
    "技术资料": "05_产品规格与技术资料",
    "检测报告": "06_产品认证与检测报告",
    "认证证书": "06_产品认证与检测报告",
    "产品认证": "06_产品认证与检测报告",
    "报价单": "07_报价与交易条件",
    "交易条件": "07_报价与交易条件",
    "包装图": "08_包装物流与交付资料",
    "包装资料": "08_包装物流与交付资料",
    "物流资料": "08_包装物流与交付资料",
    "宣传PPT": "09_宣传展示资料",
    "英文资料": "09_宣传展示资料",
    "英文PPT": "09_宣传展示资料",
    "宣传册": "09_宣传展示资料",
    "尽调照片": "10_图片视频与样品资料",
    "尽调视频": "10_图片视频与样品资料",
    "产品图片": "10_图片视频与样品资料",
    "宣传视频": "10_图片视频与样品资料",
    "产品视频": "10_图片视频与样品资料",
    "样品资料": "10_图片视频与样品资料",
    "案例资料": "11_客户案例与项目案例",
    "客户案例": "11_客户案例与项目案例",
    "项目案例": "11_客户案例与项目案例",
    "合作需求": "12_合作需求与撮合记录",
    "撮合记录": "12_合作需求与撮合记录",
    "合同协议": "13_风险审核与归档确认",
    "风险审核": "13_风险审核与归档确认",
    "归档确认": "13_风险审核与归档确认",
    "其他文件": "13_风险审核与归档确认",
}

ENTERPRISE_TYPE_ALLOWED_EXTENSIONS = {
    "企业介绍": {".pdf", ".ppt", ".pptx", ".doc", ".docx", ".jpg", ".jpeg", ".png", ".webp"},
    "营业执照": {".pdf", ".jpg", ".jpeg", ".png"},
    "BP/融资材料": {".pdf", ".ppt", ".pptx", ".doc", ".docx", ".xls", ".xlsx"},
    "工厂照片": {".jpg", ".jpeg", ".png", ".webp", ".mp4", ".mov", ".avi", ".mkv", ".webm"},
    "资质荣誉": {".pdf", ".jpg", ".jpeg", ".png", ".webp", ".zip"},
    "财务/经营资料": {".pdf", ".doc", ".docx", ".xls", ".xlsx", ".csv", ".zip"},
    "其他": {".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx", ".jpg", ".jpeg", ".png", ".webp", ".zip", ".rar", ".7z", ".txt", ".csv", ".mp4", ".mov"},
}

PROJECT_STAGE_OPTIONS = [
    "待补充资料",
    "已完成入库",
    "已推荐给外资",
    "外资初步感兴趣",
    "已发送资料",
    "已寄送样品",
    "已报价",
    "商务谈判中",
    "合同签署中",
    "已成交",
    "暂停/终止",
]

SAMPLE_STATUS_OPTIONS = [
    "未涉及",
    "待寄样",
    "已寄样",
    "样品反馈中",
    "样品通过",
    "样品未通过",
]

QUOTATION_STATUS_OPTIONS = [
    "未报价",
    "已初步报价",
    "已更新报价",
    "报价接受",
    "报价未接受",
]

CONTRACT_STATUS_OPTIONS = [
    "未开始",
    "起草中",
    "审核中",
    "已签署",
    "未成交",
]


def create_app():
    app = Flask(__name__)
    app.config["SECRET_KEY"] = "replace-with-a-secure-secret-key"
    app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{BASE_DIR / 'trade_agent.db'}"
    app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
    app.config["UPLOAD_ROOT"] = BASE_DIR / "uploads" / "企业库"
    app.config["UPLOAD_ENTERPRISE_ROOT"] = BASE_DIR / "uploads" / "企业库"
    app.config["MAX_CONTENT_LENGTH"] = MAX_UPLOAD_SIZE
    app.config["BACKUP_ROOT"] = BASE_DIR / "backups"
    app.config["DATABASE_BACKUP_ROOT"] = app.config["BACKUP_ROOT"] / "database"
    app.config["FILES_BACKUP_ROOT"] = app.config["BACKUP_ROOT"] / "files"
    app.config["BACKUP_RETENTION_DAYS"] = 30
    app.config["BACKUP_RETENTION_COUNT"] = 30

    db.init_app(app)

    if not getattr(db, "_soft_delete_filter_registered", False):
        @event.listens_for(Session, "do_orm_execute")
        def _add_soft_delete_filter(execute_state):
            if execute_state.is_select and not execute_state.execution_options.get("include_deleted"):
                execute_state.statement = execute_state.statement.options(
                    with_loader_criteria(
                        AuditSoftDeleteMixin,
                        lambda cls: cls.is_deleted.is_(False),
                        include_aliases=True,
                    )
                )

        @event.listens_for(Session, "before_flush")
        def _fill_audit_fields(session_obj, flush_context, instances):
            operator = "system"
            if has_request_context():
                operator = session.get("用户") or "system"
            for obj in session_obj.new:
                if isinstance(obj, AuditSoftDeleteMixin):
                    if not obj.created_by:
                        obj.created_by = operator
                    if not obj.updated_by:
                        obj.updated_by = operator
            for obj in session_obj.dirty:
                if isinstance(obj, AuditSoftDeleteMixin):
                    obj.updated_by = operator

        db._soft_delete_filter_registered = True

    def 当前用户():
        用户名 = session.get("用户")
        if not 用户名:
            return None
        return User.query.filter_by(username=用户名).first()

    def 记录审计日志(action, target_type, target_id=None, detail=None):
        用户 = 当前用户()
        db.session.add(
            AuditLog(
                action=action,
                target_type=target_type,
                target_id=target_id,
                user_name=用户.username if 用户 else "system",
                detail=detail,
            )
        )

    def 软删除对象(对象):
        对象.soft_delete(当前操作者名称())

    def login_required(view_func):
        @wraps(view_func)
        def wrapped(*args, **kwargs):
            if not session.get("用户"):
                flash("请先登录后再访问。", "warning")
                return redirect(url_for("登录"))
            return view_func(*args, **kwargs)

        return wrapped

    def admin_required(view_func):
        @wraps(view_func)
        def wrapped(*args, **kwargs):
            用户 = 当前用户()
            if not 用户:
                flash("请先登录后再访问。", "warning")
                return redirect(url_for("登录"))
            if 用户.role != "管理员":
                flash("仅管理员可执行此操作。", "danger")
                return redirect(url_for("dashboard"))
            return view_func(*args, **kwargs)

        return wrapped

    def 计算文件哈希(文件路径):
        sha256 = hashlib.sha256()
        with open(文件路径, "rb") as 文件:
            for 数据块 in iter(lambda: 文件.read(1024 * 1024), b""):
                sha256.update(数据块)
        return sha256.hexdigest()

    def 当前操作者名称():
        if not has_request_context():
            return "system"
        用户 = 当前用户()
        return 用户.username if 用户 else (session.get("用户") or "system")

    def 登记备份记录(备份类型, 文件路径, status="成功", error_message=None, created_by=None):
        文件路径 = Path(文件路径)
        相对路径 = str(文件路径.relative_to(BASE_DIR)) if 文件路径.is_absolute() and 文件路径.is_relative_to(BASE_DIR) else str(文件路径)
        记录 = BackupRecord(
            backup_type=备份类型,
            file_path=相对路径,
            file_size=文件路径.stat().st_size if 文件路径.exists() else 0,
            file_hash=计算文件哈希(文件路径) if 文件路径.exists() else "",
            status=status,
            error_message=error_message,
            created_by=created_by or 当前操作者名称(),
            updated_by=created_by or 当前操作者名称(),
        )
        db.session.add(记录)
        return 记录

    def 清理备份保留策略(备份类型=None):
        保留天数 = int(app.config.get("BACKUP_RETENTION_DAYS", 30))
        保留数量 = int(app.config.get("BACKUP_RETENTION_COUNT", 30))
        截止时间 = datetime.utcnow() - timedelta(days=保留天数)
        查询 = BackupRecord.query.filter(BackupRecord.is_deleted.is_(False))
        if 备份类型:
            查询 = 查询.filter(BackupRecord.backup_type == 备份类型)
        记录列表 = 查询.order_by(BackupRecord.created_at.desc()).all()
        for 序号, 记录 in enumerate(记录列表):
            if (记录.created_at and 记录.created_at < 截止时间) or 序号 >= 保留数量:
                文件路径 = BASE_DIR / 记录.file_path
                if 文件路径.exists() and 文件路径.is_file():
                    文件路径.unlink()
                记录.soft_delete("system")

    def 构建数据库备份(created_by=None):
        备份目录 = app.config["DATABASE_BACKUP_ROOT"]
        备份目录.mkdir(parents=True, exist_ok=True)
        时间戳 = datetime.now().strftime("%Y%m%d_%H%M%S")
        源数据库 = BASE_DIR / "trade_agent.db"
        备份路径 = 备份目录 / f"database_backup_{时间戳}.sqlite"
        shutil.copy2(源数据库, 备份路径)
        登记备份记录("database", 备份路径, created_by=created_by)
        清理备份保留策略("database")
        return 备份路径

    def 构建上传目录备份(created_by=None):
        备份目录 = app.config["FILES_BACKUP_ROOT"]
        备份目录.mkdir(parents=True, exist_ok=True)
        时间戳 = datetime.now().strftime("%Y%m%d_%H%M%S")
        备份路径 = 备份目录 / f"uploads_backup_{时间戳}.zip"
        上传根目录 = BASE_DIR / "uploads"
        with ZipFile(备份路径, "w", compression=ZIP_DEFLATED) as zipf:
            if 上传根目录.exists():
                for 文件路径 in 上传根目录.rglob("*"):
                    if 文件路径.is_file():
                        zipf.write(文件路径, arcname=文件路径.relative_to(上传根目录.parent))
        登记备份记录("uploads", 备份路径, created_by=created_by)
        清理备份保留策略("uploads")
        return 备份路径

    def 执行每日自动备份(created_by="system"):
        数据库文件 = None
        上传文件 = None
        if not 今日是否已有备份("database"):
            数据库文件 = 构建数据库备份(created_by=created_by)
            记录审计日志("自动备份", "backup", detail=f"type=database,filename={数据库文件.name}")
        if not 今日是否已有备份("uploads"):
            上传文件 = 构建上传目录备份(created_by=created_by)
            记录审计日志("自动备份", "backup", detail=f"type=uploads,filename={上传文件.name}")
        db.session.commit()
        return 数据库文件, 上传文件

    def 今日是否已有备份(备份类型):
        今日零点 = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        return BackupRecord.query.filter(
            BackupRecord.backup_type == 备份类型,
            BackupRecord.status == "成功",
            BackupRecord.is_deleted.is_(False),
            BackupRecord.created_at >= 今日零点,
        ).count() > 0

    def 查询最新备份时间(备份类型):
        记录 = BackupRecord.query.filter(
            BackupRecord.backup_type == 备份类型,
            BackupRecord.status == "成功",
            BackupRecord.is_deleted.is_(False),
        ).order_by(BackupRecord.created_at.desc()).first()
        return 记录.created_at if 记录 else None

    def 启动定时备份任务():
        if app.config.get("BACKUP_SCHEDULER_STARTED"):
            return
        app.config["BACKUP_SCHEDULER_STARTED"] = True

        def 任务循环():
            time.sleep(10)
            while True:
                try:
                    with app.app_context():
                        执行每日自动备份(created_by="system")
                except Exception as exc:
                    with app.app_context():
                        db.session.rollback()
                        db.session.add(BackupRecord(
                            backup_type="scheduled",
                            file_path="",
                            file_size=0,
                            file_hash="",
                            status="失败",
                            error_message=str(exc),
                            created_by="system",
                            updated_by="system",
                        ))
                        db.session.commit()
                time.sleep(24 * 60 * 60)

        threading.Thread(target=任务循环, name="daily-backup", daemon=True).start()

    def 构建产品资料缺失提示(product, enterprise=None, product_files=None, skus=None):
        extra = 兼容产品基础信息字段(product, product.product_extra_fields or {})
        product_files = product_files or []
        skus = skus or []
        资料缺失提示 = []

        if not (product.main_image and product.main_image.strip()):
            资料缺失提示.append("缺少产品主图")
        if not (product.product_name_cn and product.product_name_cn.strip()):
            资料缺失提示.append("缺少产品名称")
        if not (product.enterprise_id and enterprise):
            资料缺失提示.append("缺少所属企业")
        if not (product.product_category and product.product_category.strip()):
            资料缺失提示.append("缺少产品品类")
        if not (product.export_suitability and product.export_suitability.strip()):
            资料缺失提示.append("缺少出口适配判断")
        if not (product.recommendation_level and product.recommendation_level.strip()):
            资料缺失提示.append("缺少推荐等级")
        if not ((product.moq and str(product.moq).strip()) or (extra.get("trade_moq") and str(extra.get("trade_moq")).strip())):
            资料缺失提示.append("缺少MOQ")
        if not (
            (product.delivery_cycle and product.delivery_cycle.strip())
            or (product.production_cycle and product.production_cycle.strip())
            or (extra.get("trade_mass_cycle") and str(extra.get("trade_mass_cycle")).strip())
        ):
            资料缺失提示.append("缺少交期")

        认证情况 = (extra.get("cert_status") or product.certification_status or "").strip()
        if 认证情况 in {"待补充", "未核验"}:
            资料缺失提示.append("认证资料待补充")
        if not product_files:
            资料缺失提示.append("缺少产品附件")
        if not skus:
            资料缺失提示.append("缺少SKU明细")

        return 资料缺失提示

    def 构建产品附件统计与状态(product_files, product_extra_values=None):
        product_extra_values = product_extra_values or {}
        类型计数 = {}
        for doc in product_files or []:
            类型计数[doc.document_type] = 类型计数.get(doc.document_type, 0) + 1
        展示统计 = [{"label": label, "count": 类型计数.get(raw_type, 0), "raw_type": raw_type} for label, raw_type in PRODUCT_MATERIAL_TYPE_MAPPING]
        已覆盖类型 = {item["raw_type"] for item in 展示统计}
        for 文件类型, 数量 in sorted(类型计数.items()):
            if 文件类型 in 已覆盖类型:
                continue
            展示统计.append({"label": 文件类型, "count": 数量, "raw_type": 文件类型})
        状态建议 = {}
        for 字段, 文件类型 in PRODUCT_MATERIAL_STATUS_FIELD_MAPPING.items():
            当前值 = (product_extra_values.get(字段) or "").strip()
            数量 = 类型计数.get(文件类型, 0)
            if 数量 > 0:
                建议值 = "已上传"
            elif 当前值 == "待补充":
                建议值 = "待补充"
            else:
                建议值 = "未上传"
            状态建议[字段] = {"count": 数量, "recommended": 建议值}
        return 展示统计, 状态建议

    def 附加文件元信息(文档列表):
        for 文档 in 文档列表 or []:
            文件路径 = BASE_DIR / (文档.file_path or "")
            文档.file_size_bytes = 文件路径.stat().st_size if 文件路径.exists() and 文件路径.is_file() else None
        return 文档列表

    @app.template_filter("currency")
    def currency_filter(value, currency="USD"):
        if value is None or isinstance(value, Undefined):
            return "-"
        try:
            numeric_value = float(value)
        except (TypeError, ValueError):
            return "-"
        return f"{currency} {numeric_value:,.2f}"

    @app.context_processor
    def inject_user_context():
        return {
            "当前用户名": session.get("用户"),
            "当前角色": session.get("角色"),
            "是管理员": session.get("角色") == "管理员",
            "ENTRY_SOURCE_OPTIONS": ENTRY_SOURCE_OPTIONS,
            "ENTRY_STAGE_OPTIONS": ENTRY_STAGE_OPTIONS,
        }

    @app.before_request
    def enforce_login():
        if request.endpoint in {"登录", "static"}:
            return None
        if request.endpoint and request.endpoint.startswith("static"):
            return None
        if not session.get("用户"):
            return redirect(url_for("登录"))
        return None

    @app.errorhandler(413)
    def request_entity_too_large(_error):
        flash("上传文件过大，单文件限制为 100MB。", "danger")
        return redirect(url_for("document_upload"))

    def 行业下拉选项():
        return INDUSTRY_OPTIONS

    def 行业默认名称(行业代码):
        item = INDUSTRY_MAP.get((行业代码 or "").strip())
        if not item:
            return ""
        return item["name"]

    def 解析行业(行业代码, 手动行业名称=None):
        item = INDUSTRY_MAP.get((行业代码 or "").strip())
        if not item:
            return None, None
        行业名称 = (手动行业名称 or "").strip() or item["name"]
        return item["code"], 行业名称

    def 行业专项字段组(行业代码):
        return INDUSTRY_EXTRA_FIELD_CONFIG.get((行业代码 or "").strip(), [])

    def 提取字段配置(行业代码):
        配置 = []
        for 分组 in COMMON_ENTERPRISE_FIELD_GROUPS:
            配置.extend(分组.get("fields", []))
        for 分组 in 行业专项字段组(行业代码):
            配置.extend(分组.get("fields", []))
        return 配置

    def 读取多选值(form, key):
        return [item.strip() for item in form.getlist(key) if item and item.strip()]

    def 提取企业扩展字段(form, 行业代码):
        数据 = {}
        for 字段 in 提取字段配置(行业代码):
            键名 = 字段["key"]
            if 字段.get("type") in {"checkbox_group", "checkbox_tags"}:
                值 = 读取多选值(form, 键名)
            elif 字段.get("type") == "dynamic_contacts":
                姓名列表 = [item.strip() for item in form.getlist("dynamic_contacts_name[]")]
                部门列表 = [item.strip() for item in form.getlist("dynamic_contacts_department[]")]
                职位列表 = [item.strip() for item in form.getlist("dynamic_contacts_position[]")]
                手机列表 = [item.strip() for item in form.getlist("dynamic_contacts_mobile[]")]
                邮箱列表 = [item.strip() for item in form.getlist("dynamic_contacts_email[]")]
                微信列表 = [item.strip() for item in form.getlist("dynamic_contacts_wechat[]")]
                负责内容列表 = [item.strip() for item in form.getlist("dynamic_contacts_responsibility[]")]
                主联系人索引 = (form.get("dynamic_contacts_primary_index") or "").strip()
                值 = []
                最大长度 = max(
                    len(姓名列表),
                    len(部门列表),
                    len(职位列表),
                    len(手机列表),
                    len(邮箱列表),
                    len(微信列表),
                    len(负责内容列表),
                )
                for idx in range(最大长度):
                    行 = {
                        "name": 姓名列表[idx] if idx < len(姓名列表) else "",
                        "department": 部门列表[idx] if idx < len(部门列表) else "",
                        "position": 职位列表[idx] if idx < len(职位列表) else "",
                        "mobile": 手机列表[idx] if idx < len(手机列表) else "",
                        "email": 邮箱列表[idx] if idx < len(邮箱列表) else "",
                        "wechat": 微信列表[idx] if idx < len(微信列表) else "",
                        "responsibility": 负责内容列表[idx] if idx < len(负责内容列表) else "",
                        "is_primary_contact": str(idx) == 主联系人索引,
                    }
                    if any(行.get(k) for k in ("name", "department", "position", "mobile", "email", "wechat", "responsibility")):
                        值.append(行)
            else:
                原值 = form.get(键名, "")
                值 = 原值.strip() if isinstance(原值, str) else 原值
            if 值 in (None, "", []):
                continue
            数据[键名] = 值
        for 额外键 in ("source_channels", "enterprise_stage", "enterprise_tag_notes"):
            if 额外键 == "source_channels":
                额外值 = 读取多选值(form, 额外键)
            else:
                原值 = form.get(额外键, "")
                额外值 = 原值.strip() if isinstance(原值, str) else 原值
            if 额外值 in (None, "", []):
                continue
            数据[额外键] = 额外值
        return 清理企业扩展字段(数据)

    def 估算人员规模(employee_count):
        if not employee_count:
            return ""
        if employee_count <= 20:
            return "1-20人"
        if employee_count <= 50:
            return "21-50人"
        if employee_count <= 100:
            return "51-100人"
        if employee_count <= 300:
            return "101-300人"
        if employee_count <= 500:
            return "301-500人"
        if employee_count <= 1000:
            return "501-1000人"
        return "1000人以上"

    def 兼容企业基础信息字段(企业, 扩展字段):
        ext = 清理企业扩展字段(扩展字段 or {})
        if not ext.get("company_short_name"):
            ext["company_short_name"] = (ext.get("short_name") or "").strip()
        if not ext.get("company_full_name"):
            ext["company_full_name"] = (企业.company_name or "").strip()
        if not ext.get("registered_name"):
            ext["registered_name"] = (getattr(企业, "registered_name", None) or ext.get("company_full_name") or 企业.company_name or "").strip()
        if not ext.get("founded_date"):
            ext["founded_date"] = 企业.founded_date.strftime("%Y-%m-%d") if 企业.founded_date else ""
        if not ext.get("legal_representative"):
            ext["legal_representative"] = (
                getattr(企业, "legal_representative", None)
                or ext.get("legal_person")
                or ext.get("corporate_representative")
                or ext.get("法定代表人")
                or ""
            ).strip()
        if not ext.get("financing_stage"):
            ext["financing_stage"] = (ext.get("financing_round") or "").strip()
        if not ext.get("valuation"):
            ext["valuation"] = (ext.get("market_value") or ext.get("valuation_amount") or "").strip()
        if not ext.get("employee_count_range"):
            ext["employee_count_range"] = 估算人员规模(企业.employee_count) or (ext.get("employee_size") or "").strip()
        if not ext.get("core_products"):
            ext["core_products"] = (企业.main_products or ext.get("main_products") or "").strip()
        if not ext.get("website"):
            ext["website"] = (ext.get("official_website") or "").strip()
        if not ext.get("operating_status"):
            ext["operating_status"] = (ext.get("business_status") or "未知").strip()
        if not ext.get("company_type"):
            ext["company_type"] = (企业.company_type or "").strip()
        if not ext.get("business_registration_number"):
            ext["business_registration_number"] = (
                ext.get("industry_commerce_registration_number")
                or ext.get("business_license_registration_no")
                or ext.get("registration_number")
                or ""
            ).strip()
        if not ext.get("business_term_start") and getattr(企业, "business_term_start", None):
            ext["business_term_start"] = 企业.business_term_start.strftime("%Y-%m-%d")
        if not ext.get("business_term_end") and getattr(企业, "business_term_end", None):
            ext["business_term_end"] = 企业.business_term_end.strftime("%Y-%m-%d")
        if not ext.get("business_term"):
            起始日期 = (ext.get("business_term_start") or "").strip()
            终止日期 = (ext.get("business_term_end") or "").strip()
            ext["business_term"] = f"{起始日期}至{终止日期}" if 起始日期 and 终止日期 else 起始日期
        if not ext.get("business_term"):
            ext["business_term"] = (
                ext.get("business_validity_period")
                or ext.get("operating_term")
                or ext.get("营业期限")
                or ""
            ).strip()
        if ext.get("business_term") and not ext.get("business_term_start"):
            日期片段 = re.findall(r"\d{4}-\d{1,2}-\d{1,2}", ext.get("business_term") or "")
            if 日期片段:
                ext["business_term_start"] = 日期片段[0]
            if len(日期片段) > 1 and not ext.get("business_term_end"):
                ext["business_term_end"] = 日期片段[1]
        if not ext.get("paid_in_capital"):
            ext["paid_in_capital"] = (
                ext.get("actual_paid_capital")
                or ext.get("real_paid_capital")
                or ext.get("实缴资本")
                or ""
            ).strip()
        if not ext.get("approval_date"):
            ext["approval_date"] = (ext.get("approval_time") or ext.get("核准时间") or "").strip()
        if not ext.get("industry"):
            ext["industry"] = (ext.get("primary_industry") or 企业.industry_category or "").strip()
        if not ext.get("registration_authority"):
            ext["registration_authority"] = (ext.get("registration_organ") or ext.get("登记机关") or "").strip()
        if not ext.get("business_scope"):
            ext["business_scope"] = (
                getattr(企业, "business_scope", None)
                or ext.get("business_range")
                or ext.get("经营范围")
                or ""
            ).strip()
        if not ext.get("one_sentence_intro"):
            ext["one_sentence_intro"] = (ext.get("slogan") or ext.get("core_value") or "").strip()
        if not ext.get("enterprise_description"):
            ext["enterprise_description"] = (ext.get("company_profile") or ext.get("main_business") or 企业.main_business or "").strip()
        enterprise_tags = ext.get("enterprise_tags")
        if isinstance(enterprise_tags, str):
            enterprise_tags = [item.strip() for item in re.split(r"[、,，;/|]+", enterprise_tags) if item and item.strip()]
        if not enterprise_tags:
            legacy_tags = ext.get("tags") or ext.get("enterprise_tag_notes") or ext.get("industry_tags") or ""
            enterprise_tags = [item.strip() for item in re.split(r"[、,，;/|]+", legacy_tags) if item and item.strip()]
        if enterprise_tags:
            ext["enterprise_tags"] = enterprise_tags
        return ext

    def 企业核心完整度字段(企业, 扩展字段):
        ext = 兼容企业基础信息字段(企业, 扩展字段 or {})
        基本信息字段 = [
            ("企业全称", ext.get("company_full_name") or 企业.company_name),
            ("企业简称", ext.get("company_short_name")),
            ("行业分类", 行业显示名称(企业.industry_code, 企业.industry_category)),
            ("细分行业", 企业.sub_industry),
            ("核心产品", ext.get("core_products") or 企业.main_products),
            ("企业网址", ext.get("website")),
            ("运营状态", ext.get("operating_status")),
            ("公司类型", ext.get("company_type") or 企业.company_type),
            ("一句话简介", ext.get("one_sentence_intro")),
            ("企业介绍", ext.get("enterprise_description") or 企业.main_business),
            ("省份", 企业.province),
            ("城市", 企业.city),
        ]
        工商信息字段 = [
            ("注册名称", ext.get("registered_name")),
            ("法人代表", ext.get("legal_representative")),
            ("成立日期", ext.get("founded_date") or (企业.founded_date.strftime("%Y-%m-%d") if 企业.founded_date else "")),
            ("统一社会信用代码", ext.get("unified_social_credit_code") or 企业.unified_social_credit_code),
            ("注册资本", ext.get("registered_capital") or 企业.registered_capital),
            ("企业类型", ext.get("company_type") or 企业.company_type),
            ("注册地址", ext.get("registered_address") or 企业.registered_address),
            ("营业期限起始日期", ext.get("business_term_start")),
            ("营业期限终止日期", ext.get("business_term_end")),
        ]
        return {
            "A": [(字段名, 字段已填写(字段值)) for 字段名, 字段值 in 基本信息字段],
            "B": [(字段名, 字段已填写(字段值)) for 字段名, 字段值 in 工商信息字段],
        }

    ENTERPRISE_REQUIRED_FIELD_RULES = [
        ("基本信息-企业编号", lambda 企业, ext: 企业.enterprise_code),
        ("基本信息-行业大类", lambda 企业, ext: 企业.industry_code),
        ("基本信息-企业全称", lambda 企业, ext: ext.get("company_full_name") or 企业.company_name),
        ("工商信息-企业类型", lambda 企业, ext: ext.get("company_type") or 企业.company_type),
        ("工商信息-统一社会信用代码", lambda 企业, ext: ext.get("unified_social_credit_code") or 企业.unified_social_credit_code),
        ("工商信息-法定代表人", lambda 企业, ext: ext.get("legal_representative")),
        ("工商信息-注册资本", lambda 企业, ext: ext.get("registered_capital") or 企业.registered_capital),
        ("工商信息-成立日期", lambda 企业, ext: ext.get("founded_date") or (企业.founded_date.strftime("%Y-%m-%d") if 企业.founded_date else "")),
        ("工商信息-营业期限起始日期", lambda 企业, ext: ext.get("business_term_start")),
        ("工商信息-注册地址", lambda 企业, ext: ext.get("registered_address") or 企业.registered_address),
    ]

    def 企业必填缺失字段(企业, 扩展字段):
        ext = 兼容企业基础信息字段(企业, 扩展字段 or {})
        return [字段名 for 字段名, 取值 in ENTERPRISE_REQUIRED_FIELD_RULES if not 字段已填写(取值(企业, ext))]

    def 企业提交审核缺失字段(企业, 扩展字段):
        return 企业必填缺失字段(企业, 扩展字段)

    def 企业重复风险检查(company_name, unified_social_credit_code, exclude_enterprise_id=None):
        风险提示 = []
        企业全称 = (company_name or "").strip()
        统一社会信用代码 = (unified_social_credit_code or "").strip()

        if 企业全称:
            企业全称查询 = Enterprise.query.filter(Enterprise.company_name == 企业全称)
            if exclude_enterprise_id:
                企业全称查询 = 企业全称查询.filter(Enterprise.id != exclude_enterprise_id)
            if 企业全称查询.first():
                风险提示.append("企业可能已存在")

        if 统一社会信用代码:
            信用代码查询 = Enterprise.query.filter(Enterprise.unified_social_credit_code == 统一社会信用代码)
            if exclude_enterprise_id:
                信用代码查询 = 信用代码查询.filter(Enterprise.id != exclude_enterprise_id)
            if 信用代码查询.first():
                风险提示.append("该统一社会信用代码已存在")

        return 风险提示

    def 字段已填写(值):
        if 值 is None:
            return False
        if isinstance(值, str):
            return bool(值.strip())
        if isinstance(值, (list, tuple, set, dict)):
            return len(值) > 0
        return True

    def 提取本次上传文件类型():
        类型列表 = request.form.getlist("enterprise_upload_type")
        文件列表 = request.files.getlist("enterprise_upload_file")
        已上传类型 = set()
        for idx, 上传文件 in enumerate(文件列表):
            文件类型 = (类型列表[idx] if idx < len(类型列表) else "").strip()
            if 文件类型 and 上传文件 and 上传文件.filename:
                已上传类型.add(文件类型)
        return 已上传类型

    def 计算企业资料完整度(企业, 扩展字段, 本次上传类型=None):
        核心字段 = 企业核心完整度字段(企业, 扩展字段 or {})
        基本已填 = sum(1 for _, ok in 核心字段["A"] if ok)
        工商已填 = sum(1 for _, ok in 核心字段["B"] if ok)
        基本总数 = len(核心字段["A"])
        工商总数 = len(核心字段["B"])
        总字段数 = 基本总数 + 工商总数
        分数 = int(round(((基本已填 + 工商已填) / 总字段数) * 100)) if 总字段数 else 0

        if 分数 >= 80:
            颜色 = "success"
        elif 分数 >= 50:
            颜色 = "warning"
        else:
            颜色 = "danger"

        缺失项 = [字段名 for 字段名, ok in [*核心字段["A"], *核心字段["B"]] if not ok]
        return {
            "score": 分数,
            "label": f"{分数}%",
            "color": 颜色,
            "missing_items": 缺失项,
            "tabs": {
                "A": {"done": 基本已填, "total": 基本总数},
                "B": {"done": 工商已填, "total": 工商总数},
            },
        }

    def 产品行业专项字段组(行业代码):
        return INDUSTRY_PRODUCT_EXTRA_FIELD_CONFIG.get((行业代码 or "").strip(), [])

    def 产品扩展字段配置(行业代码):
        配置 = []
        for 分组 in COMMON_PRODUCT_FIELD_GROUPS:
            配置.extend(分组.get("fields", []))
        for 分组 in 产品行业专项字段组(行业代码):
            配置.extend(分组.get("fields", []))
        return 配置

    def 提取产品扩展字段(form, 行业代码):
        数据 = {}
        for 字段 in 产品扩展字段配置(行业代码):
            键名 = 字段["key"]
            if 字段.get("type") == "checkbox_group":
                值 = 读取多选值(form, 键名)
            else:
                原值 = form.get(键名, "")
                值 = 原值.strip() if isinstance(原值, str) else 原值
            if 值 in (None, "", []):
                continue
            数据[键名] = 值
        return 数据

    def 构建产品扩展信息分组(行业代码, 扩展字段):
        扩展字段 = 扩展字段 or {}
        分组数据 = []
        for 分组 in [*COMMON_PRODUCT_FIELD_GROUPS, *产品行业专项字段组(行业代码)]:
            字段列表 = []
            for 字段 in 分组.get("fields", []):
                值 = 扩展字段.get(字段["key"])
                if 值 in (None, "", []):
                    continue
                if isinstance(值, list):
                    值 = "、".join(值)
                字段列表.append({"label": 字段["label"], "value": 值})
            if 字段列表:
                分组数据.append({"title": 分组["title"], "fields": 字段列表})
        return 分组数据

    def 校验企业附件扩展名(document_type, extension):
        允许集合 = ENTERPRISE_TYPE_ALLOWED_EXTENSIONS.get((document_type or "").strip())
        if not 允许集合:
            return
        if extension.lower() not in 允许集合:
            可用列表 = " / ".join(sorted(允许集合))
            raise ValueError(f"{document_type} 仅支持：{可用列表}")

    def 生成企业附件文件名(enterprise, document_type, extension, 自定义名称, 上传序号):
        日期文本 = datetime.now().strftime("%Y%m%d")
        企业编号 = 清洗路径片段(enterprise.enterprise_code or "E000")
        企业简称 = 清洗路径片段((enterprise.company_name or "").strip() or "未命名企业")
        说明文本 = 清洗路径片段(自定义名称 or "")
        if document_type == "营业执照":
            return f"{企业编号}_{企业简称}_营业执照_{日期文本}{extension}"
        if document_type == "企业介绍":
            资料说明 = 说明文本 or "企业介绍"
            return f"{企业编号}_{企业简称}_{资料说明}_{日期文本}{extension}"
        if document_type == "BP/融资材料":
            资料说明 = 说明文本 or "BP融资材料"
            return f"{企业编号}_{企业简称}_{资料说明}_{日期文本}{extension}"
        if document_type == "资质荣誉":
            证书名称 = 说明文本 or "资质荣誉"
            return f"{企业编号}_{企业简称}_{证书名称}_{日期文本}{extension}"
        if document_type == "工厂照片":
            return f"{企业编号}_{企业简称}_工厂照片_{上传序号:02d}_{日期文本}{extension}"
        if document_type == "财务/经营资料":
            资料说明 = 说明文本 or "财务经营资料"
            return f"{企业编号}_{企业简称}_{资料说明}_{日期文本}{extension}"
        if document_type == "其他":
            资料说明 = 说明文本 or "其他"
            return f"{企业编号}_{企业简称}_{资料说明}_{日期文本}{extension}"
        return None

    def 生成产品附件文件名(enterprise, product, document_type, extension, 自定义名称):
        日期文本 = datetime.now().strftime("%Y%m%d")
        企业编号 = 清洗路径片段(enterprise.enterprise_code or "E000")
        产品编号 = 清洗路径片段(product.product_code or "P000")
        文件分类 = 清洗路径片段(document_type or "其他")
        文件说明 = 清洗路径片段(自定义名称 or "未命名文件")
        return f"{企业编号}_{产品编号}_{文件分类}_{文件说明}_{日期文本}{extension}"

    def 处理表单文件上传(enterprise, 类型字段名, 名称字段名, 文件字段名, product=None, use_enterprise_naming=False, use_product_naming=False):
        类型列表 = request.form.getlist(类型字段名)
        名称列表 = request.form.getlist(名称字段名)
        文件列表 = request.files.getlist(文件字段名)
        当前上传人 = (session.get("用户") or "未署名").strip() or "未署名"
        上传成功数 = 0

        for 索引, 上传文件 in enumerate(文件列表):
            if not 上传文件 or not 上传文件.filename:
                continue
            文件类型 = (类型列表[索引] if 索引 < len(类型列表) else "").strip()
            自定义名称 = (名称列表[索引] if 索引 < len(名称列表) else "").strip()
            if not 文件类型:
                raise ValueError("存在未选择文件类型的上传项。")
            if use_enterprise_naming and 文件类型 not in ENTERPRISE_UPLOAD_TYPES:
                raise ValueError(f"企业附件分类“{文件类型}”不在允许范围内。")
            if use_product_naming and 文件类型 not in PRODUCT_UPLOAD_TYPES:
                raise ValueError(f"产品附件分类“{文件类型}”不在允许范围内。")
            原始名称 = Path(上传文件.filename).stem
            文件名称 = 清洗路径片段(自定义名称 or 原始名称) or "未命名文件"
            扩展名 = Path(上传文件.filename).suffix.lower()
            if use_enterprise_naming:
                校验企业附件扩展名(文件类型, 扩展名)
            目标文件名 = 生成企业附件文件名(enterprise, 文件类型, 扩展名, 文件名称, 上传成功数 + 1) if use_enterprise_naming else None
            if use_product_naming:
                if not product:
                    raise ValueError("产品附件上传失败：缺少产品信息。")
                目标文件名 = 生成产品附件文件名(enterprise, product, 文件类型, 扩展名, 文件名称)
            保存文件并登记记录(
                app=app,
                enterprise=enterprise,
                product=product,
                上传文件=上传文件,
                document_type=文件类型,
                document_name=文件名称,
                uploaded_by=当前上传人,
                filename_override=目标文件名,
            )
            上传成功数 += 1
        return 上传成功数

    @app.route("/")
    def dashboard():
        企业数量 = Enterprise.query.count()
        产品数量 = Product.query.count()
        已上传文件数量 = Document.query.count()
        最近新增企业 = Enterprise.query.order_by(Enterprise.created_at.desc()).limit(5).all()
        最近新增产品 = Product.query.order_by(Product.created_at.desc()).limit(5).all()

        return render_template(
            "dashboard.html",
            企业数量=企业数量,
            产品数量=产品数量,
            已上传文件数量=已上传文件数量,
            最近新增企业=最近新增企业,
            最近新增产品=最近新增产品,
        )

    @app.route("/companies")
    def enterprise_library():
        return redirect(url_for("enterprise_list"))

    @app.route("/product-library")
    def product_library():
        return redirect(url_for("product_list"))

    @app.route("/backups")
    def backup_management():
        return redirect(url_for("backup_tools"))

    @app.route("/backup", methods=["GET", "POST"])
    @admin_required
    def backup_tools():
        app.config["DATABASE_BACKUP_ROOT"].mkdir(parents=True, exist_ok=True)
        app.config["FILES_BACKUP_ROOT"].mkdir(parents=True, exist_ok=True)
        if request.method == "POST":
            action = request.form.get("action", "")
            if action == "backup_db":
                文件路径 = 构建数据库备份()
                记录审计日志("备份", "backup", detail=f"type=database,filename={文件路径.name}")
                db.session.commit()
                flash(f"数据库备份成功：{文件路径.name}", "success")
            elif action == "backup_uploads":
                文件路径 = 构建上传目录备份()
                记录审计日志("备份", "backup", detail=f"type=uploads,filename={文件路径.name}")
                db.session.commit()
                flash(f"企业文件资料备份成功：{文件路径.name}", "success")
            elif action == "backup_all":
                数据库文件 = 构建数据库备份()
                文件资料文件 = 构建上传目录备份()
                记录审计日志("备份", "backup", detail=f"type=database,filename={数据库文件.name}")
                记录审计日志("备份", "backup", detail=f"type=uploads,filename={文件资料文件.name}")
                db.session.commit()
                flash(f"完整备份成功：{数据库文件.name}，{文件资料文件.name}", "success")
            else:
                flash("未知备份动作。", "danger")
            return redirect(url_for("backup_tools"))

        备份记录列表 = BackupRecord.query.filter(BackupRecord.is_deleted.is_(False)).order_by(BackupRecord.created_at.desc()).limit(50).all()
        文件列表 = [
            {
                "name": Path(记录.file_path).name,
                "type": 记录.backup_type,
                "mtime": 记录.created_at,
                "size": 记录.file_size,
                "hash": 记录.file_hash,
                "status": 记录.status,
            }
            for 记录 in 备份记录列表
        ]
        操作日志 = (
            AuditLog.query.filter_by(target_type="backup").order_by(AuditLog.created_at.desc()).limit(30).all()
        )
        导出日志 = ExportLog.query.order_by(ExportLog.created_at.desc()).limit(30).all()
        return render_template(
            "backup.html",
            files=文件列表,
            db_path=BASE_DIR / "trade_agent.db",
            upload_path=app.config["UPLOAD_ENTERPRISE_ROOT"],
            latest_db_backup_time=查询最新备份时间("database"),
            latest_files_backup_time=查询最新备份时间("uploads"),
            logs=操作日志,
            export_logs=导出日志,
        )

    @app.get("/backup/keyword-suggestions")
    @admin_required
    def backup_keyword_suggestions():
        keyword = request.args.get("q", "", type=str).strip()
        export_type = request.args.get("export_type", "", type=str).strip()
        if not keyword:
            return jsonify({"results": []})

        like = f"%{keyword}%"
        results = []
        include_enterprises = export_type in {"enterprise_table", "enterprise_package", ""}
        include_products = export_type in {"product_table", "product_card", ""}

        if include_enterprises:
            enterprises = (
                Enterprise.query.filter(
                    or_(
                        Enterprise.company_name.ilike(like),
                        Enterprise.english_name.ilike(like),
                        Enterprise.enterprise_code.ilike(like),
                    )
                )
                .order_by(Enterprise.updated_at.desc())
                .limit(8)
                .all()
            )
            for enterprise in enterprises:
                subtitle_parts = [enterprise.enterprise_code, enterprise.english_name, (enterprise.industry_category or 行业默认名称(enterprise.industry_code))]
                results.append(
                    {
                        "type": "enterprise",
                        "type_label": "企业",
                        "title": enterprise.company_name or enterprise.enterprise_code,
                        "subtitle": " · ".join(part for part in subtitle_parts if part),
                        "value": enterprise.company_name or enterprise.enterprise_code or enterprise.english_name,
                    }
                )

        if include_products:
            products = (
                Product.query.join(Enterprise, Product.enterprise_id == Enterprise.id)
                .filter(
                    or_(
                        Product.product_name_cn.ilike(like),
                        Product.product_name_en.ilike(like),
                        Product.product_code.ilike(like),
                        Product.model.ilike(like),
                        Enterprise.company_name.ilike(like),
                        Enterprise.english_name.ilike(like),
                        Enterprise.enterprise_code.ilike(like),
                    )
                )
                .order_by(Product.updated_at.desc())
                .limit(8)
                .all()
            )
            for product in products:
                enterprise_name = product.enterprise.company_name if product.enterprise else ""
                subtitle_parts = [product.product_code, product.product_name_en, product.model, enterprise_name]
                results.append(
                    {
                        "type": "product",
                        "type_label": "产品",
                        "title": product.product_name_cn or product.product_name_en or product.product_code,
                        "subtitle": " · ".join(part for part in subtitle_parts if part),
                        "value": product.product_name_cn or product.product_name_en or product.product_code or product.model,
                    }
                )

        return jsonify({"results": results[:10]})

    @app.get("/backup/download/<string:backup_type>/<path:filename>")
    @admin_required
    def download_backup(backup_type, filename):
        if backup_type == "database" and filename.startswith("database_backup_") and filename.endswith(".sqlite"):
            目录 = app.config["DATABASE_BACKUP_ROOT"]
        elif backup_type in {"files", "uploads"} and filename.startswith("uploads_backup_") and filename.endswith(".zip"):
            目录 = app.config["FILES_BACKUP_ROOT"]
        else:
            flash("非法备份文件请求。", "danger")
            return redirect(url_for("backup_tools"))
        return send_from_directory(目录, filename, as_attachment=True)

    @app.get("/excel/export/<string:export_key>")
    def excel_export(export_key):
        """兼容旧导出入口，统一转到新的可审计导出接口。"""
        if export_key == "enterprises":
            return redirect(url_for("export_enterprises_all"))
        if export_key == "products":
            return redirect(url_for("export_products_all"))
        flash("不支持的导出类型。", "danger")
        return redirect(url_for("dashboard"))

    def 记录导出日志(export_type, export_scope, filters=None, record_count=0, file_name=None, file_path=None, status="成功", error_message=None, related_enterprise_id=None, related_product_id=None):
        用户 = 当前用户()
        db.session.add(
            ExportLog(
                user_id=用户.id if 用户 else None,
                username=用户.username if 用户 else (session.get("用户") or "system"),
                export_type=export_type,
                export_scope=export_scope,
                related_enterprise_id=related_enterprise_id,
                related_product_id=related_product_id,
                filters_json=json.dumps(filters or {}, ensure_ascii=False),
                record_count=record_count or 0,
                file_name=file_name,
                file_path=file_path,
                ip_address=request.headers.get("X-Forwarded-For", request.remote_addr or "").split(",")[0].strip(),
                status=status,
                error_message=error_message,
            )
        )

    def 发送导出Excel(export_type, export_scope, sheets, filters=None, related_enterprise_id=None, related_product_id=None):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"{export_type}_{export_scope}_{timestamp}.xlsx"
        record_count = sum(len(rows) for _, _, rows in sheets)
        try:
            buffer = 构建Excel文件(sheets)
            记录导出日志(export_type, export_scope, filters=filters, record_count=record_count, file_name=file_name, related_enterprise_id=related_enterprise_id, related_product_id=related_product_id)
            db.session.commit()
            if record_count == 0:
                flash("当前筛选条件下无数据，已导出空表。", "warning")
            return send_file(buffer, as_attachment=True, download_name=file_name, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        except Exception as exc:
            db.session.rollback()
            记录导出日志(export_type, export_scope, filters=filters, record_count=0, file_name=file_name, status="失败", error_message=str(exc), related_enterprise_id=related_enterprise_id, related_product_id=related_product_id)
            db.session.commit()
            flash(f"导出失败：{str(exc)}", "danger")
            return redirect(request.referrer or url_for("dashboard"))

    def 导出文件安全名称(*parts, default="导出文件"):
        raw_name = "_".join(str(part or "").strip() for part in parts if str(part or "").strip()) or default
        safe_name = re.sub(r'[\\/:*?"<>|]+', "_", raw_name)
        return safe_name[:120] or default

    def 企业资料包Sheets(enterprise):
        enterprise_rows = 构建企业总表Sheet(Enterprise.query.filter_by(id=enterprise.id), [key for key, _ in ENTERPRISE_EXPORT_COLUMNS])
        product_rows = 构建产品总表Sheet(Product.query.join(Enterprise, Product.enterprise_id == Enterprise.id).filter(Product.enterprise_id == enterprise.id), [key for key, _ in PRODUCT_EXPORT_COLUMNS])
        sku_rows = []
        for product in Product.query.filter_by(enterprise_id=enterprise.id).all():
            sku_rows.extend(构建SKU明细Sheet(product.id))
        attachment_rows = 构建附件清单Sheet(enterprise_id=enterprise.id, include_enterprise_products=True)
        return [
            ("企业基础信息", [label for _, label in ENTERPRISE_EXPORT_COLUMNS], enterprise_rows),
            ("产品清单", [label for _, label in PRODUCT_EXPORT_COLUMNS], product_rows),
            ("SKU明细", SKU_EXPORT_HEADERS, sku_rows),
            ("附件清单", ATTACHMENT_EXPORT_HEADERS, attachment_rows),
        ]

    def 产品资料卡Sheets(product):
        product_rows = 构建产品总表Sheet(Product.query.join(Enterprise, Product.enterprise_id == Enterprise.id).filter(Product.id == product.id), [key for key, _ in PRODUCT_EXPORT_COLUMNS])
        enterprise_rows = 构建企业总表Sheet(Enterprise.query.filter_by(id=product.enterprise_id), [key for key, _ in ENTERPRISE_EXPORT_COLUMNS])
        sku_rows = 构建SKU明细Sheet(product.id)
        attachment_rows = 构建附件清单Sheet(product_id=product.id)
        return [
            ("产品基础信息", [label for _, label in PRODUCT_EXPORT_COLUMNS], product_rows),
            ("所属企业信息", [label for _, label in ENTERPRISE_EXPORT_COLUMNS], enterprise_rows),
            ("SKU明细", SKU_EXPORT_HEADERS, sku_rows),
            ("附件清单", ATTACHMENT_EXPORT_HEADERS, attachment_rows),
        ]

    def 发送资料包Zip(export_type, export_scope, items, sheet_builder, filename_builder, filters=None):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"{export_type}_{export_scope}_{timestamp}.zip"
        try:
            buffer = BytesIO()
            with ZipFile(buffer, "w", ZIP_DEFLATED) as zip_file:
                for item in items:
                    excel_buffer = 构建Excel文件(sheet_builder(item))
                    zip_file.writestr(filename_builder(item), excel_buffer.getvalue())
            buffer.seek(0)
            记录导出日志(export_type, export_scope, filters=filters, record_count=len(items), file_name=file_name)
            db.session.commit()
            if not items:
                flash("当前筛选条件下无数据，已导出空压缩包。", "warning")
            return send_file(buffer, as_attachment=True, download_name=file_name, mimetype="application/zip")
        except Exception as exc:
            db.session.rollback()
            记录导出日志(export_type, export_scope, filters=filters, record_count=0, file_name=file_name, status="失败", error_message=str(exc))
            db.session.commit()
            flash(f"导出失败：{str(exc)}", "danger")
            return redirect(request.referrer or url_for("backup_tools"))

    @app.get("/export/enterprises/all")
    def export_enterprises_all():
        columns = [key for key, _ in ENTERPRISE_EXPORT_COLUMNS]
        headers = [label for _, label in ENTERPRISE_EXPORT_COLUMNS]
        rows = 构建企业总表Sheet(Enterprise.query, columns)
        return 发送导出Excel("企业总表", "全量导出", [("企业基础信息", headers, rows)], filters={})

    @app.get("/export/enterprises/backup-selected")
    def export_enterprises_backup_selected():
        columns = [key for key, _ in ENTERPRISE_EXPORT_COLUMNS]
        headers = [label for _, label in ENTERPRISE_EXPORT_COLUMNS]
        query = 构建企业查询(request.args)
        filters = {key: value for key, value in request.args.items() if value}
        rows = 构建企业总表Sheet(query, columns)
        scope = "文本筛选导出" if filters.get("keyword") else "全量导出"
        return 发送导出Excel("企业总表", scope, [("企业基础信息", headers, rows)], filters=filters)

    @app.get("/export/enterprises/filtered")
    def export_enterprises_filtered():
        columns, headers = 解析导出列(request.args, ENTERPRISE_EXPORT_COLUMNS)
        query = 构建企业查询(request.args)
        filters = {key: value for key, value in request.args.items() if key not in {"columns", "page"}}
        rows = 构建企业总表Sheet(query, columns)
        return 发送导出Excel("企业总表", "筛选导出", [("企业基础信息", headers, rows)], filters=filters)

    @app.get("/export/products/all")
    def export_products_all():
        columns = [key for key, _ in PRODUCT_EXPORT_COLUMNS]
        headers = [label for _, label in PRODUCT_EXPORT_COLUMNS]
        rows = 构建产品总表Sheet(Product.query.join(Enterprise, Product.enterprise_id == Enterprise.id), columns)
        return 发送导出Excel("产品总表", "全量导出", [("产品清单", headers, rows)], filters={})

    @app.get("/export/products/backup-selected")
    def export_products_backup_selected():
        columns = [key for key, _ in PRODUCT_EXPORT_COLUMNS]
        headers = [label for _, label in PRODUCT_EXPORT_COLUMNS]
        query = 构建产品查询(request.args, default_status="")
        filters = {key: value for key, value in request.args.items() if value}
        rows = 构建产品总表Sheet(query, columns)
        scope = "文本筛选导出" if filters.get("keyword") or filters.get("enterprise") else "全量导出"
        return 发送导出Excel("产品总表", scope, [("产品清单", headers, rows)], filters=filters)

    @app.get("/export/products/filtered")
    def export_products_filtered():
        columns, headers = 解析导出列(request.args, PRODUCT_EXPORT_COLUMNS)
        query = 构建产品查询(request.args)
        filters = {key: value for key, value in request.args.items() if key not in {"columns", "page"}}
        rows = 构建产品总表Sheet(query, columns)
        return 发送导出Excel("产品总表", "筛选导出", [("产品清单", headers, rows)], filters=filters)

    @app.get("/export/skus/all")
    def export_skus_all():
        rows = 构建SKU明细Sheet()
        return 发送导出Excel("SKU 明细", "全量导出", [("SKU明细", SKU_EXPORT_HEADERS, rows)], filters={})

    @app.get("/export/attachments/all")
    def export_attachments_all():
        rows = 构建附件清单Sheet()
        return 发送导出Excel("附件清单", "全量导出", [("附件清单", ATTACHMENT_EXPORT_HEADERS, rows)], filters={})

    @app.get("/export/import-errors/<int:batch_id>")
    def export_import_errors(batch_id):
        batch, rows = 构建导入错误报告Sheet(batch_id)
        if not batch:
            abort(404)
        return 发送导出Excel("导入错误报告", "单批次", [("导入错误报告", IMPORT_ERROR_HEADERS, rows)], filters={"batch_id": batch_id})

    @app.get("/export/import-supplements/<int:batch_id>")
    def export_import_supplements(batch_id):
        batch, rows = 构建待补充资料清单Sheet(batch_id)
        if not batch:
            abort(404)
        return 发送导出Excel("待补充资料清单", "单批次", [("待补充资料清单", SUPPLEMENT_LIST_HEADERS, rows)], filters={"batch_id": batch_id})

    @app.get("/export/enterprise-packages/selected")
    def export_enterprise_packages_selected():
        query = 构建企业查询(request.args)
        enterprises = query.order_by(Enterprise.updated_at.desc()).all()
        filters = {key: value for key, value in request.args.items() if value}
        scope = "文本筛选导出" if filters.get("keyword") else "全量导出"

        def filename_builder(enterprise):
            return f"{导出文件安全名称(enterprise.enterprise_code, enterprise.company_name, default='企业资料包')}_企业资料包.xlsx"

        return 发送资料包Zip("企业资料包", scope, enterprises, 企业资料包Sheets, filename_builder, filters=filters)

    @app.get("/export/product-cards/selected")
    def export_product_cards_selected():
        query = 构建产品查询(request.args, default_status="")
        products = query.order_by(Product.updated_at.desc()).all()
        filters = {key: value for key, value in request.args.items() if value}
        scope = "文本筛选导出" if filters.get("keyword") or filters.get("enterprise") else "全量导出"

        def filename_builder(product):
            enterprise_name = product.enterprise.company_name if product.enterprise else ""
            return f"{导出文件安全名称(product.product_code, product.product_name_cn, enterprise_name, default='产品资料卡')}_产品资料卡.xlsx"

        return 发送资料包Zip("产品资料卡", scope, products, 产品资料卡Sheets, filename_builder, filters=filters)

    @app.get("/export/enterprise/<int:enterprise_id>/package")
    def export_enterprise_package(enterprise_id):
        enterprise = Enterprise.query.get_or_404(enterprise_id)
        return 发送导出Excel("单企业资料包", "单企业", 企业资料包Sheets(enterprise), filters={"enterprise_id": enterprise.id}, related_enterprise_id=enterprise.id)

    @app.get("/export/product/<int:product_id>/card")
    def export_product_card(product_id):
        product = Product.query.get_or_404(product_id)
        return 发送导出Excel("单产品资料卡", "单产品", 产品资料卡Sheets(product), filters={"product_id": product.id}, related_enterprise_id=product.enterprise_id, related_product_id=product.id)


    def 企业导入模板数据():
        表头 = [字段 for 字段, _ in 企业导入字段提示()]
        样例映射 = {
            "企业编号": "ENT-20260430-001",
            "行业大类": "电子电器与智能硬件",
            "企业全称": "示例科技股份有限公司",
            "企业类型": "有限责任公司",
            "统一社会信用代码": "91310000MA1EXAMPLE",
            "法定代表人": "张三",
            "注册资本": "1000万元人民币",
            "成立日期": "2020-01-15",
            "营业期限": "2020-01-15",
            "注册地址": "广东省深圳市南山区示例路 1 号",
            "联系人": "李四",
            "联系电话": "13800138000",
        }
        样例 = [样例映射.get(字段, "") for 字段 in 表头]
        return 表头, 样例

    def 产品导入模板数据():
        表头 = [字段 for 字段, _ in 产品导入字段提示()]
        样例映射 = {
            "产品编号": "PRD-20260430-001",
            "产品名称": "智能温湿度传感器",
            "所属企业名称": "示例科技股份有限公司",
            "产品品类": "工业传感器",
            "产品类型": "温湿度传感器",
            "产能-周期（天）": "30",
            "产能-实际完工合格件数（件）": "10000",
            "MOQ": "100",
            "交期": "15天",
            "均价": "USD 18-22",
        }
        样例 = [样例映射.get(字段, "") for 字段 in 表头]
        return 表头, 样例

    def 构建导入模板文件(表头, 样例, sheet_name, 文件名前缀):
        workbook = Workbook()
        工作表 = workbook.active
        工作表.title = sheet_name
        工作表.append(表头)
        工作表.append(样例)
        缓冲区 = BytesIO()
        workbook.save(缓冲区)
        缓冲区.seek(0)
        return send_file(
            缓冲区,
            as_attachment=True,
            download_name=f"{文件名前缀}_导入模板.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.get("/excel/import/enterprises/template")
    def download_enterprise_import_template():
        表头, 样例 = 企业导入模板数据()
        return 构建导入模板文件(表头, 样例, "企业导入模板", "企业")

    @app.get("/excel/import/products/template")
    def download_product_import_template():
        表头, 样例 = 产品导入模板数据()
        return 构建导入模板文件(表头, 样例, "产品导入模板", "产品")

    @app.route("/excel/import/enterprises", methods=["GET", "POST"])
    def import_enterprises():
        返回地址 = request.args.get("next", url_for("enterprise_list"), type=str)
        if not 返回地址.startswith("/"):
            返回地址 = url_for("enterprise_list")
        if request.method == "POST":
            上传文件 = request.files.get("file")
            if not 上传文件 or not 上传文件.filename:
                flash("请先选择企业 Excel 文件。", "danger")
                return redirect(url_for("import_enterprises", next=返回地址))
            批次 = 创建导入批次(上传文件, "enterprise")
            try:
                成功条数, 失败列表 = 导入企业Excel(上传文件, 批次)
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
                return redirect(url_for("import_enterprises", next=返回地址))
            记录审计日志("导入 Excel", "enterprise", detail=f"batch={批次.id}, success={成功条数}, failed={len(失败列表)}")
            db.session.commit()
            return render_template("import_result.html", 标题="企业导入结果", 成功条数=成功条数, 失败列表=失败列表, 返回地址=返回地址, 返回文案="返回企业库", 导入批次=批次)
        return render_template(
            "import_form.html",
            标题="企业 Excel 导入",
            提示="支持按企业编号更新或新增企业。",
            返回地址=返回地址,
            返回文案="返回企业库",
            字段提示=企业导入字段提示(),
            必填字段=["企业编号", "企业全称"],
            模板下载地址=url_for("download_enterprise_import_template"),
            模板下载文案="下载企业导入模板",
        )

    @app.route("/excel/import/products", methods=["GET", "POST"])
    def import_products():
        返回地址 = request.args.get("next", url_for("product_list"), type=str)
        if not 返回地址.startswith("/"):
            返回地址 = url_for("product_list")
        if request.method == "POST":
            上传文件 = request.files.get("file")
            if not 上传文件 or not 上传文件.filename:
                flash("请先选择产品 Excel 文件。", "danger")
                return redirect(url_for("import_products", next=返回地址))
            批次 = 创建导入批次(上传文件, "product")
            try:
                成功条数, 失败列表 = 导入产品Excel(上传文件, 批次)
            except ValueError as exc:
                db.session.rollback()
                flash(str(exc), "danger")
                return redirect(url_for("import_products", next=返回地址))
            记录审计日志("导入 Excel", "product", detail=f"batch={批次.id}, success={成功条数}, failed={len(失败列表)}")
            db.session.commit()
            return render_template("import_result.html", 标题="产品导入结果", 成功条数=成功条数, 失败列表=失败列表, 返回地址=返回地址, 返回文案="返回产品库", 导入批次=批次)
        return render_template(
            "import_form.html",
            标题="产品 Excel 导入",
            提示="支持按产品编号更新或新增产品；上传时产品编号和所属企业名称为必填项，其余字段可为空。",
            返回地址=返回地址,
            返回文案="返回产品库",
            字段提示=产品导入字段提示(),
            必填字段=["产品编号", "所属企业名称"],
            模板下载地址=url_for("download_product_import_template"),
            模板下载文案="下载产品导入模板",
        )

    @app.route("/enterprises")
    def enterprise_list():
        page = request.args.get("page", 1, type=int)
        keyword = request.args.get("keyword", "", type=str).strip()
        industry = request.args.get("industry", "", type=str).strip()

        查询 = Enterprise.query
        if keyword:
            查询 = 查询.filter(
                or_(
                    Enterprise.company_name.ilike(f"%{keyword}%"),
                    func.json_extract(Enterprise.enterprise_extra_fields, "$.company_full_name").ilike(f"%{keyword}%"),
                )
            )
        if industry:
            查询 = 查询.filter(Enterprise.industry_code == industry)

        分页 = 查询.order_by(Enterprise.updated_at.desc()).paginate(page=page, per_page=PER_PAGE, error_out=False)

        def 提取JSON枚举值(json_key):
            path = f"$.{json_key}"
            rows = db.session.query(func.json_extract(Enterprise.enterprise_extra_fields, path)).filter(
                func.json_extract(Enterprise.enterprise_extra_fields, path).isnot(None),
                func.json_extract(Enterprise.enterprise_extra_fields, path) != "",
            ).distinct().all()
            clean_values = []
            for (raw,) in rows:
                if raw is None:
                    continue
                value = str(raw).strip()
                if value.startswith('"') and value.endswith('"'):
                    value = value[1:-1]
                if value and value not in clean_values:
                    clean_values.append(value)
            return sorted(clean_values)
        行业列表 = 行业下拉选项()

        企业ID列表 = [企业.id for 企业 in 分页.items]
        外贸负责人映射 = {}
        if 企业ID列表:
            联系人列表 = Contact.query.filter(Contact.enterprise_id.in_(企业ID列表), Contact.contact_type == "外贸负责人").all()
            for 联系人 in 联系人列表:
                if 联系人.enterprise_id not in 外贸负责人映射:
                    外贸负责人映射[联系人.enterprise_id] = 联系人.name
        完整度映射 = {}
        if 企业ID列表:
            文件映射 = {}
            for 文件 in Document.query.filter(Document.enterprise_id.in_(企业ID列表)).all():
                文件映射.setdefault(文件.enterprise_id, set()).add(文件.document_type)
            for 企业 in 分页.items:
                完整度映射[企业.id] = 计算企业资料完整度(
                    企业,
                    兼容企业基础信息字段(企业, 企业.enterprise_extra_fields or {}),
                    文件映射.get(企业.id, set()),
                )
                企业.industry_display_name = (企业.industry_category or "").strip() or 行业默认名称(企业.industry_code)

        return render_template(
            "enterprise_list.html",
            分页=分页,
            筛选={
                "keyword": keyword,
                "industry": industry,
            },
            行业列表=行业列表,
            外贸负责人映射=外贸负责人映射,
            完整度映射=完整度映射,
            enterprise_export_columns=ENTERPRISE_EXPORT_COLUMNS,
        )

    @app.route("/enterprises/new", methods=["GET", "POST"])
    def enterprise_new():
        if request.method == "POST":
            操作动作 = (request.form.get("action") or "save").strip()
            重复确认通过 = (request.form.get("duplicate_confirmed") or "").strip() == "1"
            行业代码, 行业名称 = 解析行业(request.form.get("industry_code"), request.form.get("industry_category"))
            扩展字段 = 提取企业扩展字段(request.form, 行业代码)
            扩展字段 = 兼容企业基础信息字段(Enterprise(), 扩展字段)
            企业性质列表 = 扩展字段.get("enterprise_natures", [])
            联系人子表单 = 扩展字段.get("dynamic_contacts", [])
            主联系人数 = sum(1 for item in 联系人子表单 if item.get("is_primary_contact"))
            出口经验 = 扩展字段.get("export_experience")
            企业编号 = (request.form.get("enterprise_code") or 扩展字段.get("enterprise_code") or "").strip()
            企业名称 = (扩展字段.get("company_full_name") or "").strip() or f"未命名企业-{企业编号 or '待编号'}"
            企业 = Enterprise(
                enterprise_code=企业编号,
                company_name=企业名称,
                english_name=扩展字段.get("english_name") or None,
                unified_social_credit_code=扩展字段.get("unified_social_credit_code") or None,
                registered_name=扩展字段.get("registered_name") or 企业名称,
                legal_representative=扩展字段.get("legal_representative") or None,
                founded_date=读取日期(扩展字段.get("founded_date")),
                registered_capital=扩展字段.get("registered_capital") or None,
                business_term_start=读取日期(扩展字段.get("business_term_start")),
                business_term_end=读取日期(扩展字段.get("business_term_end")),
                registered_address=扩展字段.get("registered_address") or None,
                business_scope=扩展字段.get("business_scope") or None,
                business_address=扩展字段.get("business_address") or None,
                province=request.form.get("province", "").strip() or None,
                city=request.form.get("city", "").strip() or None,
                district=request.form.get("district", "").strip() or None,
                company_type=(扩展字段.get("company_type") or "").strip() or None,
                industry_code=行业代码,
                industry_category=行业名称,
                sub_industry=request.form.get("sub_industry", "").strip() or None,
                main_products=扩展字段.get("core_products") or None,
                main_business=扩展字段.get("enterprise_description") or None,
                is_manufacturer="制造商" in 企业性质列表,
                is_trader="贸易商" in 企业性质列表,
                is_brand_owner="品牌商" in 企业性质列表,
                is_oem_odm="OEM/ODM工厂" in 企业性质列表,
                is_service_provider="服务商" in 企业性质列表,
                is_high_tech=读取布尔(request.form, "is_high_tech"),
                is_specialized_new=读取布尔(request.form, "is_specialized_new"),
                is_listed_or_pre_ipo=读取布尔(request.form, "is_listed_or_pre_ipo"),
                has_foreign_trade_experience=出口经验 == "是",
                export_countries=扩展字段.get("export_countries") or None,
                target_markets=扩展字段.get("target_countries") or None,
                annual_capacity=扩展字段.get("annual_capacity_level") or None,
                employee_count=读取整数(request.form.get("employee_count")) or None,
                factory_area=扩展字段.get("factory_area_range") or None,
                main_equipment=request.form.get("main_equipment", "").strip() or None,
                annual_revenue=读取金额(扩展字段.get("annual_sales")),
                export_revenue=读取金额(扩展字段.get("annual_exports")),
                service_needs=request.form.get("service_needs", "").strip() or None,
                risk_notes=request.form.get("risk_notes", "").strip() or None,
                enterprise_extra_fields=扩展字段,
                status=ENTERPRISE_DEFAULT_STATUS,
                project_owner=request.form.get("project_owner", "").strip() or None,
            )

            外贸负责人 = (扩展字段.get("trade_lead") or "").strip()
            完整度信息 = 计算企业资料完整度(企业, 扩展字段, set())
            扩展字段["material_completeness_score"] = 完整度信息["score"]
            扩展字段["material_completeness"] = 完整度信息["label"]
            扩展字段["material_missing_items"] = 完整度信息["missing_items"]
            企业.enterprise_extra_fields = 扩展字段

            缺失必填字段 = 企业必填缺失字段(企业, 扩展字段)
            if 缺失必填字段:
                flash(f"已按草稿保存，待补充必填项：{'、'.join(缺失必填字段)}", "warning")
            if 主联系人数 > 1:
                flash("联系人子表单中“是否主联系人”只能选择一位。", "danger")
                return render_template("enterprise_form.html", 模式="new", 企业=None, 行业列表=行业下拉选项(), 通用字段组=COMMON_ENTERPRISE_FIELD_GROUPS, 行业字段配置=INDUSTRY_EXTRA_FIELD_CONFIG, 企业扩展字段=扩展字段, 企业文件类型选项=ENTERPRISE_UPLOAD_TYPES, 重复风险提示=[])

            if not 企业编号:
                flash("企业编号为必填项，请手动输入后再保存。", "danger")
                return render_template("enterprise_form.html", 模式="new", 企业=None, 行业列表=行业下拉选项(), 通用字段组=COMMON_ENTERPRISE_FIELD_GROUPS, 行业字段配置=INDUSTRY_EXTRA_FIELD_CONFIG, 企业扩展字段=扩展字段, 企业文件类型选项=ENTERPRISE_UPLOAD_TYPES, 重复风险提示=[])
            if Enterprise.query.filter_by(enterprise_code=企业编号).first():
                flash("企业编号已存在，请修改后再保存。", "danger")
                return render_template("enterprise_form.html", 模式="new", 企业=None, 行业列表=行业下拉选项(), 通用字段组=COMMON_ENTERPRISE_FIELD_GROUPS, 行业字段配置=INDUSTRY_EXTRA_FIELD_CONFIG, 企业扩展字段=扩展字段, 企业文件类型选项=ENTERPRISE_UPLOAD_TYPES, 重复风险提示=["企业编号已存在，请修改后再保存。"])

            重复风险提示 = 企业重复风险检查(企业.company_name, 企业.unified_social_credit_code)
            if 操作动作 == "save" and 重复风险提示:
                for 提示 in 重复风险提示:
                    flash(提示, "warning")

            if 操作动作 == "submit_review":
                缺失字段 = 企业提交审核缺失字段(企业, 扩展字段)
                if 缺失字段:
                    flash(f"提交审核失败，缺失字段：{'、'.join(缺失字段)}", "danger")
                    return render_template("enterprise_form.html", 模式="new", 企业=None, 行业列表=行业下拉选项(), 通用字段组=COMMON_ENTERPRISE_FIELD_GROUPS, 行业字段配置=INDUSTRY_EXTRA_FIELD_CONFIG, 企业扩展字段=扩展字段, 企业文件类型选项=ENTERPRISE_UPLOAD_TYPES, 重复风险提示=[])
                if 重复风险提示 and not 重复确认通过:
                    return render_template("enterprise_form.html", 模式="new", 企业=None, 行业列表=行业下拉选项(), 通用字段组=COMMON_ENTERPRISE_FIELD_GROUPS, 行业字段配置=INDUSTRY_EXTRA_FIELD_CONFIG, 企业扩展字段=扩展字段, 企业文件类型选项=ENTERPRISE_UPLOAD_TYPES, 重复风险提示=重复风险提示, 触发重复风险确认=True)
            企业.status = ENTERPRISE_DEFAULT_STATUS

            db.session.add(企业)
            db.session.flush()
            同步联系人子表单(企业.id, 联系人子表单)
            if 外贸负责人:
                db.session.add(
                    Contact(
                        enterprise_id=企业.id,
                        contact_type="外贸负责人",
                        name=外贸负责人,
                        position="外贸负责人",
                    )
                )
            try:
                上传数 = 处理表单文件上传(
                    enterprise=企业,
                    类型字段名="enterprise_upload_type",
                    名称字段名="enterprise_upload_name",
                    文件字段名="enterprise_upload_file",
                    use_enterprise_naming=True,
                )
            except ValueError as exc:
                flash(str(exc), "danger")
                return render_template("enterprise_form.html", 模式="new", 企业=None, 行业列表=行业下拉选项(), 通用字段组=COMMON_ENTERPRISE_FIELD_GROUPS, 行业字段配置=INDUSTRY_EXTRA_FIELD_CONFIG, 企业扩展字段=扩展字段, 企业文件类型选项=ENTERPRISE_UPLOAD_TYPES, 重复风险提示=[])
            记录审计日志("新增企业", "enterprise", target_id=企业.id, detail=企业.company_name)
            db.session.commit()
            flash(f"企业 {企业.company_name or 企业.unified_social_credit_code or 企业.enterprise_code} 保存成功，上传文件 {上传数} 个。", "success")
            return redirect(url_for("enterprise_edit", id=企业.id))

        return render_template("enterprise_form.html", 模式="new", 企业=None, 行业列表=行业下拉选项(), 通用字段组=COMMON_ENTERPRISE_FIELD_GROUPS, 行业字段配置=INDUSTRY_EXTRA_FIELD_CONFIG, 企业扩展字段={}, 企业文件类型选项=ENTERPRISE_UPLOAD_TYPES, 已上传附件列表=[], 重复风险提示=[])

    @app.route("/enterprises/<int:id>")
    def enterprise_detail(id):
        企业 = Enterprise.query.get_or_404(id)
        企业.industry_display_name = (企业.industry_category or "").strip() or 行业默认名称(企业.industry_code)
        联系人列表 = Contact.query.filter_by(enterprise_id=id).all()
        产品列表 = Product.query.filter_by(enterprise_id=id).all()
        产品映射 = {p.id: p.product_name_cn for p in 产品列表}
        资质列表 = []
        资质展示列表 = [构建证照展示项(资质) for 资质 in 资质列表]
        文件列表 = 附加文件元信息(Document.query.filter_by(enterprise_id=id).all())
        文件类型集合 = {item.document_type for item in 文件列表}
        扩展字段 = 兼容企业基础信息字段(企业, 企业.enterprise_extra_fields or {})
        完整度信息 = 计算企业资料完整度(企业, 扩展字段, 文件类型集合)
        缺失资料提示 = 完整度信息["missing_items"]

        通用分组映射 = {group.get("key"): group for group in COMMON_ENTERPRISE_FIELD_GROUPS}
        行业分组列表 = 行业专项字段组(企业.industry_code)
        详情Tabs = [
            {"key": "basic", "title": "基本信息", "groups": [通用分组映射.get("A")] if 通用分组映射.get("A") else []},
            {"key": "business", "title": "工商信息", "groups": [通用分组映射.get("B")] if 通用分组映射.get("B") else []},
            {"key": "contact", "title": "联系信息", "groups": [通用分组映射.get("C")] if 通用分组映射.get("C") else []},
            {"key": "operations", "title": "经营情况", "groups": [item for item in [通用分组映射.get("D"), *行业分组列表] if item]},
            {"key": "production", "title": "生产能力", "groups": [通用分组映射.get("E")] if 通用分组映射.get("E") else []},
            {"key": "trade", "title": "外贸能力", "groups": [通用分组映射.get("F")] if 通用分组映射.get("F") else []},
            {"key": "finance", "title": "财务信用", "groups": [通用分组映射.get("G")] if 通用分组映射.get("G") else []},
            {"key": "compliance", "title": "资质合规", "groups": [通用分组映射.get("H")] if 通用分组映射.get("H") else []},
            {"key": "attachments", "title": "附件资料", "groups": []},
        ]

        return render_template(
            "enterprise_detail.html",
            企业=企业,
            扩展字段=扩展字段,
            联系人列表=联系人列表,
            产品列表=产品列表,
            资质列表=资质列表,
            资质展示列表=资质展示列表,
            文件列表=文件列表,
            产品映射=产品映射,
            缺失资料提示=缺失资料提示,
            完整度信息=完整度信息,
            详情Tabs=详情Tabs,
        )

    @app.get("/enterprises/<int:id>/export")
    def enterprise_export(id):
        企业 = Enterprise.query.get_or_404(id)
        扩展字段 = 兼容企业基础信息字段(企业, 企业.enterprise_extra_fields or {})
        表头 = ["企业编号", "企业名称", "行业分类", "省份", "城市", "企业性质", "主营产品", "是否有出口经验", "年销售额区间", "年出口额区间", "资料完整度", "最近更新时间"]
        企业性质 = "、".join(
            [
                名称
                for 名称, 开关 in [
                    ("制造商", 企业.is_manufacturer),
                    ("贸易商", 企业.is_trader),
                    ("品牌商", 企业.is_brand_owner),
                    ("OEM/ODM工厂", 企业.is_oem_odm),
                    ("服务商", 企业.is_service_provider),
                ]
                if 开关
            ]
        ) or (企业.company_type or "-")
        行数据 = [[
            企业.enterprise_code,
            企业.company_name,
            行业显示名称(企业.industry_code, 企业.industry_category) or "-",
            企业.province or "-",
            企业.city or "-",
            企业性质,
            企业.main_products or "-",
            "是" if 企业.has_foreign_trade_experience else "否",
            扩展字段.get("annual_sales") or "-",
            扩展字段.get("annual_exports") or "-",
            扩展字段.get("material_completeness") or "-",
            企业.updated_at.strftime("%Y-%m-%d %H:%M"),
        ]]
        缓冲区 = BytesIO()
        文本缓冲 = [",".join([csv_safe(v) for v in 表头])]
        文本缓冲.extend([",".join([csv_safe(v) for v in row]) for row in 行数据])
        缓冲区.write(("\n".join(文本缓冲)).encode("utf-8-sig"))
        缓冲区.seek(0)
        return send_file(
            缓冲区,
            as_attachment=True,
            download_name=f"{企业.enterprise_code}_{企业.company_name}_企业信息.csv",
            mimetype="text/csv",
        )

    @app.get("/enterprises/<int:id>/folder")
    def enterprise_folder_download(id):
        企业 = Enterprise.query.get_or_404(id)
        企业目录 = BASE_DIR / 构建企业归档目录(企业)
        if not 企业目录.exists():
            flash("企业文件夹尚未创建，请先上传企业文件。", "warning")
            return redirect(url_for("enterprise_detail", id=企业.id))
        缓冲区 = BytesIO()
        with ZipFile(缓冲区, "w", compression=ZIP_DEFLATED) as zipf:
            for 文件路径 in 企业目录.rglob("*"):
                if 文件路径.is_file():
                    zipf.write(文件路径, arcname=文件路径.relative_to(企业目录.parent))
        缓冲区.seek(0)
        return send_file(
            缓冲区,
            as_attachment=True,
            download_name=f"{企业.enterprise_code}_{企业.company_name}_企业文件夹.zip",
            mimetype="application/zip",
        )

    @app.route("/enterprises/<int:id>/edit", methods=["GET", "POST"])
    def enterprise_edit(id):
        企业 = Enterprise.query.get_or_404(id)
        外贸联系人 = Contact.query.filter_by(enterprise_id=id, contact_type="外贸负责人").first()

        if request.method == "POST":
            操作动作 = (request.form.get("action") or "save").strip()
            重复确认通过 = (request.form.get("duplicate_confirmed") or "").strip() == "1"
            行业代码, 行业名称 = 解析行业(request.form.get("industry_code"), request.form.get("industry_category"))
            扩展字段 = 提取企业扩展字段(request.form, 行业代码)
            扩展字段 = 兼容企业基础信息字段(企业, 扩展字段)
            企业性质列表 = 扩展字段.get("enterprise_natures", [])
            联系人子表单 = 扩展字段.get("dynamic_contacts", [])
            主联系人数 = sum(1 for item in 联系人子表单 if item.get("is_primary_contact"))
            企业.company_name = (扩展字段.get("company_full_name") or "").strip() or 企业.company_name or f"未命名企业-{企业.enterprise_code}"
            企业.english_name = 扩展字段.get("english_name") or None
            企业.unified_social_credit_code = 扩展字段.get("unified_social_credit_code") or None
            企业.registered_name = 扩展字段.get("registered_name") or 企业.company_name
            企业.legal_representative = 扩展字段.get("legal_representative") or None
            企业.founded_date = 读取日期(扩展字段.get("founded_date"))
            企业.registered_capital = 扩展字段.get("registered_capital") or None
            企业.business_term_start = 读取日期(扩展字段.get("business_term_start"))
            企业.business_term_end = 读取日期(扩展字段.get("business_term_end"))
            企业.registered_address = 扩展字段.get("registered_address") or None
            企业.business_scope = 扩展字段.get("business_scope") or None
            企业.business_address = 扩展字段.get("business_address") or None
            企业.province = request.form.get("province", "").strip() or None
            企业.city = request.form.get("city", "").strip() or None
            企业.district = request.form.get("district", "").strip() or None
            企业.company_type = (扩展字段.get("company_type") or "").strip() or None
            企业.industry_code = 行业代码
            企业.industry_category = 行业名称
            企业.sub_industry = request.form.get("sub_industry", "").strip() or None
            企业.main_products = 扩展字段.get("core_products") or None
            企业.main_business = 扩展字段.get("enterprise_description") or None
            企业.is_manufacturer = "制造商" in 企业性质列表
            企业.is_trader = "贸易商" in 企业性质列表
            企业.is_brand_owner = "品牌商" in 企业性质列表
            企业.is_oem_odm = "OEM/ODM工厂" in 企业性质列表
            企业.is_service_provider = "服务商" in 企业性质列表
            企业.is_high_tech = 读取布尔(request.form, "is_high_tech")
            企业.is_specialized_new = 读取布尔(request.form, "is_specialized_new")
            企业.is_listed_or_pre_ipo = 读取布尔(request.form, "is_listed_or_pre_ipo")
            企业.has_foreign_trade_experience = 扩展字段.get("export_experience") == "是"
            企业.export_countries = 扩展字段.get("export_countries") or None
            企业.target_markets = 扩展字段.get("target_countries") or None
            企业.annual_capacity = 扩展字段.get("annual_capacity_level") or None
            企业.employee_count = 读取整数(request.form.get("employee_count"))
            企业.factory_area = 扩展字段.get("factory_area_range") or None
            企业.main_equipment = request.form.get("main_equipment", "").strip() or None
            企业.annual_revenue = 读取金额(扩展字段.get("annual_sales"))
            企业.export_revenue = 读取金额(扩展字段.get("annual_exports"))
            企业.service_needs = request.form.get("service_needs", "").strip() or None
            企业.risk_notes = request.form.get("risk_notes", "").strip() or None
            企业.enterprise_extra_fields = 扩展字段
            企业.status = ENTERPRISE_DEFAULT_STATUS
            企业.project_owner = request.form.get("project_owner", "").strip() or None
            企业.updated_at = datetime.utcnow()

            外贸负责人 = (扩展字段.get("trade_lead") or "").strip()
            完整度信息 = 计算企业资料完整度(企业, 扩展字段, 提取本次上传文件类型())
            扩展字段["material_completeness_score"] = 完整度信息["score"]
            扩展字段["material_completeness"] = 完整度信息["label"]
            扩展字段["material_missing_items"] = 完整度信息["missing_items"]
            企业.enterprise_extra_fields = 扩展字段
            if 外贸负责人 and not 外贸联系人:
                外贸联系人 = Contact(
                    enterprise_id=企业.id,
                    contact_type="外贸负责人",
                    name=外贸负责人,
                    position="外贸负责人",
                )
                db.session.add(外贸联系人)
            elif 外贸联系人:
                if 外贸负责人:
                    外贸联系人.name = 外贸负责人
                else:
                    软删除对象(外贸联系人)
            同步联系人子表单(企业.id, 联系人子表单)

            缺失必填字段 = 企业必填缺失字段(企业, 扩展字段)
            if 缺失必填字段:
                flash(f"已按草稿保存，待补充必填项：{'、'.join(缺失必填字段)}", "warning")
            if 主联系人数 > 1:
                flash("联系人子表单中“是否主联系人”只能选择一位。", "danger")
                return render_template("enterprise_form.html", 模式="edit", 企业=企业, 外贸负责人=外贸负责人, 行业列表=行业下拉选项(), 通用字段组=COMMON_ENTERPRISE_FIELD_GROUPS, 行业字段配置=INDUSTRY_EXTRA_FIELD_CONFIG, 企业扩展字段=兼容企业基础信息字段(企业, 企业.enterprise_extra_fields or {}), 企业文件类型选项=ENTERPRISE_UPLOAD_TYPES, 重复风险提示=[])
            重复风险提示 = 企业重复风险检查(企业.company_name, 企业.unified_social_credit_code, exclude_enterprise_id=企业.id)
            if 操作动作 == "save" and 重复风险提示:
                for 提示 in 重复风险提示:
                    flash(提示, "warning")
            if 操作动作 == "submit_review":
                缺失字段 = 企业提交审核缺失字段(企业, 扩展字段)
                if 缺失字段:
                    flash(f"提交审核失败，缺失字段：{'、'.join(缺失字段)}", "danger")
                    return render_template("enterprise_form.html", 模式="edit", 企业=企业, 外贸负责人=外贸负责人, 行业列表=行业下拉选项(), 通用字段组=COMMON_ENTERPRISE_FIELD_GROUPS, 行业字段配置=INDUSTRY_EXTRA_FIELD_CONFIG, 企业扩展字段=兼容企业基础信息字段(企业, 企业.enterprise_extra_fields or {}), 企业文件类型选项=ENTERPRISE_UPLOAD_TYPES, 重复风险提示=[])
                if 重复风险提示 and not 重复确认通过:
                    return render_template("enterprise_form.html", 模式="edit", 企业=企业, 外贸负责人=外贸负责人, 行业列表=行业下拉选项(), 通用字段组=COMMON_ENTERPRISE_FIELD_GROUPS, 行业字段配置=INDUSTRY_EXTRA_FIELD_CONFIG, 企业扩展字段=兼容企业基础信息字段(企业, 企业.enterprise_extra_fields or {}), 企业文件类型选项=ENTERPRISE_UPLOAD_TYPES, 重复风险提示=重复风险提示, 触发重复风险确认=True)
            企业.status = ENTERPRISE_DEFAULT_STATUS

            try:
                上传数 = 处理表单文件上传(
                    enterprise=企业,
                    类型字段名="enterprise_upload_type",
                    名称字段名="enterprise_upload_name",
                    文件字段名="enterprise_upload_file",
                    use_enterprise_naming=True,
                )
            except ValueError as exc:
                flash(str(exc), "danger")
                return render_template("enterprise_form.html", 模式="edit", 企业=企业, 外贸负责人=外贸负责人, 行业列表=行业下拉选项(), 通用字段组=COMMON_ENTERPRISE_FIELD_GROUPS, 行业字段配置=INDUSTRY_EXTRA_FIELD_CONFIG, 企业扩展字段=兼容企业基础信息字段(企业, 企业.enterprise_extra_fields or {}), 企业文件类型选项=ENTERPRISE_UPLOAD_TYPES, 重复风险提示=[])
            记录审计日志("编辑企业", "enterprise", target_id=企业.id, detail=企业.company_name)
            db.session.commit()
            flash(f"企业信息更新成功，新增上传文件 {上传数} 个。", "success")
            return redirect(url_for("enterprise_edit", id=企业.id))

        return render_template(
            "enterprise_form.html",
            模式="edit",
            企业=企业,
            外贸负责人=外贸联系人.name if 外贸联系人 else "",
            行业列表=行业下拉选项(),
            通用字段组=COMMON_ENTERPRISE_FIELD_GROUPS,
            行业字段配置=INDUSTRY_EXTRA_FIELD_CONFIG,
            企业扩展字段=兼容企业基础信息字段(企业, 企业.enterprise_extra_fields or {}),
            企业文件类型选项=ENTERPRISE_UPLOAD_TYPES,
            已上传附件列表=Document.query.filter_by(enterprise_id=企业.id).order_by(Document.uploaded_at.desc(), Document.id.desc()).all(),
            重复风险提示=[],
        )

    @app.route("/enterprises/<int:id>/delete", methods=["POST"])
    @admin_required
    def enterprise_delete(id):
        企业 = Enterprise.query.get_or_404(id)
        if request.form.get("confirm_delete") != "YES":
            flash("请勾选二次确认后再删除企业。", "warning")
            return redirect(url_for("enterprise_list"))
        记录审计日志("删除企业", "enterprise", target_id=企业.id, detail=企业.company_name)
        软删除对象(企业)
        for product in Product.query.filter_by(enterprise_id=企业.id, is_deleted=False).all():
            软删除对象(product)
            for sku in ProductSKU.query.filter_by(product_id=product.id, is_deleted=False).all():
                软删除对象(sku)
        for document in Document.query.filter(Document.is_deleted.is_(False), Document.enterprise_id == 企业.id).all():
            软删除对象(document)
        db.session.commit()
        flash("企业已删除", "success")
        return redirect(url_for("enterprise_list"))

    @app.route("/登录", methods=["GET", "POST"])
    def 登录():
        if session.get("用户"):
            return redirect(url_for("dashboard"))
        if request.method == "POST":
            用户名 = request.form.get("用户名", "").strip()
            密码 = request.form.get("密码", "").strip()

            用户 = User.query.filter_by(username=用户名).first()
            if 用户 and check_password_hash(用户.password, 密码):
                session["用户"] = 用户.username
                session["角色"] = 用户.role
                flash("登录成功", "success")
                return redirect(url_for("dashboard"))

            flash("用户名或密码错误", "danger")

        return render_template("login.html")

    @app.route("/退出")
    def 退出():
        session.pop("用户", None)
        session.pop("角色", None)
        flash("已退出登录", "info")
        return redirect(url_for("登录"))

    @app.route("/products")
    def product_list():
        page = request.args.get("page", 1, type=int)
        q_enterprise = request.args.get("enterprise", "").strip()
        q_keyword = request.args.get("keyword", "").strip()
        q_category = request.args.get("category", "").strip()
        q_product_type = request.args.get("product_type", "").strip()
        q_export_suitability = request.args.get("export_suitability", "").strip()
        q_recommendation_level = request.args.get("recommendation_level", "").strip()
        q_certification = request.args.get("certification_status", "").strip()
        q_target_market = request.args.get("target_market", "").strip()
        q_industry = request.args.get("industry", "").strip()
        q_status = request.args.get("status", "").strip() or "active"

        query = Product.query.join(Enterprise, Product.enterprise_id == Enterprise.id)

        if q_enterprise:
            if q_enterprise.isdigit():
                query = query.filter(Product.enterprise_id == int(q_enterprise))
            else:
                query = query.filter(Enterprise.company_name.ilike(f"%{q_enterprise}%"))
        if q_keyword:
            keyword_like = f"%{q_keyword}%"
            query = query.filter(
                or_(
                    Product.product_name_cn.ilike(keyword_like),
                    Product.product_name_en.ilike(keyword_like),
                    Product.model.ilike(keyword_like),
                )
            )
        if q_category:
            query = query.filter(Product.product_category.ilike(f"%{q_category}%"))
        if q_product_type:
            query = query.filter(Product.product_type == q_product_type)
        if q_export_suitability:
            query = query.filter(Product.export_suitability == q_export_suitability)
        if q_recommendation_level:
            query = query.filter(Product.recommendation_level == q_recommendation_level)
        if q_certification:
            query = query.filter(Product.certification_status == q_certification)
        if q_target_market:
            query = query.filter(Product.target_market.ilike(f"%{q_target_market}%"))
        if q_industry:
            query = query.filter(Product.industry_code == q_industry)
        if q_status in {"active", "inactive"}:
            query = query.filter(Product.status == q_status)

        分页 = query.order_by(Product.updated_at.desc()).paginate(page=page, per_page=PER_PAGE, error_out=False)
        enterprises = Enterprise.query.order_by(Enterprise.company_name.asc()).all()
        categories = [
            r[0]
            for r in db.session.query(Product.product_category)
            .filter(Product.product_category.isnot(None), Product.product_category != "")
            .distinct()
            .order_by(Product.product_category)
            .all()
        ]
        target_market_options = [
            r[0]
            for r in db.session.query(Product.target_market)
            .filter(Product.target_market.isnot(None), Product.target_market != "")
            .distinct()
            .order_by(Product.target_market)
            .all()
        ]
        return render_template(
            "products/list.html",
            分页=分页,
            enterprises=enterprises,
            categories=categories,
            industries=行业下拉选项(),
            product_type_options=Product.PRODUCT_TYPE_OPTIONS,
            export_suitability_options=Product.EXPORT_SUITABILITY_OPTIONS,
            recommendation_level_options=Product.RECOMMENDATION_LEVEL_OPTIONS,
            certification_status_options=Product.CERTIFICATION_STATUS_OPTIONS,
            target_market_options=target_market_options,
            filters=request.args,
            product_export_columns=PRODUCT_EXPORT_COLUMNS,
        )

    @app.get("/products/check-code")
    def check_product_code():
        product_code = request.args.get("product_code", "", type=str).strip()
        current_id = request.args.get("current_id", type=int)
        exists = False
        if product_code:
            query = Product.query.filter(Product.product_code == product_code)
            if current_id:
                query = query.filter(Product.id != current_id)
            exists = query.first() is not None
        return jsonify({"exists": exists})

    @app.route("/products/new", methods=["GET", "POST"])
    def product_new():
        enterprises = Enterprise.query.order_by(Enterprise.company_name.asc()).all()
        当前标签 = request.form.get("_active_tab") or request.args.get("tab", "overview")
        if request.method == "POST":
            enterprise_id = request.form.get("enterprise_id", type=int)
            enterprise = Enterprise.query.get(enterprise_id) if enterprise_id else None
            if not enterprise:
                flash("请选择有效的所属企业。", "danger")
                return render_template(
                    "products/form.html",
                    form_action=url_for("product_new"),
                    enterprises=enterprises,
                    form_title="新增产品",
                    sections=PRODUCT_FORM_SECTIONS,
                    product=None,
                    industries=行业下拉选项(),
                    common_extra_groups=COMMON_PRODUCT_FIELD_GROUPS,
                    industry_extra_groups=INDUSTRY_PRODUCT_EXTRA_FIELD_CONFIG,
                    product_extra_values={},
                    产品文件类型选项=PRODUCT_UPLOAD_TYPES,
                    is_new_product=True,
                    initial_tab=当前标签,
                )

            product = Product(
                enterprise_id=enterprise.id,
            )
            try:
                fill_product_from_form(product, request.form)
            except ValueError as exc:
                flash(str(exc), "danger")
                return render_template(
                    "products/form.html",
                    form_action=url_for("product_new"),
                    enterprises=enterprises,
                    form_title="新增产品",
                    sections=PRODUCT_FORM_SECTIONS,
                    product=product,
                    industries=行业下拉选项(),
                    common_extra_groups=COMMON_PRODUCT_FIELD_GROUPS,
                    industry_extra_groups=INDUSTRY_PRODUCT_EXTRA_FIELD_CONFIG,
                    product_extra_values=兼容产品基础信息字段(product, product.product_extra_fields or {}),
                    产品文件类型选项=PRODUCT_UPLOAD_TYPES,
                    is_new_product=True,
                    initial_tab=当前标签,
                )
            db.session.add(product)
            db.session.flush()
            try:
                附件主图 = request.form.get("main_image_from_attachment", "").strip()
                if 附件主图:
                    product.main_image = 附件主图
                上传主图 = 处理产品主图上传(product, enterprise)
                if 上传主图:
                    product.main_image = 上传主图
                上传数 = 处理表单文件上传(
                    enterprise=enterprise,
                    product=product,
                    类型字段名="product_upload_type",
                    名称字段名="product_upload_name",
                    文件字段名="product_upload_file",
                    use_product_naming=True,
                )
            except ValueError as exc:
                flash(str(exc), "danger")
                return render_template(
                    "products/form.html",
                    form_action=url_for("product_new"),
                    enterprises=enterprises,
                    form_title="新增产品",
                    sections=PRODUCT_FORM_SECTIONS,
                    product=product,
                    industries=行业下拉选项(),
                    common_extra_groups=COMMON_PRODUCT_FIELD_GROUPS,
                    industry_extra_groups=INDUSTRY_PRODUCT_EXTRA_FIELD_CONFIG,
                    product_extra_values=兼容产品基础信息字段(product, product.product_extra_fields or {}),
                    产品文件类型选项=PRODUCT_UPLOAD_TYPES,
                    is_new_product=True,
                    product_files=[],
                    initial_tab=当前标签,
                )
            记录审计日志("新增产品", "product", target_id=product.id, detail=product.product_name_cn)
            db.session.commit()
            flash(f"产品已保存，编号：{product.product_code}，上传文件 {上传数} 个。", "success")
            return redirect(url_for("product_edit", product_id=product.id, tab=当前标签 or "overview"))

        return render_template(
            "products/form.html",
            form_action=url_for("product_new"),
            enterprises=enterprises,
            form_title="新增产品",
            sections=PRODUCT_FORM_SECTIONS,
            product=None,
            industries=行业下拉选项(),
            common_extra_groups=COMMON_PRODUCT_FIELD_GROUPS,
            industry_extra_groups=INDUSTRY_PRODUCT_EXTRA_FIELD_CONFIG,
            product_extra_values={},
            产品文件类型选项=PRODUCT_UPLOAD_TYPES,
            is_new_product=True,
            product_files=[],
            initial_tab=当前标签,
        )

    @app.route("/products/<int:product_id>")
    def product_detail(product_id):
        product = Product.query.get_or_404(product_id)
        enterprise = Enterprise.query.get(product.enterprise_id)
        sku_filters = 读取SKU筛选条件(request.args)
        skus = 查询SKU列表(product.id, request.args).all()
        sku_filter_options = SKU筛选选项(product.id)
        certificates = []
        product_files = 附加文件元信息(Document.query.filter_by(product_id=product.id).order_by(Document.uploaded_at.desc()).all())
        product_extra_values = 兼容产品基础信息字段(product, product.product_extra_fields or {})
        附件统计, _ = 构建产品附件统计与状态(product_files, product_extra_values)
        匹配需求列表 = []
        archive_code = f"{enterprise.enterprise_code}_{product.product_code}" if enterprise else product.product_code
        资料缺失提示 = 构建产品资料缺失提示(product, enterprise=enterprise, product_files=product_files, skus=skus)
        return render_template(
            "products/detail.html",
            product=product,
            skus=skus,
            enterprise=enterprise,
            certificates=certificates,
            product_files=product_files,
            archive_code=archive_code,
            匹配需求列表=匹配需求列表,
            资料缺失提示=资料缺失提示,
            sku_filters=sku_filters,
            sku_filter_options=sku_filter_options,
            附件统计=附件统计,
            product_extra_display_groups=构建产品扩展信息分组(
                product.industry_code, product_extra_values
            ),
        )

    @app.post("/products/<int:product_id>/skus")
    def product_sku_create(product_id):
        product = Product.query.get_or_404(product_id)
        sku = ProductSKU(product_id=product.id)
        填充SKU字段(sku, request.form, product=product)
        if not sku.sku_name:
            flash("SKU名称为必填项。", "danger")
            return redirect(获取SKU返回地址(product.id))
        db.session.add(sku)
        db.session.commit()
        记录审计日志("新增SKU", "product_sku", target_id=sku.id, detail=f"{product.product_code}:{sku.sku_code}")
        flash(f"SKU 已新增：{sku.sku_code}", "success")
        return redirect(获取SKU返回地址(product.id))

    @app.post("/products/<int:product_id>/skus/<int:sku_id>/update")
    def product_sku_update(product_id, sku_id):
        product = Product.query.get_or_404(product_id)
        sku = ProductSKU.query.filter_by(id=sku_id, product_id=product.id).first_or_404()
        填充SKU字段(sku, request.form, product=product, 自动生成编码=False)
        if not sku.sku_name:
            flash("SKU名称为必填项。", "danger")
            return redirect(获取SKU返回地址(product.id))
        db.session.commit()
        记录审计日志("编辑SKU", "product_sku", target_id=sku.id, detail=f"{product.product_code}:{sku.sku_code}")
        flash(f"SKU 已更新：{sku.sku_code}", "success")
        return redirect(获取SKU返回地址(product.id))

    @app.post("/products/<int:product_id>/skus/<int:sku_id>/delete")
    def product_sku_delete(product_id, sku_id):
        product = Product.query.get_or_404(product_id)
        sku = ProductSKU.query.filter_by(id=sku_id, product_id=product.id).first_or_404()
        code = sku.sku_code
        软删除对象(sku)
        db.session.commit()
        记录审计日志("删除SKU", "product_sku", target_id=sku_id, detail=f"{product.product_code}:{code}")
        flash(f"SKU 已删除：{code}", "success")
        return redirect(获取SKU返回地址(product.id))

    @app.post("/products/<int:product_id>/skus/batch-update")
    def product_sku_batch_update(product_id):
        product = Product.query.get_or_404(product_id)
        sku_ids = request.form.getlist("sku_ids")
        if not sku_ids:
            flash("请先勾选至少一个 SKU。", "warning")
            return redirect(获取SKU返回地址(product.id))
        try:
            ids = [int(item) for item in sku_ids]
        except ValueError:
            flash("批量更新参数无效。", "danger")
            return redirect(获取SKU返回地址(product.id))
        skus = ProductSKU.query.filter(ProductSKU.product_id == product.id, ProductSKU.id.in_(ids)).all()
        if not skus:
            flash("未找到可更新的 SKU。", "warning")
            return redirect(获取SKU返回地址(product.id))
        for sku in skus:
            for 字段名 in ("moq", "delivery_cycle", "currency", "stock_status", "notes"):
                值 = request.form.get(字段名, "").strip()
                if 值:
                    setattr(sku, 字段名, 值)
            if request.form.get("price", "").strip():
                sku.price = 读取金额(request.form.get("price"))
            if request.form.get("fob_price", "").strip():
                sku.fob_price = 读取金额(request.form.get("fob_price"))
                sku.price = sku.fob_price
            样品值 = request.form.get("sample_available", "").strip()
            if 样品值 in {"0", "1"}:
                sku.sample_available = 样品值 == "1"
            定制值 = request.form.get("customization_supported", "").strip()
            if 定制值 in {"0", "1"}:
                sku.customization_supported = 定制值 == "1"
        db.session.commit()
        记录审计日志("批量编辑SKU", "product_sku", target_id=product.id, detail=f"count={len(skus)}")
        flash(f"已批量更新 {len(skus)} 个 SKU。", "success")
        return redirect(获取SKU返回地址(product.id))

    @app.get("/products/<int:product_id>/skus/export")
    def product_sku_export(product_id):
        product = Product.query.get_or_404(product_id)
        skus = 查询SKU列表(product.id, request.args).all()
        工作簿 = Workbook()
        工作表 = 工作簿.active
        工作表.title = "SKU明细"
        表头 = SKU导入导出字段()
        工作表.append(表头)
        for sku in skus:
            工作表.append([
                sku.sku_code,
                sku.sku_name,
                sku.model,
                sku.specification,
                sku.color,
                sku.size,
                sku.material,
                sku.unit,
                sku.package_spec,
                sku.moq,
                float(sku.price) if sku.price is not None else "",
                float(sku.exw_price) if sku.exw_price is not None else "",
                float(sku.fob_price) if sku.fob_price is not None else "",
                float(sku.cif_price) if sku.cif_price is not None else "",
                float(sku.ddp_price) if sku.ddp_price is not None else "",
                sku.gross_weight,
                sku.net_weight,
                sku.weight,
                sku.delivery_cycle,
                sku.currency or "USD",
                sku.stock_status,
                "是" if sku.sample_available else "否",
                "是" if sku.customization_supported else "否",
                sku.notes,
                sku.created_at.strftime("%Y-%m-%d %H:%M:%S") if sku.created_at else "",
                sku.updated_at.strftime("%Y-%m-%d %H:%M:%S") if sku.updated_at else "",
            ])
        缓冲区 = BytesIO()
        工作簿.save(缓冲区)
        缓冲区.seek(0)
        文件名 = f"{product.product_code}_SKU明细.xlsx"
        return send_file(
            缓冲区,
            as_attachment=True,
            download_name=文件名,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.get("/products/<int:product_id>/skus/import-template")
    def product_sku_import_template(product_id):
        product = Product.query.get_or_404(product_id)
        工作簿 = Workbook()
        工作表 = 工作簿.active
        工作表.title = "SKU导入模板"
        工作表.append(SKU导入导出字段())
        工作表.append([
            f"{product.product_code}-S001",
            "示例SKU",
            "M-001",
            "220V",
            "黑色",
            "L",
            "铝合金",
            "件",
            "20pcs/箱",
            "100",
            12.8,
            10.5,
            12.8,
            14.2,
            15.6,
            "0.6kg",
            "0.5kg",
            "0.5kg",
            "30天",
            "USD",
            "现货",
            "是",
            "否",
            "示例备注",
            "",
            "",
        ])
        缓冲区 = BytesIO()
        工作簿.save(缓冲区)
        缓冲区.seek(0)
        return send_file(
            缓冲区,
            as_attachment=True,
            download_name=f"{product.product_code}_SKU导入模板.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    @app.post("/products/<int:product_id>/skus/import")
    def product_sku_import(product_id):
        product = Product.query.get_or_404(product_id)
        上传文件 = request.files.get("file")
        if not 上传文件 or not 上传文件.filename:
            flash("请先选择 SKU Excel 文件。", "danger")
            return redirect(获取SKU返回地址(product.id))
        成功条数, 失败列表 = 导入SKUExcel(product, 上传文件)
        记录审计日志("导入SKU", "product_sku", target_id=product.id, detail=f"success={成功条数}, failed={len(失败列表)}")
        return render_template(
            "import_result.html",
            标题=f"{product.product_name_cn} SKU导入结果",
            成功条数=成功条数,
            失败列表=失败列表,
            返回地址=url_for("product_edit", product_id=product.id, tab="sku"),
            返回文案="返回 SKU 明细",
        )

    @app.get("/products/<int:product_id>/export")
    def product_export(product_id):
        product = Product.query.get_or_404(product_id)
        enterprise = Enterprise.query.get(product.enterprise_id)
        product_extra = 兼容产品基础信息字段(product, product.product_extra_fields or {})
        表头 = ["产品编号", "产品名称", "所属企业", "行业分类", "产品类别", "HS编码", "MOQ", "生产周期", "FOB价", "主要认证", "是否可样品", "是否支持定制", "是否适合跨境电商", "是否适合工程采购", "是否适合经销代理", "最近更新时间"]
        行数据 = [[
            product.product_code,
            product.product_name_cn,
            enterprise.company_name if enterprise else "-",
            行业显示名称(product.industry_code, product.industry_name) or "-",
            product.product_category or "-",
            product.hs_code or "-",
            product.moq or product_extra.get("trade_moq") or "-",
            product.production_cycle or product_extra.get("trade_mass_cycle") or "-",
            f"{product.currency or 'USD'} {product.fob_price}" if product.fob_price is not None else "-",
            product.certifications or "-",
            "否" if product_extra.get("trade_sample_policy") == "不支持样品" else "是",
            product_extra.get("support_customization") or ("是" if product.customization_supported else "否"),
            product_extra.get("fit_cross_border") or "待判断",
            product_extra.get("fit_engineering") or "待判断",
            product_extra.get("fit_distributor") or "待判断",
            product.updated_at.strftime("%Y-%m-%d %H:%M"),
        ]]
        缓冲区 = BytesIO()
        文本缓冲 = [",".join([csv_safe(v) for v in 表头])]
        文本缓冲.extend([",".join([csv_safe(v) for v in row]) for row in 行数据])
        缓冲区.write(("\n".join(文本缓冲)).encode("utf-8-sig"))
        缓冲区.seek(0)
        return send_file(
            缓冲区,
            as_attachment=True,
            download_name=f"{product.product_code}_{product.product_name_cn}_产品信息.csv",
            mimetype="text/csv",
        )

    @app.route("/products/<int:product_id>/edit", methods=["GET", "POST"])
    def product_edit(product_id):
        product = Product.query.get_or_404(product_id)
        enterprises = Enterprise.query.order_by(Enterprise.company_name.asc()).all()
        product_files = 附加文件元信息(Document.query.filter_by(product_id=product.id).order_by(Document.uploaded_at.desc()).all())
        sku_list = 查询SKU列表(product.id, request.args).all()
        enterprise = Enterprise.query.get(product.enterprise_id) if product.enterprise_id else None
        product_extra_values = 兼容产品基础信息字段(product, product.product_extra_fields or {})
        附件统计, 状态建议 = 构建产品附件统计与状态(product_files, product_extra_values)
        资料缺失提示 = 构建产品资料缺失提示(product, enterprise=enterprise, product_files=product_files, skus=sku_list)
        当前标签 = request.form.get("_active_tab") or request.args.get("tab", "overview")
        if request.method == "POST":
            enterprise_id = request.form.get("enterprise_id", type=int)
            enterprise = Enterprise.query.get(enterprise_id) if enterprise_id else None
            if not enterprise:
                flash("请选择有效的所属企业。", "danger")
                return render_template(
                    "products/form.html",
                    form_action=url_for("product_edit", product_id=product.id),
                    enterprises=enterprises,
                    form_title="编辑产品",
                    sections=PRODUCT_FORM_SECTIONS,
                    product=product,
                    industries=行业下拉选项(),
                    common_extra_groups=COMMON_PRODUCT_FIELD_GROUPS,
                    industry_extra_groups=INDUSTRY_PRODUCT_EXTRA_FIELD_CONFIG,
                    product_extra_values=product_extra_values if product else {},
                    产品文件类型选项=PRODUCT_UPLOAD_TYPES,
                    is_new_product=False,
                    product_files=product_files,
                    initial_tab=当前标签,
                    资料缺失提示=资料缺失提示,
                )

            product.enterprise_id = enterprise.id
            try:
                fill_product_from_form(product, request.form)
            except ValueError as exc:
                flash(str(exc), "danger")
                return render_template(
                    "products/form.html",
                    form_action=url_for("product_edit", product_id=product.id),
                    enterprises=enterprises,
                    form_title="编辑产品",
                    sections=PRODUCT_FORM_SECTIONS,
                    product=product,
                    industries=行业下拉选项(),
                    common_extra_groups=COMMON_PRODUCT_FIELD_GROUPS,
                    industry_extra_groups=INDUSTRY_PRODUCT_EXTRA_FIELD_CONFIG,
                    product_extra_values=product_extra_values if product else {},
                    产品文件类型选项=PRODUCT_UPLOAD_TYPES,
                    is_new_product=False,
                    product_files=product_files,
                    initial_tab=当前标签,
                    资料缺失提示=资料缺失提示,
                )
            try:
                附件主图 = request.form.get("main_image_from_attachment", "").strip()
                if 附件主图:
                    product.main_image = 附件主图
                上传主图 = 处理产品主图上传(product, enterprise)
                if 上传主图:
                    product.main_image = 上传主图
                上传数 = 处理表单文件上传(
                    enterprise=enterprise,
                    product=product,
                    类型字段名="product_upload_type",
                    名称字段名="product_upload_name",
                    文件字段名="product_upload_file",
                    use_product_naming=True,
                )
            except ValueError as exc:
                flash(str(exc), "danger")
                return render_template(
                    "products/form.html",
                    form_action=url_for("product_edit", product_id=product.id),
                    enterprises=enterprises,
                    form_title="编辑产品",
                    sections=PRODUCT_FORM_SECTIONS,
                    product=product,
                    industries=行业下拉选项(),
                    common_extra_groups=COMMON_PRODUCT_FIELD_GROUPS,
                    industry_extra_groups=INDUSTRY_PRODUCT_EXTRA_FIELD_CONFIG,
                    product_extra_values=product_extra_values if product else {},
                    产品文件类型选项=PRODUCT_UPLOAD_TYPES,
                    is_new_product=False,
                    product_files=product_files,
                    initial_tab=当前标签,
                    资料缺失提示=资料缺失提示,
                )
            记录审计日志("编辑产品", "product", target_id=product.id, detail=product.product_name_cn)
            db.session.commit()
            flash(f"产品信息已更新，新增上传文件 {上传数} 个。", "success")
            return redirect(url_for("product_edit", product_id=product.id, tab=当前标签))

        return render_template(
            "products/form.html",
            form_action=url_for("product_edit", product_id=product.id),
            enterprises=enterprises,
            form_title="编辑产品",
            sections=PRODUCT_FORM_SECTIONS,
            product=product,
            industries=行业下拉选项(),
            common_extra_groups=COMMON_PRODUCT_FIELD_GROUPS,
            industry_extra_groups=INDUSTRY_PRODUCT_EXTRA_FIELD_CONFIG,
            product_extra_values=product_extra_values,
            产品文件类型选项=PRODUCT_UPLOAD_TYPES,
            is_new_product=False,
            product_files=product_files,
            initial_tab=当前标签,
            sku_list=sku_list,
            sku_filters=读取SKU筛选条件(request.args),
            sku_filter_options=SKU筛选选项(product.id),
            资料缺失提示=资料缺失提示,
            附件统计=附件统计,
            资料状态建议=状态建议,
        )

    @app.post("/products/<int:product_id>/delete")
    @admin_required
    def product_delete(product_id):
        product = Product.query.get_or_404(product_id)
        if request.form.get("confirm_delete") != "YES":
            flash("请勾选二次确认后再删除产品。", "warning")
            return redirect(url_for("product_list"))
        记录审计日志("删除产品", "product", target_id=product.id, detail=product.product_name_cn)
        软删除对象(product)
        for sku in ProductSKU.query.filter_by(product_id=product.id, is_deleted=False).all():
            软删除对象(sku)
        for document in Document.query.filter_by(product_id=product.id, is_deleted=False).all():
            软删除对象(document)
        db.session.commit()
        flash("产品已删除。", "info")
        return redirect(url_for("product_list"))

    @app.post("/products/<int:product_id>/attachments/<int:document_id>/set-main-image")
    def product_attachment_set_main_image(product_id, document_id):
        product = Product.query.get_or_404(product_id)
        document = Document.query.filter_by(id=document_id, product_id=product.id).first_or_404()
        if document.document_type != "产品图片":
            flash("仅产品图片可设为主图。", "warning")
            return redirect(url_for("product_edit", product_id=product.id, tab="attachment"))
        product.main_image = f"/{document.file_path}"
        记录审计日志("设为产品主图", "product", target_id=product.id, detail=document.document_name)
        db.session.commit()
        flash("主图已更新。", "success")
        return redirect(url_for("product_edit", product_id=product.id, tab="attachment"))

    @app.post("/products/<int:product_id>/attachments/<int:document_id>/delete")
    def product_attachment_delete(product_id, document_id):
        product = Product.query.get_or_404(product_id)
        document = Document.query.filter_by(id=document_id, product_id=product.id).first_or_404()
        if product.main_image == f"/{document.file_path}":
            product.main_image = None
        记录审计日志("删除产品附件", "document", target_id=document.id, detail=document.document_name)
        软删除对象(document)
        db.session.commit()
        flash("附件已删除。", "success")
        return redirect(url_for("product_edit", product_id=product.id, tab="attachment"))

    @app.post("/products/<int:product_id>/toggle-status")
    def product_toggle_status(product_id):
        product = Product.query.get_or_404(product_id)
        action = request.form.get("action", "").strip()
        if action not in {"disable", "enable"}:
            flash("无效操作。", "danger")
            return redirect(url_for("product_list"))
        product.status = "inactive" if action == "disable" else "active"
        记录审计日志("更新产品状态", "product", target_id=product.id, detail=f"{product.product_name_cn} -> {product.status}")
        db.session.commit()
        flash("产品状态已更新。", "success")
        return redirect(url_for("product_list"))

    @app.route("/documents")
    def document_list():
        enterprise_id = request.args.get("enterprise_id", type=int)
        product_id = request.args.get("product_id", type=int)
        document_type = request.args.get("document_type", "", type=str).strip()
        upload_date = request.args.get("upload_date", "", type=str).strip()
        keyword = request.args.get("keyword", "", type=str).strip()

        query = Document.query
        if enterprise_id:
            query = query.filter(Document.enterprise_id == enterprise_id)
        if product_id:
            query = query.filter(Document.product_id == product_id)
        if document_type:
            query = query.filter(Document.document_type == document_type)
        if upload_date:
            日期 = 读取日期(upload_date)
            if 日期:
                query = query.filter(db.func.date(Document.uploaded_at) == 日期)
            else:
                flash("上传日期格式无效，已忽略该筛选。", "warning")
        if keyword:
            query = query.filter(Document.document_name.ilike(f"%{keyword}%"))

        documents = query.order_by(Document.uploaded_at.desc(), Document.id.desc()).all()
        enterprises = Enterprise.query.order_by(Enterprise.company_name.asc()).all()
        products = Product.query.filter_by(enterprise_id=enterprise_id).order_by(Product.product_name_cn.asc()).all() if enterprise_id else Product.query.order_by(Product.product_name_cn.asc()).all()
        enterprise_map = {item.id: item for item in Enterprise.query.filter(Enterprise.id.in_({d.enterprise_id for d in documents if d.enterprise_id})).all()} if documents else {}
        product_map = {item.id: item for item in Product.query.filter(Product.id.in_({d.product_id for d in documents if d.product_id})).all()} if documents else {}

        return render_template(
            "documents/list.html",
            documents=documents,
            enterprises=enterprises,
            products=products,
            enterprise_map=enterprise_map,
            product_map=product_map,
            filters={
                "enterprise_id": enterprise_id,
                "product_id": product_id,
                "document_type": document_type,
                "upload_date": upload_date,
                "keyword": keyword,
            },
            document_types=DOCUMENT_TYPE_OPTIONS,
        )

    @app.route("/documents/upload", methods=["GET", "POST"])
    def document_upload():
        enterprises = Enterprise.query.order_by(Enterprise.company_name.asc()).all()
        all_products = Product.query.order_by(Product.product_name_cn.asc()).all()
        enterprise_options = [
            {
                "id": e.id,
                "code": e.enterprise_code or "",
                "name": e.company_name or "",
                "label": f"{e.enterprise_code or ''} - {e.company_name or ''}",
            }
            for e in enterprises
        ]
        product_options = [
            {"id": p.id, "enterprise_id": p.enterprise_id, "label": f"{p.product_code} - {p.product_name_cn}"}
            for p in all_products
        ]

        def 渲染上传页(products, form_data, document_types):
            return render_template(
                "documents/upload.html",
                enterprises=enterprises,
                products=products,
                all_products=all_products,
                enterprise_options=enterprise_options,
                product_options=product_options,
                document_types=document_types,
                enterprise_document_types=ENTERPRISE_DOCUMENT_TYPE_OPTIONS,
                product_document_types=PRODUCT_DOCUMENT_TYPE_OPTIONS,
                form_data=form_data,
            )

        if request.method == "POST":
            enterprise_id = request.form.get("enterprise_id", type=int)
            enterprise = Enterprise.query.get(enterprise_id) if enterprise_id else None
            if not enterprise:
                flash("所属企业为必填项。", "danger")
                return 渲染上传页([], request.form, ENTERPRISE_DOCUMENT_TYPE_OPTIONS)

            product_id = request.form.get("product_id", type=int)
            product = Product.query.filter_by(id=product_id, enterprise_id=enterprise.id).first() if product_id else None
            if product_id and not product:
                flash("所属产品不属于当前企业，请重新选择。", "danger")
                return 渲染上传页(
                    Product.query.filter_by(enterprise_id=enterprise.id).order_by(Product.product_name_cn.asc()).all(),
                    request.form,
                    PRODUCT_DOCUMENT_TYPE_OPTIONS if product_id else ENTERPRISE_DOCUMENT_TYPE_OPTIONS,
                )

            document_type = request.form.get("document_type", "").strip()
            document_name = request.form.get("document_name", "").strip()
            version = request.form.get("version", "").strip() or "V01"
            uploaded_by = request.form.get("uploaded_by", "").strip() or "未署名"
            notes = request.form.get("notes", "").strip() or None
            if not document_type:
                flash("文件类型为必填项。", "danger")
                return 渲染上传页(
                    Product.query.filter_by(enterprise_id=enterprise.id).order_by(Product.product_name_cn.asc()).all(),
                    request.form,
                    PRODUCT_DOCUMENT_TYPE_OPTIONS if product_id else ENTERPRISE_DOCUMENT_TYPE_OPTIONS,
                )
            允许文件类型 = PRODUCT_UPLOAD_TYPES if product else ENTERPRISE_UPLOAD_TYPES
            if document_type not in 允许文件类型:
                flash("文件类型不在当前归档对象允许范围内。", "danger")
                return 渲染上传页(
                    Product.query.filter_by(enterprise_id=enterprise.id).order_by(Product.product_name_cn.asc()).all(),
                    request.form,
                    PRODUCT_DOCUMENT_TYPE_OPTIONS if product else ENTERPRISE_DOCUMENT_TYPE_OPTIONS,
                )
            if not document_name:
                flash("文件名称为必填项。", "danger")
                return 渲染上传页(
                    Product.query.filter_by(enterprise_id=enterprise.id).order_by(Product.product_name_cn.asc()).all(),
                    request.form,
                    PRODUCT_DOCUMENT_TYPE_OPTIONS if product_id else ENTERPRISE_DOCUMENT_TYPE_OPTIONS,
                )
            上传文件 = request.files.get("file")
            if not 上传文件 or not 上传文件.filename:
                flash("请选择需要上传的文件。", "danger")
                return 渲染上传页(
                    Product.query.filter_by(enterprise_id=enterprise.id).order_by(Product.product_name_cn.asc()).all(),
                    request.form,
                    PRODUCT_DOCUMENT_TYPE_OPTIONS if product_id else ENTERPRISE_DOCUMENT_TYPE_OPTIONS,
                )

            扩展名 = Path(上传文件.filename).suffix.lower()
            if 扩展名 in BLOCKED_EXTENSIONS:
                flash("文件类型不允许上传。", "danger")
                return 渲染上传页(
                    Product.query.filter_by(enterprise_id=enterprise.id).order_by(Product.product_name_cn.asc()).all(),
                    request.form,
                    PRODUCT_DOCUMENT_TYPE_OPTIONS if product_id else ENTERPRISE_DOCUMENT_TYPE_OPTIONS,
                )
            if request.content_length and request.content_length > MAX_UPLOAD_SIZE:
                flash("单文件大小不能超过 100MB。", "danger")
                return 渲染上传页(
                    Product.query.filter_by(enterprise_id=enterprise.id).order_by(Product.product_name_cn.asc()).all(),
                    request.form,
                    PRODUCT_DOCUMENT_TYPE_OPTIONS if product_id else ENTERPRISE_DOCUMENT_TYPE_OPTIONS,
                )

            try:
                document = 保存文件并登记记录(
                    app=app,
                    enterprise=enterprise,
                    product=product,
                    上传文件=上传文件,
                    document_type=document_type,
                    document_name=清洗路径片段(document_name),
                    uploaded_by=uploaded_by,
                    notes=notes,
                )
            except ValueError as exc:
                flash(str(exc), "danger")
                return 渲染上传页(
                    Product.query.filter_by(enterprise_id=enterprise.id).order_by(Product.product_name_cn.asc()).all(),
                    request.form,
                    PRODUCT_DOCUMENT_TYPE_OPTIONS if product_id else ENTERPRISE_DOCUMENT_TYPE_OPTIONS,
                )
            document.version = version
            记录审计日志("上传文件", "document", target_id=document.id, detail=document.document_name)
            db.session.commit()
            flash("文件上传并归档成功。", "success")
            return redirect(url_for("document_list"))

        default_enterprise_id = request.args.get("enterprise_id", type=int)
        default_product_id = request.args.get("product_id", type=int)
        products = Product.query.filter_by(enterprise_id=default_enterprise_id).order_by(Product.product_name_cn.asc()).all() if default_enterprise_id else []
        return 渲染上传页(
            products,
            {
                "enterprise_id": default_enterprise_id,
                "product_id": default_product_id,
            } if default_enterprise_id else {},
            PRODUCT_DOCUMENT_TYPE_OPTIONS if default_product_id else ENTERPRISE_DOCUMENT_TYPE_OPTIONS,
        )

    @app.route("/documents/<int:document_id>/download")
    def document_download(document_id):
        document = Document.query.get_or_404(document_id)
        文件路径 = BASE_DIR / document.file_path
        if not 文件路径.exists():
            flash("文件不存在，可能已被移动或删除。", "danger")
            return redirect(url_for("document_list"))
        return send_from_directory(
            文件路径.parent,
            文件路径.name,
            as_attachment=True,
            download_name=document.original_filename or 文件路径.name,
        )

    @app.post("/documents/<int:document_id>/delete")
    @admin_required
    def document_delete(document_id):
        document = Document.query.get_or_404(document_id)
        返回地址 = request.form.get("next", "").strip()
        if request.form.get("confirm_delete") != "YES":
            flash("请勾选二次确认后再删除文件。", "warning")
            return redirect(返回地址 or url_for("document_list"))
        记录审计日志("删除文件", "document", target_id=document.id, detail=document.document_name)
        软删除对象(document)
        db.session.commit()
        flash("文件已删除。", "success")
        return redirect(返回地址 or url_for("document_list"))

    def init_db_command():
        """初始化数据库（可选：flask init-db）。"""

        init_db(app)

    启动定时备份任务()
    return app


def 生成企业编号():
    最新企业 = Enterprise.query.order_by(Enterprise.id.desc()).first()
    if not 最新企业:
        return "E001"

    最新编号 = 最新企业.enterprise_code or "E000"
    try:
        当前数字 = int(最新编号[1:])
    except (ValueError, TypeError):
        当前数字 = Enterprise.query.count()
    return f"E{当前数字 + 1:03d}"


def 获取企业行业分布():
    rows = (
        db.session.query(Enterprise.industry_category, func.count(Enterprise.id))
        .filter(Enterprise.industry_category.isnot(None), Enterprise.industry_category != "")
        .group_by(Enterprise.industry_category)
        .order_by(func.count(Enterprise.id).desc())
        .limit(10)
        .all()
    )
    return {"labels": [row[0] for row in rows], "data": [row[1] for row in rows]}


def 获取产品分类分布():
    rows = (
        db.session.query(Product.product_category, func.count(Product.id))
        .filter(Product.product_category.isnot(None), Product.product_category != "")
        .group_by(Product.product_category)
        .order_by(func.count(Product.id).desc())
        .limit(10)
        .all()
    )
    return {"labels": [row[0] for row in rows], "data": [row[1] for row in rows]}


def generate_product_code(enterprise_id):
    最后产品 = Product.query.filter_by(enterprise_id=enterprise_id).order_by(Product.id.desc()).first()
    if not 最后产品:
        return "P001"
    match = re.search(r"(\d+)$", 最后产品.product_code or "")
    seq = int(match.group(1)) + 1 if match else Product.query.filter_by(enterprise_id=enterprise_id).count() + 1
    return f"P{seq:03d}"


def generate_sku_code(product):
    最后SKU = ProductSKU.query.filter_by(product_id=product.id).order_by(ProductSKU.id.desc()).first()
    if not 最后SKU:
        seq = 1
    else:
        match = re.search(r"-S(\d+)$", 最后SKU.sku_code or "")
        seq = int(match.group(1)) + 1 if match else ProductSKU.query.filter_by(product_id=product.id).count() + 1
    return f"{product.product_code}-S{seq:03d}"


def 产品行业专项字段组(行业代码):
    return INDUSTRY_PRODUCT_EXTRA_FIELD_CONFIG.get((行业代码 or "").strip(), [])


def 产品扩展字段配置(行业代码):
    配置 = []
    for 分组 in COMMON_PRODUCT_FIELD_GROUPS:
        配置.extend(分组.get("fields", []))
    for 分组 in 产品行业专项字段组(行业代码):
        配置.extend(分组.get("fields", []))
    return 配置


def 提取产品扩展字段(form, 行业代码, 原扩展字段=None):
    数据 = dict(原扩展字段 or {})
    for 字段 in 产品扩展字段配置(行业代码):
        键名 = 字段["key"]
        if 键名 not in form:
            continue
        if 字段.get("type") == "checkbox_group":
            值 = [item.strip() for item in form.getlist(键名) if item and item.strip()]
        else:
            原值 = form.get(键名, "")
            值 = 原值.strip() if isinstance(原值, str) else 原值
        if 值 in (None, "", []):
            数据.pop(键名, None)
            continue
        数据[键名] = 值
    return 数据


def 构建产品扩展信息分组(行业代码, 扩展字段):
    扩展字段 = 扩展字段 or {}
    分组数据 = []
    for 分组 in [*COMMON_PRODUCT_FIELD_GROUPS, *产品行业专项字段组(行业代码)]:
        字段列表 = []
        for 字段 in 分组.get("fields", []):
            值 = 扩展字段.get(字段["key"])
            if 值 in (None, "", []):
                continue
            if isinstance(值, list):
                值 = "、".join(值)
            字段列表.append({"label": 字段["label"], "value": 值})
        if 字段列表:
            分组数据.append({"title": 分组["title"], "fields": 字段列表})
    return 分组数据


def 兼容产品基础信息字段(product, 扩展字段):
    数据 = dict(扩展字段 or {})
    映射 = {
        "core_selling_points": ("desc_core_selling_points", product.product_selling_points),
        "trade_sample_policy": ("sample_policy", product.sample_policy),
        "support_customization": ("customization_supported", "是" if product.customization_supported else "否"),
        "cert_status": ("certification_status", product.certification_status),
        "positioning_scenarios": ("desc_scenarios", product.application_scenario),
    }
    for 扩展键, (结构化键, 结构化值) in 映射.items():
        扩展值 = 数据.get(扩展键)
        if 扩展值 in (None, "", []):
            回填值 = 结构化值
            if 回填值 not in (None, "", []):
                数据[扩展键] = 回填值
        elif 结构化值 in (None, "", []):
            数据[结构化键] = 扩展值
    return 数据


def fill_product_from_form(product, form):
    product_code = form.get("product_code", "").strip()
    if not product_code:
        raise ValueError("产品编号为必填项")
    duplicate_query = Product.query.filter(Product.product_code == product_code)
    if product.id:
        duplicate_query = duplicate_query.filter(Product.id != product.id)
    if duplicate_query.first():
        raise ValueError("产品编号已存在，请更换后再保存")
    product.product_code = product_code

    enterprise = Enterprise.query.get(product.enterprise_id) if product.enterprise_id else None
    选中行业代码 = (enterprise.industry_code if enterprise else "") or (form.get("industry_code") or "").strip()
    行业信息 = INDUSTRY_MAP.get(选中行业代码) if 选中行业代码 else None
    if not 行业信息 and enterprise:
        行业信息 = {"code": enterprise.industry_code, "name": enterprise.industry_category}
    if not 行业信息:
        raise ValueError("所属企业未配置行业分类，请先完善企业信息。")

    product.industry_code = 行业信息["code"]
    product.industry_name = 行业信息["name"]
    product.product_name_cn = form.get("product_name_cn", "").strip()
    product.product_name_en = form.get("product_name_en", "").strip() or None
    product.product_category = form.get("product_category", "").strip() or None
    product.hs_code = form.get("hs_code", "").strip() or None
    product.capacity_cycle_days = 读取整数(form.get("capacity_cycle_days"))
    product.capacity_qualified_pieces = 读取整数(form.get("capacity_qualified_pieces"))
    product.function_description = form.get("function_description", "").strip() or None
    if "application_scenario" in form:
        旧应用场景 = form.get("application_scenario", "").strip()
    else:
        旧应用场景 = (product.application_scenario or "").strip()
    应用场景标签 = [item.strip() for item in form.getlist("positioning_scenarios") if item and item.strip()]
    其他应用场景 = form.get("positioning_scenario_other", "").strip()
    if "其他" in 应用场景标签 and 其他应用场景:
        应用场景展示值 = [其他应用场景 if item == "其他" else item for item in 应用场景标签]
    else:
        应用场景展示值 = 应用场景标签
    product.application_scenario = "、".join(应用场景展示值) if 应用场景展示值 else (旧应用场景 or None)
    定制支持文本 = form.get("support_customization", "").strip()
    if 定制支持文本:
        product.customization_supported = 定制支持文本 == "是"
    else:
        product.customization_supported = 读取布尔(form, "customization_supported")
    目标市场标签 = [item.strip() for item in form.getlist("target_market_tags") if item and item.strip()]
    product.target_market = "、".join(目标市场标签) if 目标市场标签 else (form.get("target_market", "").strip() or None)
    product.main_image = form.get("main_image", "").strip() or product.main_image
    product.product_type = form.get("product_type", "").strip() or None
    product.export_suitability = form.get("export_suitability", "").strip() or None
    product.recommendation_level = form.get("recommendation_level", "").strip() or None
    product.existing_sales_countries = form.get("existing_sales_countries", "").strip() or None
    product.certifications = form.get("certifications", "").strip() or None
    认证标签 = [item.strip() for item in form.getlist("cert_product") if item and item.strip()]
    if 认证标签:
        product.certifications = "、".join(认证标签)
    product.certification_status = form.get("cert_status", "").strip() or form.get("certification_status", "").strip() or None
    核心卖点 = form.get("core_selling_points", "").strip() or form.get("product_selling_points", "").strip()
    product.product_selling_points = 核心卖点 or None
    product.notes = form.get("notes", "").strip() or None
    product.product_extra_fields = 提取产品扩展字段(form, product.industry_code, 原扩展字段=product.product_extra_fields)
    if 核心卖点:
        product.product_extra_fields["desc_core_selling_points"] = 核心卖点
    if 应用场景标签:
        product.product_extra_fields["desc_scenarios"] = "、".join(应用场景展示值)
        if 其他应用场景:
            product.product_extra_fields["positioning_scenario_other"] = 其他应用场景
    目标客户标签 = [item.strip() for item in form.getlist("target_customer_tags") if item and item.strip()]
    if 目标客户标签:
        product.product_extra_fields["desc_target_customer"] = "、".join(目标客户标签)
    if 定制支持文本:
        product.product_extra_fields["support_customization"] = 定制支持文本
    product.product_extra_fields = 兼容产品基础信息字段(product, product.product_extra_fields)
    if not product.product_name_cn:
        raise ValueError("产品中文名为必填项")


def 同步联系人子表单(enterprise_id, dynamic_contacts):
    for contact in Contact.query.filter(
        Contact.enterprise_id == enterprise_id,
        Contact.contact_type == "联系人",
        Contact.is_deleted.is_(False),
    ).all():
        contact.soft_delete(session.get("用户") or "system")
    for item in dynamic_contacts or []:
        姓名 = (item.get("name") or "").strip()
        if not 姓名:
            continue
        电话 = (item.get("mobile") or "").strip()
        db.session.add(
            Contact(
                enterprise_id=enterprise_id,
                contact_type="联系人",
                name=姓名,
                department=(item.get("department") or "").strip() or None,
                position=(item.get("position") or "").strip() or None,
                mobile=电话 or None,
                phone=电话 or None,
                email=(item.get("email") or "").strip() or None,
                wechat=(item.get("wechat") or "").strip() or None,
                responsibility=(item.get("responsibility") or "").strip() or None,
                is_primary_contact=bool(item.get("is_primary_contact")),
            )
        )


def 处理产品主图上传(product, enterprise):
    主图文件 = request.files.get("main_image_upload")
    if not 主图文件 or not 主图文件.filename:
        return None
    当前上传人 = (session.get("用户") or "未署名").strip() or "未署名"
    文档 = 保存文件并登记记录(
        app=app,
        enterprise=enterprise,
        product=product,
        上传文件=主图文件,
        document_type="产品图片",
        document_name="产品主图",
        uploaded_by=当前上传人,
    )
    return f"/{文档.file_path}"


def 填充SKU字段(sku, form, product=None, 自动生成编码=True):
    product = product or (Product.query.get(sku.product_id) if sku.product_id else None)
    def 取值(字段名, 默认值=None):
        if 字段名 not in form:
            return 默认值
        return form.get(字段名, "")

    if 自动生成编码:
        sku.sku_code = (取值("sku_code", "").strip() or (generate_sku_code(product) if product else None))
    else:
        sku.sku_code = 取值("sku_code", "").strip() or sku.sku_code
    sku.sku_name = 取值("sku_name", sku.sku_name or "").strip() or None
    sku.model = 取值("model", sku.model or "").strip() or None
    sku.specification = 取值("specification", sku.specification or "").strip() or None
    sku.color = 取值("color", sku.color or "").strip() or None
    sku.size = 取值("size", sku.size or "").strip() or None
    sku.material = 取值("material", sku.material or "").strip() or None
    sku.unit = 取值("unit", sku.unit or "").strip() or None
    sku.package_spec = 取值("package_spec", sku.package_spec or "").strip() or None
    sku.moq = 取值("moq", sku.moq or "").strip() or None
    if "price" in form:
        sku.price = 读取金额(取值("price"))
    if "exw_price" in form:
        sku.exw_price = 读取金额(取值("exw_price"))
    if "fob_price" in form:
        sku.fob_price = 读取金额(取值("fob_price"))
        sku.price = sku.fob_price
    if "cif_price" in form:
        sku.cif_price = 读取金额(取值("cif_price"))
    if "ddp_price" in form:
        sku.ddp_price = 读取金额(取值("ddp_price"))
    sku.gross_weight = 取值("gross_weight", sku.gross_weight or "").strip() or None
    sku.net_weight = 取值("net_weight", sku.net_weight or "").strip() or None
    sku.weight = 取值("weight", sku.weight or "").strip() or None
    sku.delivery_cycle = 取值("delivery_cycle", sku.delivery_cycle or "").strip() or None
    sku.currency = 取值("currency", sku.currency or "USD").strip() or "USD"
    sku.stock_status = 取值("stock_status", sku.stock_status or "").strip() or None
    if "sample_available" in form or sku.id is None:
        sku.sample_available = 读取布尔(form, "sample_available")
    if "customization_supported" in form or sku.id is None:
        sku.customization_supported = 读取布尔(form, "customization_supported")
    sku.notes = 取值("notes", sku.notes or "").strip() or None


def 获取SKU返回地址(product_id):
    回跳地址 = request.form.get("next") or request.args.get("next")
    if 回跳地址:
        return 回跳地址
    return url_for("product_edit", product_id=product_id, tab="sku")


def SKU导入导出字段():
    return [
        "SKU编号",
        "SKU名称",
        "型号",
        "规格",
        "颜色",
        "尺寸",
        "材质",
        "单位",
        "包装规格",
        "MOQ",
        "单价",
        "EXW价",
        "FOB价",
        "CIF价",
        "DDP价",
        "毛重",
        "净重",
        "重量",
        "交期",
        "币种",
        "库存状态",
        "是否可样品",
        "是否支持定制",
        "备注",
        "创建时间",
        "更新时间",
    ]


def 读取SKU筛选条件(source):
    return {
        "keyword": source.get("sku_keyword", "", type=str).strip(),
        "color": source.get("sku_color", "", type=str).strip(),
        "size": source.get("sku_size", "", type=str).strip(),
        "stock_status": source.get("sku_stock_status", "", type=str).strip(),
        "sample_available": source.get("sku_sample_available", "", type=str).strip(),
        "customization_supported": source.get("sku_customization_supported", "", type=str).strip(),
    }


def 查询SKU列表(product_id, source):
    filters = 读取SKU筛选条件(source)
    query = ProductSKU.query.filter_by(product_id=product_id)
    if filters["keyword"]:
        like = f"%{filters['keyword']}%"
        query = query.filter(or_(ProductSKU.sku_name.ilike(like), ProductSKU.model.ilike(like)))
    if filters["color"]:
        query = query.filter(ProductSKU.color == filters["color"])
    if filters["size"]:
        query = query.filter(ProductSKU.size == filters["size"])
    if filters["stock_status"]:
        query = query.filter(ProductSKU.stock_status == filters["stock_status"])
    if filters["sample_available"] in {"是", "否"}:
        query = query.filter(ProductSKU.sample_available.is_(filters["sample_available"] == "是"))
    if filters["customization_supported"] in {"是", "否"}:
        query = query.filter(ProductSKU.customization_supported.is_(filters["customization_supported"] == "是"))
    return query.order_by(ProductSKU.id.asc())


def SKU筛选选项(product_id):
    全部SKU = ProductSKU.query.filter_by(product_id=product_id).all()
    return {
        "colors": sorted({item.color for item in 全部SKU if item.color}),
        "sizes": sorted({item.size for item in 全部SKU if item.size}),
        "stock_statuses": sorted({item.stock_status for item in 全部SKU if item.stock_status}),
    }


def 生成价格展示文案(product):
    片段 = []
    货币 = (product.currency or "USD").strip() or "USD"
    if product.fob_price is not None:
        片段.append(f"FOB {product.fob_price:.2f} {货币}")
    if product.cif_price is not None:
        片段.append(f"CIF {product.cif_price:.2f} {货币}")
    if product.ddp_price is not None:
        片段.append(f"DDP {product.ddp_price:.2f} {货币}")
    if product.exw_price is not None:
        片段.append(f"出厂价 {product.exw_price:.2f} {货币}")
    return " / ".join(片段) if 片段 else None


def 构建导出数据(export_key):
    导出映射 = {
        "enterprises": ("企业总表", 导出企业总表),
        "products": ("产品总表", 导出产品总表),
    }
    if export_key not in 导出映射:
        raise ValueError("unknown export key")
    文件名, 生成函数 = 导出映射[export_key]
    表头, 行数据 = 生成函数()
    return 文件名, 表头, 行数据


def 导出企业总表():
    表头 = [
        "企业编号",
        "行业分类",
        "企业全称",
        "英文名称",
        "统一社会信用代码",
        "省市",
        "企业性质",
        "主营业务",
        "核心产品",
        "年销售额",
        "年出口额",
        "是否有出口经验",
        "出口国家",
        "厂房面积",
        "员工数量",
        "产线数量",
        "产能利用率",
        "目标市场",
        "目标客户类型",
        "可接受合作模式",
        "资料完整度",
        "风险备注",
        "最近更新时间",
    ]
    rows = []
    for item in Enterprise.query.order_by(Enterprise.updated_at.desc()).all():
        ext = item.enterprise_extra_fields or {}
        企业性质 = []
        if item.is_manufacturer:
            企业性质.append("制造商")
        if item.is_trader:
            企业性质.append("贸易商")
        if item.is_brand_owner:
            企业性质.append("品牌商")
        if item.is_oem_odm:
            企业性质.append("OEM/ODM工厂")
        if item.is_service_provider:
            企业性质.append("服务商")
        rows.append([
            item.enterprise_code,
            item.industry_category,
            item.company_name,
            item.english_name,
            item.unified_social_credit_code,
            f"{item.province or '-'} / {item.city or '-'}",
            "、".join(企业性质) if 企业性质 else (item.company_type or ""),
            item.main_business,
            item.main_products,
            ext.get("annual_sales") or (float(item.annual_revenue) if item.annual_revenue is not None else ""),
            ext.get("annual_exports") or (float(item.export_revenue) if item.export_revenue is not None else ""),
            "是" if item.has_foreign_trade_experience else "否",
            item.export_countries,
            item.factory_area,
            item.employee_count,
            ext.get("production_line_count"),
            ext.get("capacity_utilization"),
            item.target_markets,
            ext.get("target_customer_types"),
            ext.get("acceptable_cooperation_modes"),
            ext.get("material_completeness"),
            item.risk_notes,
            item.updated_at.strftime("%Y-%m-%d %H:%M:%S") if item.updated_at else "",
        ])
    return 表头, rows


def 导出产品总表():
    表头 = [
        "产品编号",
        "产品名称",
        "产品英文名称",
        "所属企业编号",
        "所属企业名称",
        "行业编号",
        "行业名称",
        "产品品类",
        "产品类型",
        "HS编码",
        "品牌",
        "型号",
        "是否适合出口",
        "推荐等级",
        "目标市场",
        "合作模式",
        "产品状态",
        "上架状态",
        "MOQ",
        "交期",
        "价格展示",
        "币种",
        "样品政策",
        "是否支持定制",
        "认证情况",
        "产品认证",
        "检测报告状态",
        "质量报告状态",
        "目标市场准入文件",
        "证书有效期状态",
        "核心卖点",
        "风险提示",
        "更新时间",
    ]
    enterprise_map = {e.id: e for e in Enterprise.query.all()}
    rows = []
    for item in Product.query.order_by(Product.updated_at.desc()).all():
        enterprise = enterprise_map.get(item.enterprise_id)
        extra = item.product_extra_fields or {}
        rows.append([
            item.product_code,
            item.product_name_cn,
            item.product_name_en,
            enterprise.enterprise_code if enterprise else "",
            enterprise.company_name if enterprise else "",
            item.industry_code or (enterprise.industry_code if enterprise else ""),
            item.industry_name or (enterprise.industry_category if enterprise else ""),
            item.product_category,
            item.product_type,
            item.hs_code,
            item.brand,
            item.model,
            item.export_suitability,
            item.recommendation_level,
            item.target_market,
            "、".join(extra.get("cooperation_modes")) if isinstance(extra.get("cooperation_modes"), list) else (extra.get("cooperation_modes") or extra.get("cooperation_mode") or extra.get("trade_cooperation_mode")),
            extra.get("product_status_review"),
            item.status,
            item.moq or extra.get("trade_moq"),
            item.delivery_cycle or item.production_cycle or extra.get("trade_mass_cycle"),
            item.price_display or 生成价格展示文案(item),
            item.currency or "USD",
            item.sample_policy or extra.get("trade_sample_policy"),
            "是" if item.customization_supported else "否",
            item.certification_status or extra.get("cert_status"),
            item.certifications,
            extra.get("cert_test_report"),
            extra.get("cert_quality_report"),
            extra.get("cert_market_access"),
            extra.get("cert_validity_status"),
            item.product_selling_points,
            extra.get("risk_warning") or item.notes,
            item.updated_at.strftime("%Y-%m-%d %H:%M:%S") if item.updated_at else "",
        ])
    return 表头, rows


EXPORT_DATETIME_FORMAT = "%Y-%m-%d %H:%M:%S"


def 导出值(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime(EXPORT_DATETIME_FORMAT)
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (list, tuple, set)):
        return "、".join(str(item) for item in value if item is not None)
    return value


def 写入工作表(workbook, sheet_name, headers, rows):
    worksheet = workbook.create_sheet(title=sheet_name[:31])
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"
    for row in rows:
        worksheet.append([导出值(value) for value in row])
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 40)
    return worksheet


def 构建Excel文件(sheets):
    workbook = Workbook()
    default_sheet = workbook.active
    workbook.remove(default_sheet)
    for sheet_name, headers, rows in sheets:
        写入工作表(workbook, sheet_name, headers, rows)
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def 省市区文本(enterprise):
    return " / ".join([part for part in [enterprise.province, enterprise.city, enterprise.district] if part])


def 企业类型文本(enterprise):
    types = []
    if enterprise.is_manufacturer:
        types.append("制造商")
    if enterprise.is_trader:
        types.append("贸易商")
    if enterprise.is_brand_owner:
        types.append("品牌商")
    if enterprise.is_oem_odm:
        types.append("OEM/ODM")
    if enterprise.is_service_provider:
        types.append("服务商")
    return "、".join(types) if types else (enterprise.company_type or "")


def 合并导出列(*列组):
    """按顺序合并导出列，避免弹窗出现重复字段。"""

    合并后 = []
    已有键 = set()
    for 列列表 in 列组:
        for 键名, 标题 in 列列表:
            if 键名 in 已有键:
                continue
            合并后.append((键名, 标题))
            已有键.add(键名)
    return 合并后


def 配置字段导出列(字段组列表, 前缀=""):
    """从动态表单配置生成导出列，覆盖非必填项和行业专项属性。"""

    导出列 = []
    for 分组 in 字段组列表:
        分组标题 = 分组.get("title", "")
        for 字段 in 分组.get("fields", []):
            键名 = 字段.get("key")
            标题 = 字段.get("label")
            if not 键名 or not 标题:
                continue
            if 前缀:
                标题 = f"{前缀}-{分组标题}-{标题}" if 分组标题 else f"{前缀}-{标题}"
            导出列.append((键名, 标题))
    return 导出列


def 全部行业字段组(行业字段配置):
    """展开所有行业专项字段组，便于筛选导出跨行业数据。"""

    字段组 = []
    for 行业代码, 分组列表 in 行业字段配置.items():
        行业名称 = 行业默认名称(行业代码) or 行业代码
        for 分组 in 分组列表:
            字段组.append({
                "title": f"{行业名称}-{分组.get('title', '')}".strip("-"),
                "fields": 分组.get("fields", []),
            })
    return 字段组


EXCLUDED_EXPORT_COLUMN_LABELS = {
    "行业专项-工业机械与装备-工业机械专项-设备加工能力",
    "行业专项-工业机械与装备-工业机械专项-装配能力",
    "行业专项-工业机械与装备-工业机械专项-调试能力",
    "行业专项-工业机械与装备-工业机械专项-研发团队人数",
    "行业专项-工业机械与装备-工业机械专项-已交付设备案例",
    "行业专项-工业机械与装备-工业机械专项-海外项目经验",
    "行业专项-工业机械与装备-工业机械专项-售后备件能力",
}


def 过滤导出列(列列表):
    """移除不在企业/产品详情中展示的导出选项。"""

    return [(键名, 标题) for 键名, 标题 in 列列表 if 标题 not in EXCLUDED_EXPORT_COLUMN_LABELS]


ENTERPRISE_EXPORT_BASE_COLUMNS = [
    ("enterprise_code", "企业编号"),
    ("status", "入库状态"),
    ("project_owner", "项目负责人"),
    ("industry_code", "行业编号"),
    ("industry_category", "行业名称"),
    ("enterprise_stage", "企业当前阶段"),
    ("source_channels", "来源渠道"),
    ("company_name", "企业名称"),
    ("english_name", "英文名称"),
    ("unified_social_credit_code", "统一社会信用代码"),
    ("industry", "行业"),
    ("company_type", "企业类型"),
    ("region", "省市区"),
    ("province", "省份"),
    ("city", "城市"),
    ("district", "区县"),
    ("contact", "联系人"),
    ("phone", "联系电话"),
    ("email", "邮箱"),
    ("main_business", "主营业务"),
    ("main_products", "核心产品"),
    ("export_experience", "出口经验"),
    ("export_countries", "出口国家"),
    ("target_markets", "目标市场"),
    ("annual_capacity", "年产能"),
    ("employee_count", "员工人数"),
    ("factory_area", "工厂面积"),
    ("main_equipment", "主要设备"),
    ("annual_revenue", "年营业额"),
    ("export_revenue", "年出口额"),
    ("service_needs", "服务需求"),
    ("risk_notes", "风险备注"),
    ("material_completeness", "资料完整度"),
    ("created_at", "创建时间"),
    ("updated_at", "更新时间"),
]

ENTERPRISE_EXPORT_COLUMNS = 过滤导出列(
    合并导出列(
        ENTERPRISE_EXPORT_BASE_COLUMNS,
        配置字段导出列(COMMON_ENTERPRISE_FIELD_GROUPS),
        配置字段导出列(全部行业字段组(INDUSTRY_EXTRA_FIELD_CONFIG), "行业专项"),
    )
)


PRODUCT_EXPORT_BASE_COLUMNS = [
    ("product_code", "产品编号"),
    ("product_name_cn", "产品名称"),
    ("product_name_en", "英文名称"),
    ("enterprise", "所属企业"),
    ("enterprise_code", "所属企业编号"),
    ("industry_code", "行业编号"),
    ("industry_name", "行业名称"),
    ("product_category", "产品品类"),
    ("product_type", "产品类型"),
    ("brand", "品牌"),
    ("model", "型号"),
    ("hs_code", "HS 编码"),
    ("function_description", "产品用途"),
    ("application_scenario", "应用场景"),
    ("target_market", "目标市场"),
    ("export_suitability", "是否适合出口"),
    ("recommendation_level", "推荐等级"),
    ("existing_sales_countries", "已销售国家"),
    ("certifications", "产品认证"),
    ("certification_status", "认证情况"),
    ("product_selling_points", "核心卖点"),
    ("price_range", "价格区间"),
    ("exw_price", "出厂价"),
    ("fob_price", "FOB价"),
    ("cif_price", "CIF价"),
    ("ddp_price", "DDP参考价"),
    ("currency", "币种"),
    ("price_display", "价格展示"),
    ("moq", "MOQ"),
    ("delivery_cycle", "交付周期"),
    ("production_cycle", "生产周期"),
    ("capacity_cycle_days", "产能-周期（天）"),
    ("capacity_qualified_pieces", "产能-实际完工合格件数（件）"),
    ("sample_policy", "样品政策"),
    ("customization_supported", "是否支持定制"),
    ("notes", "备注"),
    ("main_image", "产品主图"),
    ("status", "上架状态"),
    ("created_at", "创建时间"),
    ("updated_at", "更新时间"),
]

PRODUCT_EXPORT_COLUMNS = 过滤导出列(
    合并导出列(
        PRODUCT_EXPORT_BASE_COLUMNS,
        配置字段导出列(COMMON_PRODUCT_FIELD_GROUPS),
        配置字段导出列(全部行业字段组(INDUSTRY_PRODUCT_EXTRA_FIELD_CONFIG), "行业专项"),
    )
)


def 导出值文本(value):
    """统一处理导出值，列表/字典转为便于阅读的文本。"""

    if value is None:
        return ""
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, (list, tuple, set)):
        文本列表 = []
        for item in value:
            if isinstance(item, dict):
                文本列表.append("；".join(f"{k}:{v}" for k, v in item.items() if v not in (None, "", [])))
            else:
                文本列表.append(str(item))
        return "、".join(item for item in 文本列表 if item)
    if isinstance(value, dict):
        return "；".join(f"{k}:{v}" for k, v in value.items() if v not in (None, "", []))
    return value

def 解析导出列(source, defaults):
    allowed = {key for key, _ in defaults}
    selected = [key for key in source.getlist("columns") if key in allowed]
    if not selected:
        selected = [key for key, _ in defaults]
    labels = dict(defaults)
    return selected, [labels[key] for key in selected]


def 企业主联系人映射(enterprise_ids):
    mapping = {}
    if not enterprise_ids:
        return mapping
    contacts = Contact.query.filter(Contact.enterprise_id.in_(enterprise_ids)).order_by(Contact.is_primary_contact.desc(), Contact.id.asc()).all()
    for contact in contacts:
        mapping.setdefault(contact.enterprise_id, contact)
    return mapping



def 企业导出行(enterprise, contact, completeness, columns):
    ext = 清理企业扩展字段(enterprise.enterprise_extra_fields or {})
    values = {
        "enterprise_code": enterprise.enterprise_code,
        "status": enterprise.status,
        "project_owner": enterprise.project_owner,
        "company_name": enterprise.company_name,
        "company_full_name": ext.get("company_full_name") or enterprise.company_name,
        "english_name": enterprise.english_name or ext.get("english_name"),
        "unified_social_credit_code": enterprise.unified_social_credit_code or ext.get("unified_social_credit_code"),
        "founded_date": enterprise.founded_date or ext.get("founded_date"),
        "registered_capital": enterprise.registered_capital or ext.get("registered_capital"),
        "registered_address": enterprise.registered_address or ext.get("registered_address"),
        "business_address": enterprise.business_address or ext.get("business_address"),
        "province": enterprise.province or ext.get("province"),
        "city": enterprise.city or ext.get("city"),
        "district": enterprise.district,
        "industry_code": enterprise.industry_code,
        "industry_category": enterprise.industry_category or 行业默认名称(enterprise.industry_code),
        "industry": enterprise.industry_category or ext.get("industry") or 行业默认名称(enterprise.industry_code),
        "primary_industry": ext.get("industry") or enterprise.industry_category or 行业默认名称(enterprise.industry_code),
        "core_products": ext.get("core_products") or enterprise.main_products,
        "sub_industry": enterprise.sub_industry,
        "company_type": 企业类型文本(enterprise),
        "region": 省市区文本(enterprise),
        "contact": contact.name if contact else ext.get("primary_contact_name", ""),
        "primary_contact_name": ext.get("primary_contact_name") or (contact.name if contact else ""),
        "primary_contact_position": ext.get("primary_contact_position") or (contact.position if contact else ""),
        "primary_contact_mobile": ext.get("primary_contact_mobile") or (contact.mobile if contact else ""),
        "contact_phone": ext.get("contact_phone") or (contact.phone if contact else ""),
        "contact_email": ext.get("contact_email") or (contact.email if contact else ""),
        "phone": (contact.mobile or contact.phone) if contact else ext.get("primary_contact_mobile") or ext.get("contact_phone"),
        "email": contact.email if contact else ext.get("contact_email"),
        "main_business": enterprise.main_business or ext.get("main_business_direction") or ext.get("enterprise_description"),
        "main_products": enterprise.main_products,
        "export_experience": "是" if enterprise.has_foreign_trade_experience else (ext.get("has_export_experience") or "否"),
        "has_export_experience": ext.get("has_export_experience") or ("是" if enterprise.has_foreign_trade_experience else "否"),
        "export_countries": enterprise.export_countries,
        "target_markets": enterprise.target_markets or ext.get("target_countries"),
        "target_countries": ext.get("target_countries") or enterprise.target_markets,
        "annual_capacity": enterprise.annual_capacity,
        "employee_count": enterprise.employee_count,
        "employee_count_range": ext.get("employee_count_range") or (估算人员规模(enterprise.employee_count) if enterprise.employee_count else ""),
        "factory_area": enterprise.factory_area or ext.get("factory_area"),
        "main_equipment": enterprise.main_equipment,
        "annual_revenue": enterprise.annual_revenue,
        "export_revenue": enterprise.export_revenue,
        "service_needs": enterprise.service_needs,
        "risk_notes": enterprise.risk_notes,
        "material_completeness": completeness,
        "created_at": enterprise.created_at,
        "updated_at": enterprise.updated_at,
    }
    values.update({key: value for key, value in ext.items() if key not in values or values.get(key) in (None, "", [])})
    return [导出值文本(values.get(column, "")) for column in columns]


def 构建企业查询(source):
    keyword = source.get("keyword", "", type=str).strip()
    industry = source.get("industry", "", type=str).strip()
    query = Enterprise.query
    if keyword:
        query = query.filter(or_(Enterprise.company_name.ilike(f"%{keyword}%"), Enterprise.english_name.ilike(f"%{keyword}%"), Enterprise.enterprise_code.ilike(f"%{keyword}%")))
    if industry:
        query = query.filter(Enterprise.industry_code == industry)
    return query


def 构建企业总表Sheet(query, columns):
    enterprises = query.order_by(Enterprise.updated_at.desc()).all()
    ids = [item.id for item in enterprises]
    contact_map = 企业主联系人映射(ids)
    rows = []
    for enterprise in enterprises:
        completeness = 导出企业资料完整度标签(enterprise)
        rows.append(企业导出行(enterprise, contact_map.get(enterprise.id), completeness, columns))
    return rows


def 构建产品查询(source, default_status="active"):
    query = Product.query.join(Enterprise, Product.enterprise_id == Enterprise.id)
    q_enterprise = source.get("enterprise", "", type=str).strip()
    q_keyword = source.get("keyword", "", type=str).strip()
    q_category = source.get("category", "", type=str).strip()
    q_product_type = source.get("product_type", "", type=str).strip()
    q_export_suitability = source.get("export_suitability", "", type=str).strip()
    q_recommendation_level = source.get("recommendation_level", "", type=str).strip()
    q_certification = source.get("certification_status", "", type=str).strip()
    q_target_market = source.get("target_market", "", type=str).strip()
    q_industry = source.get("industry", "", type=str).strip()
    q_status = source.get("status", "", type=str).strip() or default_status
    if q_enterprise:
        query = query.filter(Product.enterprise_id == int(q_enterprise)) if q_enterprise.isdigit() else query.filter(Enterprise.company_name.ilike(f"%{q_enterprise}%"))
    if q_keyword:
        like = f"%{q_keyword}%"
        query = query.filter(or_(Product.product_name_cn.ilike(like), Product.product_name_en.ilike(like), Product.model.ilike(like), Product.product_code.ilike(like), Enterprise.company_name.ilike(like), Enterprise.english_name.ilike(like), Enterprise.enterprise_code.ilike(like)))
    if q_category:
        query = query.filter(Product.product_category.ilike(f"%{q_category}%"))
    if q_product_type:
        query = query.filter(Product.product_type == q_product_type)
    if q_export_suitability:
        query = query.filter(Product.export_suitability == q_export_suitability)
    if q_recommendation_level:
        query = query.filter(Product.recommendation_level == q_recommendation_level)
    if q_certification:
        query = query.filter(Product.certification_status == q_certification)
    if q_target_market:
        query = query.filter(Product.target_market.ilike(f"%{q_target_market}%"))
    if q_industry:
        query = query.filter(Product.industry_code == q_industry)
    if q_status in {"active", "inactive"}:
        query = query.filter(Product.status == q_status)
    return query


def 产品价格区间(product):
    prices = [product.exw_price, product.fob_price, product.cif_price, product.ddp_price]
    numbers = [float(price) for price in prices if price is not None]
    if product.price_display:
        return product.price_display
    if not numbers:
        return ""
    currency = product.currency or "USD"
    return f"{min(numbers):.2f}-{max(numbers):.2f} {currency}" if len(numbers) > 1 else f"{numbers[0]:.2f} {currency}"


def 产品导出行(product, columns):
    enterprise = product.enterprise
    extra = 兼容产品基础信息字段(product, product.product_extra_fields or {})
    values = {
        "product_code": product.product_code,
        "product_name_cn": product.product_name_cn,
        "product_name_en": product.product_name_en,
        "enterprise": enterprise.company_name if enterprise else "",
        "enterprise_code": enterprise.enterprise_code if enterprise else "",
        "industry_code": product.industry_code,
        "industry_name": product.industry_name,
        "product_category": product.product_category,
        "product_type": product.product_type,
        "brand": product.brand,
        "model": product.model,
        "hs_code": product.hs_code,
        "function_description": product.function_description or extra.get("desc_usage"),
        "application_scenario": product.application_scenario or extra.get("desc_scenarios"),
        "target_market": product.target_market or extra.get("target_market_tags"),
        "export_suitability": product.export_suitability,
        "recommendation_level": product.recommendation_level,
        "existing_sales_countries": product.existing_sales_countries,
        "certifications": product.certifications or extra.get("cert_product"),
        "certification_status": product.certification_status or product.certifications or extra.get("cert_status"),
        "product_selling_points": product.product_selling_points or extra.get("desc_core_selling_points"),
        "price_range": 产品价格区间(product),
        "exw_price": product.exw_price or extra.get("price_exw"),
        "fob_price": product.fob_price or extra.get("price_fob"),
        "cif_price": product.cif_price or extra.get("price_cif"),
        "ddp_price": product.ddp_price or extra.get("price_ddp"),
        "currency": product.currency,
        "price_display": product.price_display,
        "moq": product.moq or extra.get("trade_moq"),
        "delivery_cycle": product.delivery_cycle or product.production_cycle or extra.get("trade_mass_cycle"),
        "production_cycle": product.production_cycle or extra.get("trade_mass_cycle"),
        "capacity_cycle_days": product.capacity_cycle_days,
        "capacity_qualified_pieces": product.capacity_qualified_pieces,
        "sample_policy": product.sample_policy or extra.get("trade_sample_policy"),
        "customization_supported": product.customization_supported if product.customization_supported else extra.get("support_customization"),
        "notes": product.notes,
        "main_image": product.main_image,
        "status": product.status,
        "created_at": product.created_at,
        "updated_at": product.updated_at,
        # 兼容产品扩展配置中的身份字段命名。
        "identity_name_cn": product.product_name_cn,
        "identity_name_en": product.product_name_en,
        "identity_model": product.model,
        "identity_sku": product.product_code,
        "identity_hs_code": product.hs_code,
        "desc_usage": product.function_description or extra.get("desc_usage"),
        "desc_scenarios": product.application_scenario or extra.get("desc_scenarios"),
        "desc_core_selling_points": product.product_selling_points or extra.get("desc_core_selling_points"),
        "price_exw": product.exw_price or extra.get("price_exw"),
        "price_fob": product.fob_price or extra.get("price_fob"),
        "price_cif": product.cif_price or extra.get("price_cif"),
        "price_ddp": product.ddp_price or extra.get("price_ddp"),
        "trade_moq": product.moq or extra.get("trade_moq"),
        "trade_sample_policy": product.sample_policy or extra.get("trade_sample_policy"),
        "trade_mass_cycle": product.delivery_cycle or product.production_cycle or extra.get("trade_mass_cycle"),
        "cert_status": product.certification_status or extra.get("cert_status"),
        "cert_product": product.certifications or extra.get("cert_product"),
        "support_customization": ("是" if product.customization_supported else "否") if product.customization_supported is not None else extra.get("support_customization"),
        "target_market_tags": product.target_market or extra.get("target_market_tags"),
        "product_status_review": extra.get("product_status_review") or product.status,
    }
    values.update({key: value for key, value in extra.items() if key not in values or values.get(key) in (None, "", [])})
    return [导出值文本(values.get(column, "")) for column in columns]


def 构建产品总表Sheet(query, columns):
    return [产品导出行(item, columns) for item in query.order_by(Product.updated_at.desc()).all()]


SKU_EXPORT_HEADERS = ["SKU 编号", "所属产品", "所属企业", "规格型号", "颜色", "尺寸", "材质", "包装规格", "单价", "币种", "MOQ", "库存/产能信息", "备注"]


def 构建SKU明细Sheet(product_id=None):
    if not hasattr(ProductSKU, "query"):
        return []
    query = ProductSKU.query.join(Product, ProductSKU.product_id == Product.id).join(Enterprise, Product.enterprise_id == Enterprise.id)
    if product_id:
        query = query.filter(ProductSKU.product_id == product_id)
    rows = []
    for sku in query.order_by(Product.product_code.asc(), ProductSKU.id.asc()).all():
        product = sku.product
        enterprise = product.enterprise if product else None
        rows.append([sku.sku_code, product.product_name_cn if product else "", enterprise.company_name if enterprise else "", sku.model or sku.specification or sku.sku_name, sku.color, sku.size, sku.material, sku.package_spec, sku.price, sku.currency, sku.moq, sku.stock_status, sku.notes])
    return rows


ATTACHMENT_EXPORT_HEADERS = ["附件名称", "原始文件名", "附件类型", "关联对象类型", "关联企业", "关联产品", "文件大小", "上传人", "上传时间", "文件路径/下载链接", "文件状态", "备注"]


def 附件清单行(document):
    enterprise = Enterprise.query.get(document.enterprise_id) if document.enterprise_id else None
    product = Product.query.get(document.product_id) if document.product_id else None
    relative_path = document.file_path or ""
    file_path = BASE_DIR / relative_path if relative_path else None
    exists = file_path.exists() if file_path else False
    size = file_path.stat().st_size if exists else ""
    link = f"/documents/{document.id}/download" if document.id else ""
    related_type = "产品附件" if document.product_id else "企业附件" if document.enterprise_id else "其他附件"
    return [document.document_name, document.original_filename, document.document_type, related_type, enterprise.company_name if enterprise else "", product.product_name_cn if product else "", size, document.uploaded_by, document.uploaded_at, link, "正常" if exists else "文件缺失", document.notes]


def 构建附件清单Sheet(enterprise_id=None, product_id=None, include_enterprise_products=False):
    query = Document.query
    if product_id:
        query = query.filter(Document.product_id == product_id)
    elif enterprise_id and include_enterprise_products:
        product_ids = [p.id for p in Product.query.filter_by(enterprise_id=enterprise_id).all()]
        query = query.filter(or_(Document.enterprise_id == enterprise_id, Document.product_id.in_(product_ids) if product_ids else False))
    elif enterprise_id:
        query = query.filter(Document.enterprise_id == enterprise_id)
    return [附件清单行(document) for document in query.order_by(Document.uploaded_at.desc()).all()]


IMPORT_ERROR_HEADERS = ["导入批次号", "导入文件名", "数据行号", "错误字段", "错误原因", "原始值", "建议修正方式", "导入时间", "操作人"]
SUPPLEMENT_LIST_HEADERS = ["导入批次号", "导入文件名", "数据行号", "企业名称", "产品名称", "联系方式", "待补充资料", "给企业的补充说明", "生成时间", "操作人"]


def 构建导入错误报告Sheet(batch_id):
    batch = ImportBatch.query.get(batch_id)
    if not batch:
        return None, []
    rows = []
    for error in batch.errors.order_by(ImportError.row_number.asc(), ImportError.id.asc()).all():
        rows.append([batch.batch_no, batch.file_name, error.row_number, error.field_name, error.error_reason, error.raw_value, error.suggestion, error.created_at, batch.operator])
    return batch, rows



def 构建待补充资料清单Sheet(batch_id):
    batch = ImportBatch.query.get(batch_id)
    if not batch:
        return None, []
    rows = []
    for item in batch.supplement_items.order_by(ImportSupplementItem.row_number.asc(), ImportSupplementItem.id.asc()).all():
        rows.append([
            batch.batch_no,
            batch.file_name,
            item.row_number,
            item.enterprise_name,
            item.product_name,
            item.contact_info,
            item.missing_items,
            item.suggestion,
            item.created_at,
            batch.operator,
        ])
    return batch, rows

def 填充项目字段(project, form):
    project.enterprise_id = form.get("enterprise_id", type=int)
    project.product_id = form.get("product_id", type=int)
    project.foreign_client_id = form.get("foreign_client_id", type=int)
    project.demand_id = form.get("demand_id", type=int)
    project.current_stage = form.get("current_stage", "").strip() or "待补充资料"
    project.first_contact_date = 读取日期(form.get("first_contact_date"))
    project.material_sent_date = 读取日期(form.get("material_sent_date"))
    project.sample_status = form.get("sample_status", "").strip() or None
    project.quotation_status = form.get("quotation_status", "").strip() or None
    project.negotiation_status = form.get("negotiation_status", "").strip() or None
    project.contract_status = form.get("contract_status", "").strip() or None
    project.deal_amount = 读取金额(form.get("deal_amount"))
    project.next_action = form.get("next_action", "").strip() or None
    project.project_owner = form.get("project_owner", "").strip() or None
    project.notes = form.get("notes", "").strip() or None


def 统一关键词集合(text):
    if not text:
        return set()
    return {item.lower() for item in re.split(r"[\s,，;；/|]+", text) if item.strip()}


def 文本匹配关键词(tokens, text):
    值 = (text or "").lower()
    return any(token in 值 for token in tokens)


def 读取生产周期天数(text):
    if not text:
        return None
    match = re.search(r"(\d+)", text)
    return int(match.group(1)) if match else None


def 计算推荐状态(score):
    if score >= 80:
        return "高度匹配"
    if score >= 60:
        return "可进一步沟通"
    return "谨慎推荐"


def 读取布尔(form, 字段名):
    return str(form.get(字段名, "")).strip().lower() in {"1", "true", "yes", "y", "是", "已", "on"}


def 读取日期(值):
    if not 值:
        return None
    try:
        return datetime.strptime(值, "%Y-%m-%d").date()
    except ValueError:
        return None


def 读取整数(值):
    if not 值:
        return None
    try:
        return int(值)
    except ValueError:
        return None


def 读取金额(值):
    if not 值:
        return None
    try:
        return Decimal(值)
    except (InvalidOperation, ValueError):
        return None


def 读取布尔文本(值):
    return str(值 or "").strip().lower() in {"1", "true", "yes", "y", "是", "已", "on"}


def csv_safe(value):
    text = "" if value is None else str(value)
    text = text.replace('"', '""')
    return f'"{text}"'


def 单元格文本(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def 读取导入表格(file_storage):
    文件名 = (file_storage.filename or "").lower()
    文件内容 = file_storage.stream.read()
    if 文件名.endswith(".xlsx"):
        workbook = load_workbook(BytesIO(文件内容), read_only=True, data_only=True)
        sheet = workbook.active
        return [[单元格文本(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    if 文件名.endswith(".csv"):
        内容 = 文件内容.decode("utf-8-sig")
        return [[单元格文本(cell) for cell in row] for row in csv.reader(内容.splitlines())]
    raise ValueError("仅支持 .xlsx 或 .csv 文件")


def 企业导入字段提示():
    return [
        ("企业编号", "对应基本信息-企业编号（必填，不可与库内或导入文件内其他企业重复）"),
        ("行业大类", "对应基本信息-行业大类"),
        ("企业全称", "对应基本信息-企业全称（必填）"),
        ("企业类型", "对应工商信息-企业类型"),
        ("统一社会信用代码", "对应工商信息-统一社会信用代码"),
        ("法定代表人", "对应工商信息-法定代表人"),
        ("注册资本", "对应工商信息-注册资本"),
        ("成立日期", "对应工商信息-成立日期，格式建议为 YYYY-MM-DD"),
        ("营业期限", "对应工商信息-营业期限起始日期，格式建议为 YYYY-MM-DD"),
        ("注册地址", "对应工商信息-注册地址"),
        ("联系人", "对应联系信息-主联系人姓名，可为空"),
        ("联系电话", "对应联系信息-电话/手机，可为空"),
    ]


def 产品导入字段提示():
    return [
        ("产品编号", "必填；用于更新匹配或新增产品，不再自动生成"),
        ("产品名称", "对应产品概览-基础信息-产品名称，兼容旧列名“产品中文名”；可为空"),
        ("所属企业名称", "必填；按企业名称匹配"),
        ("产品品类", "对应产品概览-基础信息-产品品类，兼容旧列名“产品类别”"),
        ("产品类型", "对应产品概览-基础信息-产品类型"),
        ("产能-周期（天）", "对应产品概览-基础信息-产能-周期（天）"),
        ("产能-实际完工合格件数（件）", "对应产品概览-基础信息-产能-实际完工合格件数（件）"),
        ("MOQ", "对应产品概览-交易摘要-MOQ"),
        ("交期", "对应产品概览-交易摘要-交期，兼容旧列名“批量生产周期”"),
        ("均价", "对应产品概览-交易摘要-价格展示（兼容列名“价格展示”）"),
    ]


企业导入列名别名 = {
    "行业分类": "行业大类",
    "industry_code": "行业大类",
    "industry_name": "行业大类",
}


产品导入列名别名 = {
    "product_code": "产品编号",
    "product_name_cn": "产品名称",
    "product_name_en": "产品英文名称",
    "enterprise_code": "所属企业编号",
    "enterprise_name": "所属企业名称",
    "company_name": "所属企业名称",
    "industry_code": "行业编号",
    "industry_name": "行业名称",
    "product_category": "产品品类",
    "category": "产品品类",
    "product_type": "产品类型",
    "hs_code": "HS编码",
    "brand": "品牌",
    "model": "型号",
    "target_market": "目标市场",
    "export_suitability": "是否适合出口",
    "recommendation_level": "推荐等级",
    "certification_status": "认证情况",
    "status": "上架状态",
    "moq": "MOQ",
    "delivery_cycle": "交期",
    "production_cycle": "交期",
    "capacity_cycle_days": "产能-周期（天）",
    "capacity_qualified_pieces": "产能-实际完工合格件数（件）",
    "price_display": "价格展示",
    "currency": "币种",
    "sample_policy": "样品政策",
    "customization_supported": "是否支持定制",
    "certifications": "产品认证",
    "product_selling_points": "核心卖点",
    "notes": "备注",
}


def 规范导入表头(header, 列名别名=None):
    列名别名 = 列名别名 or {}
    idx = {}
    for i, raw_name in enumerate(header):
        name = 单元格文本(raw_name).strip()
        if not name:
            continue
        idx.setdefault(name, i)
        idx.setdefault(列名别名.get(name, name), i)
    return idx


def 读取行字段(row, idx, 字段名):
    if 字段名 not in idx:
        return ""
    列索引 = idx[字段名]
    if 列索引 >= len(row):
        return ""
    return 单元格文本(row[列索引])


def 取首个存在字段值(row, idx, *字段名):
    for 字段 in 字段名:
        if 字段 in idx:
            return 读取行字段(row, idx, 字段)
    return ""


def 规范多值文本(value):
    tokens = [item.strip() for item in re.split(r"[，,]+", 单元格文本(value)) if item and item.strip()]
    return "、".join(tokens) if tokens else None


def 解析省市(value):
    text = 单元格文本(value)
    if not text:
        return None, None
    if "/" in text:
        parts = [p.strip() for p in text.split("/", 1)]
    elif "／" in text:
        parts = [p.strip() for p in text.split("／", 1)]
    else:
        return text, None
    return parts[0] or None, parts[1] or None


def 读取首个日期(值):
    text = 单元格文本(值)
    if not text:
        return None
    标准文本 = text.replace("/", "-").replace(".", "-")
    日期 = 读取日期(标准文本)
    if 日期:
        return 日期
    match = re.search(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", text)
    if not match:
        return None
    年, 月, 日 = re.split(r"[-/]", match.group(0))
    try:
        return date(int(年), int(月), int(日))
    except ValueError:
        return None


def 解析导入行业(value):
    text = 单元格文本(value)
    if not text:
        return None, None
    if text in INDUSTRY_MAP:
        item = INDUSTRY_MAP[text]
        return item["code"], item["name"]
    for item in INDUSTRY_OPTIONS:
        if text == item["name"] or text in {f'{item["code"]} {item["name"]}', f'{item["code"]}-{item["name"]}', f'{item["code"]}｜{item["name"]}'}:
            return item["code"], item["name"]
    return None, text


def 联系电话字段键(value):
    text = 单元格文本(value)
    digits = re.sub(r"\D", "", text)
    if len(digits) == 11 and digits.startswith("1"):
        return "primary_contact_mobile"
    return "contact_phone"


导入错误字段建议 = {
    "企业名称": "请填写营业执照上的企业全称。",
    "企业全称": "请填写营业执照上的企业全称。",
    "产品名称": "请填写产品中文名称。",
    "HS编码": "请补充准确 HS 编码，建议由企业外贸/报关负责人确认。",
    "联系方式": "请至少补充联系人姓名、手机/电话或邮箱中的一种有效联系方式。",
    "核心产品": "请填写企业核心产品，多个产品可用顿号分隔。",
    "所属企业": "请填写可匹配的所属企业名称或企业编号。",
}


def 创建导入批次(file_storage, import_type):
    批次 = ImportBatch(
        batch_no=f"IMP-{datetime.now().strftime('%Y%m%d%H%M%S')}-{int(time.time() * 1000) % 1000:03d}",
        file_name=file_storage.filename or "未命名导入文件",
        import_type=import_type,
        operator=session.get("用户") or "system",
    )
    db.session.add(批次)
    db.session.flush()
    return 批次


def 记录导入错误(batch, row_number, field_name, error_reason, raw_value="", suggestion=None):
    if not batch:
        return
    db.session.add(
        ImportError(
            batch_id=batch.id,
            row_number=row_number,
            field_name=field_name,
            error_reason=error_reason,
            raw_value=单元格文本(raw_value),
            suggestion=suggestion or 导入错误字段建议.get(field_name) or "请核对该字段后重新导入。",
        )
    )


def 记录待补充资料(batch, row_number, enterprise_name, product_name, missing_items, contact_info=""):
    if not batch or not missing_items:
        return
    清单文本 = "、".join(dict.fromkeys([item for item in missing_items if item]))
    db.session.add(
        ImportSupplementItem(
            batch_id=batch.id,
            row_number=row_number,
            enterprise_name=enterprise_name or "",
            product_name=product_name or "",
            contact_info=contact_info or "",
            missing_items=清单文本,
            suggestion=f"请企业补充：{清单文本}。",
        )
    )


def 导入行上下文(row, idx, 字段名列表):
    return "；".join(f"{字段名}={读取行字段(row, idx, 字段名)}" for 字段名 in 字段名列表 if 字段名 in idx)

def 导入企业Excel(file_storage, batch=None):
    rows = 读取导入表格(file_storage)
    if not rows:
        return 0, [{"行号": 1, "原因": "文件为空", "数据": {}}]
    header = [单元格文本(c) for c in rows[0]]
    idx = 规范导入表头(header, 企业导入列名别名)
    必填 = ["企业编号", "企业全称"]
    缺失 = [f for f in 必填 if f not in idx]
    if 缺失:
        原因 = f"缺少必填列: {', '.join(缺失)}"
        for 字段 in 缺失:
            记录导入错误(batch, 1, 字段, 原因, suggestion=导入错误字段建议.get(字段))
        return 0, [{"行号": 1, "原因": 原因, "数据": {}}]

    success = 0
    failed = []
    for row_num, row in enumerate(rows[1:], start=2):
        try:
            if not any(单元格文本(cell) for cell in row):
                continue
            enterprise_code = 读取行字段(row, idx, "企业编号")
            company_name = 读取行字段(row, idx, "企业全称")
            core_products = 读取行字段(row, idx, "核心产品")
            contact_name = 取首个存在字段值(row, idx, "联系人", "联系人姓名")
            contact_phone = 取首个存在字段值(row, idx, "联系电话", "手机", "电话", "联系方式")
            contact_info = contact_phone or contact_name or 取首个存在字段值(row, idx, "邮箱")
            行错误 = []
            if not enterprise_code:
                行错误.append(("企业编号", "企业编号缺失"))
            if not company_name:
                行错误.append(("企业全称", "企业名称缺失"))
            if 行错误:
                上下文 = 导入行上下文(row, idx, ["企业编号", "企业全称", "联系人", "联系电话", "手机", "电话", "邮箱", "联系方式"])
                for 字段, 原因 in 行错误:
                    记录导入错误(batch, row_num, 字段, 原因, 上下文, 导入错误字段建议.get(字段))
                raise ValueError("；".join(原因 for _, 原因 in 行错误))

            上下文 = 导入行上下文(row, idx, ["企业编号", "企业全称", "统一社会信用代码"])
            if enterprise_code in 已读取企业编号:
                原因 = f"企业编号与第 {已读取企业编号[enterprise_code]} 行重复，请修改后重新导入"
                记录导入错误(batch, row_num, "企业编号", 原因, 上下文, "请确保导入文件内企业编号唯一。")
                raise ValueError(原因)
            已读取企业编号[enterprise_code] = row_num

            if Enterprise.query.filter_by(enterprise_code=enterprise_code).first():
                原因 = "企业编号已存在，请修改后重新导入"
                记录导入错误(batch, row_num, "企业编号", 原因, 上下文, "请使用未入库的企业编号，或先在企业编辑页调整编号。")
                raise ValueError(原因)

            enterprise = Enterprise(enterprise_code=enterprise_code)
            db.session.add(enterprise)

            ext = dict(enterprise.enterprise_extra_fields or {})
            province = city = None
            if "省市" in idx:
                province, city = 解析省市(读取行字段(row, idx, "省市"))
            enterprise.company_name = company_name
            ext["company_full_name"] = company_name
            enterprise.english_name = 读取行字段(row, idx, "英文名称") or enterprise.english_name
            enterprise.unified_social_credit_code = 读取行字段(row, idx, "统一社会信用代码") or None if "统一社会信用代码" in idx else enterprise.unified_social_credit_code
            if "行业大类" in idx:
                行业文本 = 读取行字段(row, idx, "行业大类")
                if 行业文本:
                    行业代码, 行业名称 = 解析导入行业(行业文本)
                    enterprise.industry_code = 行业代码 or enterprise.industry_code
                    enterprise.industry_category = 行业名称
                    ext["industry"] = 行业名称
            enterprise.company_type = 读取行字段(row, idx, "企业类型") or enterprise.company_type
            enterprise.legal_representative = 读取行字段(row, idx, "法定代表人") or enterprise.legal_representative
            enterprise.registered_capital = 读取行字段(row, idx, "注册资本") or enterprise.registered_capital
            enterprise.founded_date = 读取首个日期(读取行字段(row, idx, "成立日期")) or enterprise.founded_date
            enterprise.business_term_start = 读取首个日期(读取行字段(row, idx, "营业期限")) or enterprise.business_term_start
            enterprise.registered_address = 读取行字段(row, idx, "注册地址") or enterprise.registered_address
            if province is not None:
                enterprise.province = province
            if city is not None:
                enterprise.city = city
            if contact_name:
                ext["primary_contact_name"] = contact_name
            if contact_phone:
                ext[联系电话字段键(contact_phone)] = contact_phone
                ext["import_contact_info"] = contact_phone
            if "企业性质" in idx:
                企业性质文本 = 单元格文本(row[idx["企业性质"]])
                enterprise.company_type = 企业性质文本 or None
                tokens = set(re.split(r"[、,，;/|]+", 企业性质文本))
                enterprise.is_manufacturer = "制造商" in tokens
                enterprise.is_trader = "贸易商" in tokens
                enterprise.is_brand_owner = "品牌商" in tokens
                enterprise.is_oem_odm = "OEM/ODM工厂" in tokens or "OEM" in tokens or "ODM" in tokens
                enterprise.is_service_provider = "服务商" in tokens
            if "主营业务" in idx:
                enterprise.main_business = 读取行字段(row, idx, "主营业务") or None
            if core_products:
                enterprise.main_products = core_products
                ext["core_products"] = core_products
            if contact_info:
                ext["import_contact_info"] = contact_info
            if "出口国家" in idx:
                enterprise.export_countries = 读取行字段(row, idx, "出口国家") or None
            if "目标市场" in idx:
                enterprise.target_markets = 读取行字段(row, idx, "目标市场") or None
            if "厂房面积" in idx:
                enterprise.factory_area = 读取行字段(row, idx, "厂房面积") or None
            if "员工数量" in idx:
                enterprise.employee_count = 读取整数(读取行字段(row, idx, "员工数量"))
            if "风险备注" in idx:
                enterprise.risk_notes = 读取行字段(row, idx, "风险备注") or None
            if "是否有出口经验" in idx:
                enterprise.has_foreign_trade_experience = 读取布尔文本(row[idx["是否有出口经验"]])
            if "年销售额" in idx:
                ext["annual_sales"] = 单元格文本(row[idx["年销售额"]]) or None
            if "年出口额" in idx:
                ext["annual_exports"] = 单元格文本(row[idx["年出口额"]]) or None
            if "产线数量" in idx:
                ext["production_line_count"] = 单元格文本(row[idx["产线数量"]]) or None
            if "产能利用率" in idx:
                ext["capacity_utilization"] = 单元格文本(row[idx["产能利用率"]]) or None
            if "目标客户类型" in idx:
                ext["target_customer_types"] = 单元格文本(row[idx["目标客户类型"]]) or None
            if "可接受合作模式" in idx:
                ext["acceptable_cooperation_modes"] = 单元格文本(row[idx["可接受合作模式"]]) or None
            if "资料完整度" in idx:
                ext["material_completeness"] = 单元格文本(row[idx["资料完整度"]]) or None
            enterprise.enterprise_extra_fields = 清理企业扩展字段(ext)
            success += 1
        except Exception as exc:
            if batch and not batch.errors.filter_by(row_number=row_num).first():
                记录导入错误(batch, row_num, "整行", str(exc), 导入行上下文(row, idx, ["企业编号", "企业全称", "核心产品"]))
            failed.append({"行号": row_num, "原因": str(exc), "数据": {"企业编号": row[idx["企业编号"]] if "企业编号" in idx else "", "企业全称": row[idx["企业全称"]] if "企业全称" in idx else ""}})
    db.session.commit()
    return success, failed


def 导入产品Excel(file_storage, batch=None):
    rows = 读取导入表格(file_storage)
    if not rows:
        return 0, [{"行号": 1, "原因": "文件为空", "数据": {}}]
    header = [单元格文本(c) for c in rows[0]]
    idx = 规范导入表头(header, 产品导入列名别名)
    缺失表头 = [字段 for 字段 in ["产品编号", "所属企业名称"] if 字段 not in idx]
    if 缺失表头:
        原因 = f"缺少必填列: {', '.join(缺失表头)}"
        记录导入错误(batch, 1, "、".join(缺失表头), 原因)
        return 0, [{"行号": 1, "原因": 原因, "数据": {}}]

    enterprise_list = Enterprise.query.all()
    enterprise_name_map = {e.company_name: e for e in enterprise_list if e.company_name}
    success = 0
    failed = []
    for row_num, row in enumerate(rows[1:], start=2):
        try:
            if not any(单元格文本(cell) for cell in row):
                continue
            enterprise_name = 读取行字段(row, idx, "所属企业名称")
            product_code = 读取行字段(row, idx, "产品编号")
            行错误 = []
            if not product_code:
                行错误.append(("产品编号", "产品编号缺失"))
            if not enterprise_name:
                行错误.append(("所属企业名称", "所属企业名称缺失"))
            if 行错误:
                上下文 = 导入行上下文(row, idx, ["产品编号", "所属企业名称", "产品名称", "产品中文名"])
                for 字段, 原因 in 行错误:
                    记录导入错误(batch, row_num, 字段, 原因, 上下文, 导入错误字段建议.get(字段))
                raise ValueError("；".join(原因 for _, 原因 in 行错误))
            enterprise = enterprise_name_map.get(enterprise_name)
            if not enterprise:
                记录导入错误(batch, row_num, "所属企业名称", f"未找到企业：{enterprise_name}", 导入行上下文(row, idx, ["产品编号", "所属企业名称", "产品名称"]))
                raise ValueError(f"未找到企业：{enterprise_name}")
            name_cn = 取首个存在字段值(row, idx, "产品名称", "产品中文名")
            hs_code_value = 读取行字段(row, idx, "HS编码")
            product = Product.query.filter_by(product_code=product_code).first()
            if not product:
                product = Product(
                    enterprise_id=enterprise.id,
                    product_code=product_code,
                )
                db.session.add(product)

            extra = dict(product.product_extra_fields or {})
            product.enterprise_id = enterprise.id
            product.product_name_cn = name_cn or product.product_name_cn or product_code
            if "产品英文名称" in idx or "产品英文名" in idx:
                product.product_name_en = 取首个存在字段值(row, idx, "产品英文名称", "产品英文名") or None
            if "行业编号" in idx:
                product.industry_code = 读取行字段(row, idx, "行业编号") or None
            行业名称 = 取首个存在字段值(row, idx, "行业名称", "行业分类")
            if 行业名称:
                product.industry_name = 行业名称
            if not product.industry_name:
                product.industry_name = enterprise.industry_category
            if not product.industry_code:
                product.industry_code = enterprise.industry_code
            product.product_category = 取首个存在字段值(row, idx, "产品品类", "产品类别") or None
            product.product_type = 读取行字段(row, idx, "产品类型") or None if "产品类型" in idx else None
            if "HS编码" in idx:
                product.hs_code = hs_code_value or None
            if "产能-周期（天）" in idx:
                product.capacity_cycle_days = 读取整数(读取行字段(row, idx, "产能-周期（天）"))
            if "产能-实际完工合格件数（件）" in idx:
                product.capacity_qualified_pieces = 读取整数(读取行字段(row, idx, "产能-实际完工合格件数（件）"))
            product.brand = 取首个存在字段值(row, idx, "品牌", "SKU") or None
            product.model = 读取行字段(row, idx, "型号") or None if "型号" in idx else None
            product.export_suitability = 读取行字段(row, idx, "是否适合出口") or None if "是否适合出口" in idx else None
            product.recommendation_level = 读取行字段(row, idx, "推荐等级") or None if "推荐等级" in idx else None
            product.target_market = 规范多值文本(取首个存在字段值(row, idx, "目标市场")) or product.target_market
            合作模式 = 规范多值文本(取首个存在字段值(row, idx, "合作模式"))
            if 合作模式:
                extra["cooperation_modes"] = [item for item in 合作模式.split("、") if item]
            业务状态 = 读取行字段(row, idx, "产品状态") if "产品状态" in idx else ""
            if 业务状态:
                extra["product_status_review"] = 业务状态
            if "上架状态" in idx:
                上架状态原值 = 读取行字段(row, idx, "上架状态").lower()
                if 上架状态原值 in {"active", "上架", "启用", "enable", "enabled"}:
                    product.status = "active"
                elif 上架状态原值 in {"inactive", "下架", "停用", "disable", "disabled"}:
                    product.status = "inactive"
            product.product_selling_points = 读取行字段(row, idx, "核心卖点") or None if "核心卖点" in idx else None
            product.moq = 读取行字段(row, idx, "MOQ") or None if "MOQ" in idx else None
            product.delivery_cycle = 取首个存在字段值(row, idx, "交期", "批量生产周期") or product.delivery_cycle
            product.price_display = 取首个存在字段值(row, idx, "均价", "价格展示") or None
            product.currency = 读取行字段(row, idx, "币种") or (product.currency or "USD") if "币种" in idx else (product.currency or "USD")
            product.sample_policy = 读取行字段(row, idx, "样品政策") or None if "样品政策" in idx else None
            product.certification_status = 读取行字段(row, idx, "认证情况") or None if "认证情况" in idx else None
            product.certifications = 规范多值文本(取首个存在字段值(row, idx, "产品认证")) or None
            if "检测报告状态" in idx:
                extra["cert_test_report"] = 读取行字段(row, idx, "检测报告状态") or None
            if "质量报告状态" in idx:
                extra["cert_quality_report"] = 读取行字段(row, idx, "质量报告状态") or None
            if "目标市场准入文件" in idx:
                extra["cert_market_access"] = 读取行字段(row, idx, "目标市场准入文件") or None
            if "证书有效期状态" in idx:
                extra["cert_validity_status"] = 读取行字段(row, idx, "证书有效期状态") or None
            if "是否支持定制" in idx:
                product.customization_supported = 读取布尔文本(读取行字段(row, idx, "是否支持定制"))
            备注 = 读取行字段(row, idx, "备注") if "备注" in idx else ""
            if 备注:
                product.notes = 备注
            product.product_extra_fields = extra
            待补充项 = []
            if not product.main_image:
                待补充项.append("产品图片")
            if not (product.certifications or product.certification_status):
                待补充项.append("认证证书")
            if not product.moq:
                待补充项.append("MOQ")
            if not product.delivery_cycle:
                待补充项.append("交期")
            if not product.product_name_en:
                待补充项.append("英文资料")
            if not product.target_market:
                待补充项.append("目标国家")
            记录待补充资料(batch, row_num, enterprise.company_name, product.product_name_cn, 待补充项, enterprise.enterprise_extra_fields.get("import_contact_info", "") if enterprise.enterprise_extra_fields else "")
            success += 1
        except Exception as exc:
            if batch and not batch.errors.filter_by(row_number=row_num).first():
                记录导入错误(batch, row_num, "整行", str(exc), 导入行上下文(row, idx, ["产品编号", "所属企业名称", "产品名称"]))
            failed.append({"行号": row_num, "原因": str(exc), "数据": {"产品编号": 读取行字段(row, idx, "产品编号"), "所属企业名称": 读取行字段(row, idx, "所属企业名称"), "产品名称": 取首个存在字段值(row, idx, "产品名称", "产品中文名")}})
    db.session.commit()
    return success, failed


def 导入SKUExcel(product, file_storage):
    rows = 读取导入表格(file_storage)
    if not rows:
        return 0, [{"行号": 1, "原因": "文件为空", "数据": {}}]
    header = [单元格文本(c) for c in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    必填 = ["SKU名称"]
    缺失 = [f for f in 必填 if f not in idx]
    if 缺失:
        return 0, [{"行号": 1, "原因": f"缺少必填列: {', '.join(缺失)}", "数据": {}}]

    success = 0
    failed = []
    for row_num, row in enumerate(rows[1:], start=2):
        try:
            if not any(单元格文本(cell) for cell in row):
                continue
            sku_code = 单元格文本(row[idx["SKU编号"]]) if "SKU编号" in idx else ""
            sku = ProductSKU.query.filter_by(product_id=product.id, sku_code=sku_code).first() if sku_code else None
            if not sku:
                sku = ProductSKU(
                    product_id=product.id,
                    sku_code=sku_code or generate_sku_code(product),
                )
                db.session.add(sku)
            sku.sku_name = 单元格文本(row[idx["SKU名称"]]) or ""
            if not sku.sku_name:
                raise ValueError("SKU名称不能为空")
            sku.model = 单元格文本(row[idx["型号"]]) or None if "型号" in idx else sku.model
            sku.specification = 单元格文本(row[idx["规格"]]) or None if "规格" in idx else sku.specification
            sku.color = 单元格文本(row[idx["颜色"]]) or None if "颜色" in idx else sku.color
            sku.size = 单元格文本(row[idx["尺寸"]]) or None if "尺寸" in idx else sku.size
            sku.material = 单元格文本(row[idx["材质"]]) or None if "材质" in idx else sku.material
            sku.unit = 单元格文本(row[idx["单位"]]) or None if "单位" in idx else sku.unit
            sku.package_spec = 单元格文本(row[idx["包装规格"]]) or None if "包装规格" in idx else sku.package_spec
            sku.moq = 单元格文本(row[idx["MOQ"]]) or None if "MOQ" in idx else sku.moq
            if "单价" in idx:
                sku.price = 读取金额(单元格文本(row[idx["单价"]]))
            if "EXW价" in idx:
                sku.exw_price = 读取金额(单元格文本(row[idx["EXW价"]]))
            if "FOB价" in idx:
                sku.fob_price = 读取金额(单元格文本(row[idx["FOB价"]]))
                if sku.price is None:
                    sku.price = sku.fob_price
            if "CIF价" in idx:
                sku.cif_price = 读取金额(单元格文本(row[idx["CIF价"]]))
            if "DDP价" in idx:
                sku.ddp_price = 读取金额(单元格文本(row[idx["DDP价"]]))
            sku.gross_weight = 单元格文本(row[idx["毛重"]]) or None if "毛重" in idx else sku.gross_weight
            sku.net_weight = 单元格文本(row[idx["净重"]]) or None if "净重" in idx else sku.net_weight
            sku.weight = 单元格文本(row[idx["重量"]]) or None if "重量" in idx else sku.weight
            sku.delivery_cycle = 单元格文本(row[idx["交期"]]) or None if "交期" in idx else sku.delivery_cycle
            sku.currency = 单元格文本(row[idx["币种"]]) or "USD" if "币种" in idx else (sku.currency or "USD")
            sku.stock_status = 单元格文本(row[idx["库存状态"]]) or None if "库存状态" in idx else sku.stock_status
            if "是否可样品" in idx:
                sku.sample_available = 读取布尔文本(row[idx["是否可样品"]])
            if "是否支持定制" in idx:
                sku.customization_supported = 读取布尔文本(row[idx["是否支持定制"]])
            sku.notes = 单元格文本(row[idx["备注"]]) or None if "备注" in idx else sku.notes
            success += 1
        except Exception as exc:
            failed.append({"行号": row_num, "原因": str(exc), "数据": [单元格文本(c) for c in row]})
    db.session.commit()
    return success, failed


def 获取证书类型选项():
    return [
        "基础证照",
        "生产资质",
        "外贸资质",
        "质量体系认证",
        "产品认证",
        "知识产权",
        "合规文件",
        "其他",
    ]


def 计算证书状态(到期日期):
    if not 到期日期:
        return "未填写"
    剩余天数 = (到期日期 - date.today()).days
    if 剩余天数 < 0:
        return "已过期"
    if 剩余天数 <= 90:
        return "即将到期"
    return "正常"


def 构建证照展示项(资质):
    证书状态 = 计算证书状态(资质.expiry_date)
    剩余天数 = (资质.expiry_date - date.today()).days if 资质.expiry_date else None
    状态样式 = {
        "已过期": "danger",
        "即将到期": "warning",
        "正常": "success",
        "未填写": "secondary",
    }
    return {
        "记录": 资质,
        "证书状态": 证书状态,
        "剩余天数": 剩余天数,
        "状态样式": 状态样式.get(证书状态, "secondary"),
    }


def 包含非法文件名字符(名称):
    return bool(re.search(r'[\\/:*?"<>|]', 名称 or ""))


def 清洗路径片段(名称):
    return re.sub(r'[\\/:*?"<>|]+', "_", (名称 or "").strip())


def 获取文件分类目录(document_type):
    return DOCUMENT_FOLDER_MAPPING.get((document_type or "").strip(), "13_风险审核与归档确认")


def 构建企业归档目录(enterprise):
    行业编号 = 清洗路径片段(enterprise.industry_code or "I00")
    企业编号 = 清洗路径片段(enterprise.enterprise_code or "E000")
    企业名称 = 清洗路径片段(enterprise.company_name or "未命名企业")
    省市 = 清洗路径片段((enterprise.province or "") + (enterprise.city or "")) or "未知地区"
    入库日期 = (enterprise.created_at.date() if enterprise.created_at else date.today()).strftime("%Y%m%d")
    目录名 = f"{行业编号}_{企业编号}_{企业名称}_{省市}_{入库日期}"
    return Path("uploads") / "企业库" / 目录名


def 初始化企业归档目录(upload_root, enterprise):
    企业相对目录 = 构建企业归档目录(enterprise)
    企业绝对目录 = BASE_DIR / 企业相对目录
    企业绝对目录.mkdir(parents=True, exist_ok=True)
    for 子目录 in ENTERPRISE_SUB_FOLDERS:
        (企业绝对目录 / 子目录).mkdir(parents=True, exist_ok=True)
    return 企业绝对目录


def 构建标准文件名(industry_code, enterprise_code, product_code, document_type, document_name, date_text, uploaded_by, extension):
    安全文件名 = 清洗路径片段(document_name)
    安全上传人 = 清洗路径片段(uploaded_by)
    片段 = [清洗路径片段(industry_code), 清洗路径片段(enterprise_code)]
    if product_code:
        片段.append(清洗路径片段(product_code))
    片段.extend([清洗路径片段(document_type), 安全文件名, date_text, 安全上传人])
    return "_".join([片段项 for 片段项 in 片段 if 片段项]) + extension


def 生成不覆盖文件路径(目标路径):
    if not 目标路径.exists():
        return 目标路径
    时间戳 = datetime.now().strftime("%H%M%S")
    return 目标路径.with_name(f"{目标路径.stem}_{时间戳}{目标路径.suffix}")


def 保存文件并登记记录(app, enterprise, product, 上传文件, document_type, document_name, uploaded_by, notes=None, related_project_id=None, filename_override=None):
    扩展名 = Path(上传文件.filename).suffix.lower()
    if 扩展名 in BLOCKED_EXTENSIONS:
        raise ValueError("文件类型不允许上传。")
    if request.content_length and request.content_length > MAX_UPLOAD_SIZE:
        raise ValueError("单文件大小不能超过 100MB。")

    安全原始文件名 = secure_filename(上传文件.filename)
    if not 安全原始文件名:
        raise ValueError("上传文件名无效，请重命名后重试。")

    企业目录 = 初始化企业归档目录(app.config["UPLOAD_ROOT"], enterprise)
    归档目录 = 企业目录 / 获取文件分类目录(document_type)
    归档目录.mkdir(parents=True, exist_ok=True)

    日期文本 = datetime.now().strftime("%Y%m%d")
    标准文件名 = filename_override or 构建标准文件名(
        industry_code=enterprise.industry_code or "I00",
        enterprise_code=enterprise.enterprise_code or "E000",
        product_code=product.product_code if product else None,
        document_type=document_type,
        document_name=document_name,
        date_text=日期文本,
        uploaded_by=uploaded_by,
        extension=扩展名,
    )
    存储路径 = 生成不覆盖文件路径(归档目录 / 标准文件名)
    上传文件.save(存储路径)

    document = Document(
        enterprise_id=enterprise.id,
        product_id=product.id if product else None,
        related_project_id=related_project_id,
        document_type=document_type,
        document_name=document_name,
        version="V01",
        file_path=str(存储路径.relative_to(BASE_DIR)),
        original_filename=安全原始文件名,
        uploaded_by=uploaded_by,
        notes=notes,
    )
    db.session.add(document)
    db.session.flush()
    return document


def 包含关键词(文本, 关键词列表):
    内容 = (文本 or "").upper()
    return any(关键词 in 内容 for 关键词 in 关键词列表)


def 生成出海方案分析(企业, 产品列表, 资质列表, 文件列表):
    认证关键词 = ["CE", "FDA", "UL", "REACH", "ROHS", "ETL", "FCC", "ISO"]
    宣传类型 = {"PPT", "IMG", "VIDEO"}
    产品资料类型 = {"PROD", "SPEC"}
    合规资料类型 = {"CERT", "AUTH", "CONTRACT"}

    有外贸经验 = bool(企业.has_foreign_trade_experience or 企业.export_revenue or 企业.export_countries)
    有目标市场认证 = any(
        包含关键词(资质.certificate_name, 认证关键词) or 包含关键词(资质.certificate_type, 认证关键词)
        for 资质 in 资质列表
    ) or any(包含关键词(产品.certifications, 认证关键词) for 产品 in 产品列表)
    有完整产品资料 = any(
        (产品.product_name_en or "").strip() and ((产品.specification or "").strip() or (产品.function_description or "").strip())
        for 产品 in 产品列表
    ) and any((文件.document_type or "").upper() in 产品资料类型 for 文件 in 文件列表)
    有清晰报价和MOQ = any(
        (产品.moq or "").strip()
        and any([产品.exw_price is not None, 产品.fob_price is not None, 产品.cif_price is not None, 产品.ddp_price is not None])
        for 产品 in 产品列表
    )
    有稳定产能和生产周期 = bool(企业.annual_capacity) and any(
        产品.capacity_qualified_pieces is not None and (产品.capacity_cycle_days is not None or (产品.production_cycle or "").strip()) for 产品 in 产品列表
    )
    有目标市场或已出口国家 = bool(企业.target_markets or 企业.export_countries) or any(
        (产品.target_market or "").strip() or (产品.existing_sales_countries or "").strip() for 产品 in 产品列表
    )
    有合规文件和证照 = bool(资质列表) or any((文件.document_type or "").upper() in 合规资料类型 for 文件 in 文件列表)
    有宣传资料 = any((文件.document_type or "").upper() in 宣传类型 for 文件 in 文件列表)

    评分项 = [
        {"名称": "是否有外贸经验", "分值": 15, "达成": 有外贸经验},
        {"名称": "是否有目标市场认证", "分值": 20, "达成": 有目标市场认证},
        {"名称": "是否有完整产品资料", "分值": 15, "达成": 有完整产品资料},
        {"名称": "是否有清晰报价和 MOQ", "分值": 10, "达成": 有清晰报价和MOQ},
        {"名称": "是否有稳定产能和生产周期", "分值": 15, "达成": 有稳定产能和生产周期},
        {"名称": "是否有目标市场或已出口国家", "分值": 10, "达成": 有目标市场或已出口国家},
        {"名称": "是否有合规文件和证照", "分值": 10, "达成": 有合规文件和证照},
        {"名称": "是否有宣传资料/PPT/影像资料", "分值": 5, "达成": 有宣传资料},
    ]
    成熟度分数 = sum(项["分值"] for 项 in 评分项 if 项["达成"])

    标签 = []
    if 成熟度分数 >= 80 and 有目标市场认证 and 有清晰报价和MOQ:
        标签.extend(["外贸成熟型", "适合重点推荐"])
    elif 成熟度分数 >= 55:
        标签.append("产品潜力型")
    if not 有完整产品资料:
        标签.append("资料待完善型")
    if not 有目标市场认证:
        标签.append("认证缺失型")
    if not 有清晰报价和MOQ:
        标签.append("价格待确认型")
    if not 有稳定产能和生产周期:
        标签.append("产能风险型")
    if "适合重点推荐" not in 标签:
        标签.append("适合先辅导再推荐")

    优势说明 = []
    风险提示 = []
    下一步建议 = []

    if 有外贸经验 and 有目标市场或已出口国家:
        优势说明.append("具备一定外贸经验与目标市场信息，可快速进入客户匹配环节。")
        下一步建议.append("优先推荐给已有明确采购需求的外资客户")
    if 有完整产品资料:
        优势说明.append("产品基础资料较完整，可支持初步评估和商机沟通。")
    else:
        风险提示.append("产品英文资料/规格书不完整，影响外资客户快速评估。")
        下一步建议.append("先补充产品英文资料、规格书和报价单")
    if 有目标市场认证:
        优势说明.append("已具备目标市场相关认证基础，有助于缩短准入周期。")
    else:
        风险提示.append("缺少关键目标市场认证，存在准入障碍。")
        下一步建议.append("先完善目标市场认证，例如 CE/FDA/UL/REACH 等")
    if not 有稳定产能和生产周期:
        风险提示.append("产能或生产周期信息不足，交付稳定性待确认。")
        下一步建议.append("需要进一步核实产能、交期和供货稳定性")
    if 成熟度分数 >= 60:
        下一步建议.append("适合参加海外展会或线上 B2B 平台测试")
        下一步建议.append("适合做目标市场准入与竞品价格分析")
    if 有清晰报价和MOQ:
        下一步建议.append("适合先进行样品测试和小批量订单撮合")
    else:
        风险提示.append("报价或 MOQ 信息不明确，影响商务推进效率。")
    if not 有合规文件和证照:
        风险提示.append("基础合规文件不充分，当前推荐风险较高。")
        下一步建议.append("暂不建议对外推荐，需补齐基础资质或合规文件")

    if not 优势说明:
        优势说明.append("企业具备一定生产与产品基础，可通过辅导逐步进入推荐池。")
    if not 风险提示:
        风险提示.append("当前未发现显著短板，建议继续维护资料与交付稳定性。")

    下一步建议 = list(dict.fromkeys(下一步建议))
    标签 = list(dict.fromkeys(标签))

    return {
        "成熟度分数": 成熟度分数,
        "评分项": 评分项,
        "标签": 标签,
        "优势说明": 优势说明,
        "风险提示": 风险提示,
        "下一步建议": 下一步建议,
    }


def init_db(app):
    """初始化数据库与基础账号。"""

    with app.app_context():
        db.create_all()
        inspector = inspect(db.engine)
        existing_tables = set(inspector.get_table_names())
        product_columns = {col["name"] for col in inspector.get_columns("products")}
        enterprise_columns = {col["name"] for col in inspector.get_columns("enterprises")}
        def 补充审计软删除字段(表名):
            columns = {col["name"] for col in inspector.get_columns(表名)}
            if "created_at" not in columns:
                db.session.execute(text(f"ALTER TABLE {表名} ADD COLUMN created_at DATETIME"))
            if "updated_at" not in columns:
                db.session.execute(text(f"ALTER TABLE {表名} ADD COLUMN updated_at DATETIME"))
            if "created_by" not in columns:
                db.session.execute(text(f"ALTER TABLE {表名} ADD COLUMN created_by VARCHAR(100)"))
            if "updated_by" not in columns:
                db.session.execute(text(f"ALTER TABLE {表名} ADD COLUMN updated_by VARCHAR(100)"))
            if "deleted_at" not in columns:
                db.session.execute(text(f"ALTER TABLE {表名} ADD COLUMN deleted_at DATETIME"))
            if "is_deleted" not in columns:
                db.session.execute(text(f"ALTER TABLE {表名} ADD COLUMN is_deleted BOOLEAN DEFAULT 0"))
            db.session.execute(text(f"UPDATE {表名} SET created_at = COALESCE(created_at, CURRENT_TIMESTAMP)"))
            db.session.execute(text(f"UPDATE {表名} SET updated_at = COALESCE(updated_at, CURRENT_TIMESTAMP)"))
            db.session.execute(text(f"UPDATE {表名} SET is_deleted = COALESCE(is_deleted, 0)"))
            db.session.execute(text(f"CREATE INDEX IF NOT EXISTS ix_{表名}_deleted_at ON {表名} (deleted_at)"))
            db.session.execute(text(f"CREATE INDEX IF NOT EXISTS ix_{表名}_is_deleted ON {表名} (is_deleted)"))

        for 表名 in existing_tables:
            补充审计软删除字段(表名)
        db.session.commit()
        inspector = inspect(db.engine)
        product_columns = {col["name"] for col in inspector.get_columns("products")}
        enterprise_columns = {col["name"] for col in inspector.get_columns("enterprises")}
        if "enterprise_extra_fields" not in enterprise_columns:
            db.session.execute(text("ALTER TABLE enterprises ADD COLUMN enterprise_extra_fields JSON"))
            enterprise_columns.add("enterprise_extra_fields")
        if "main_products" not in enterprise_columns:
            db.session.execute(text("ALTER TABLE enterprises ADD COLUMN main_products TEXT"))
            enterprise_columns.add("main_products")
        if "province" not in enterprise_columns:
            db.session.execute(text("ALTER TABLE enterprises ADD COLUMN province VARCHAR(50)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_enterprises_province ON enterprises (province)"))
        if "city" not in enterprise_columns:
            db.session.execute(text("ALTER TABLE enterprises ADD COLUMN city VARCHAR(50)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_enterprises_city ON enterprises (city)"))
        enterprise_schema_additions = {
            "english_name": ("VARCHAR(255)", None),
            "unified_social_credit_code": ("VARCHAR(64)", "ix_enterprises_unified_social_credit_code"),
            "registered_name": ("VARCHAR(255)", "ix_enterprises_registered_name"),
            "legal_representative": ("VARCHAR(100)", None),
            "founded_date": ("DATE", None),
            "registered_capital": ("VARCHAR(100)", None),
            "business_term_start": ("DATE", None),
            "business_term_end": ("DATE", None),
            "registered_address": ("VARCHAR(255)", None),
            "business_scope": ("TEXT", None),
            "business_address": ("VARCHAR(255)", None),
            "district": ("VARCHAR(50)", None),
            "company_type": ("VARCHAR(100)", None),
            "industry_code": ("VARCHAR(50)", "ix_enterprises_industry_code"),
            "industry_category": ("VARCHAR(100)", "ix_enterprises_industry_category"),
            "sub_industry": ("VARCHAR(100)", None),
            "main_business": ("TEXT", None),
            "project_owner": ("VARCHAR(100)", None),
        }
        for 列名, (字段类型, 索引名) in enterprise_schema_additions.items():
            if 列名 not in enterprise_columns:
                db.session.execute(text(f"ALTER TABLE enterprises ADD COLUMN {列名} {字段类型}"))
                enterprise_columns.add(列名)
            if 索引名:
                db.session.execute(text(f"CREATE INDEX IF NOT EXISTS {索引名} ON enterprises ({列名})"))
        for 列名 in DEPRECATED_ENTERPRISE_EXTRA_FIELD_KEYS:
            if 列名 in enterprise_columns:
                db.session.execute(text(f"ALTER TABLE enterprises DROP COLUMN {列名}"))
                enterprise_columns.remove(列名)
        if "industry_code" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN industry_code VARCHAR(50)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_products_industry_code ON products (industry_code)"))
        if "industry_name" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN industry_name VARCHAR(100)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_products_industry_name ON products (industry_name)"))
        if "product_extra_fields" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN product_extra_fields JSON"))
        if "main_image" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN main_image VARCHAR(500)"))
        if "product_type" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN product_type VARCHAR(50)"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_products_product_type ON products (product_type)"))
        if "delivery_cycle" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN delivery_cycle VARCHAR(100)"))
            if "production_cycle" in product_columns:
                db.session.execute(
                    text(
                        "UPDATE products SET delivery_cycle=production_cycle "
                        "WHERE (delivery_cycle IS NULL OR delivery_cycle='') "
                        "AND production_cycle IS NOT NULL AND production_cycle!=''"
                    )
                )
        if "export_suitability" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN export_suitability VARCHAR(50)"))
            db.session.execute(
                text("CREATE INDEX IF NOT EXISTS ix_products_export_suitability ON products (export_suitability)")
            )
        if "recommendation_level" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN recommendation_level VARCHAR(50)"))
            db.session.execute(
                text("CREATE INDEX IF NOT EXISTS ix_products_recommendation_level ON products (recommendation_level)")
            )
        if "certification_status" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN certification_status VARCHAR(50)"))
            db.session.execute(
                text("CREATE INDEX IF NOT EXISTS ix_products_certification_status ON products (certification_status)")
            )
        if "brand" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN brand VARCHAR(100)"))
            product_columns.add("brand")
        if "model" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN model VARCHAR(100)"))
            product_columns.add("model")
        if "moq" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN moq VARCHAR(50)"))
            product_columns.add("moq")
        if "production_cycle" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN production_cycle VARCHAR(100)"))
            product_columns.add("production_cycle")
        if "capacity_cycle_days" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN capacity_cycle_days INTEGER"))
            product_columns.add("capacity_cycle_days")
        if "capacity_qualified_pieces" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN capacity_qualified_pieces INTEGER"))
            product_columns.add("capacity_qualified_pieces")
        if "sample_policy" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN sample_policy VARCHAR(255)"))
            product_columns.add("sample_policy")
        if "customization_supported" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN customization_supported BOOLEAN NOT NULL DEFAULT 0"))
            product_columns.add("customization_supported")
        if "exw_price" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN exw_price NUMERIC(18, 2)"))
            product_columns.add("exw_price")
        if "fob_price" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN fob_price NUMERIC(18, 2)"))
            product_columns.add("fob_price")
        if "cif_price" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN cif_price NUMERIC(18, 2)"))
            product_columns.add("cif_price")
        if "ddp_price" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN ddp_price NUMERIC(18, 2)"))
            product_columns.add("ddp_price")
        if "currency" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN currency VARCHAR(10)"))
            product_columns.add("currency")
        if "price_display" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN price_display VARCHAR(255)"))
            product_columns.add("price_display")
        if "status" not in product_columns:
            db.session.execute(text("ALTER TABLE products ADD COLUMN status VARCHAR(20) DEFAULT 'active'"))
            db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_products_status ON products (status)"))
            db.session.execute(text("UPDATE products SET status='active' WHERE status IS NULL OR status=''"))
        if "product_skus" not in existing_tables:
            db.session.execute(
                text(
                    """
                    CREATE TABLE IF NOT EXISTS product_skus (
                        id INTEGER PRIMARY KEY,
                        product_id INTEGER NOT NULL,
                        sku_code VARCHAR(64) NOT NULL,
                        sku_name VARCHAR(255) NOT NULL,
                        model VARCHAR(100),
                        specification TEXT,
                        color VARCHAR(100),
                        size VARCHAR(100),
                        material VARCHAR(255),
                        unit VARCHAR(20),
                        package_spec VARCHAR(255),
                        moq VARCHAR(50),
                        price NUMERIC(18, 2),
                        exw_price NUMERIC(18, 2),
                        fob_price NUMERIC(18, 2),
                        cif_price NUMERIC(18, 2),
                        ddp_price NUMERIC(18, 2),
                        gross_weight VARCHAR(100),
                        net_weight VARCHAR(100),
                        weight VARCHAR(100),
                        delivery_cycle VARCHAR(100),
                        currency VARCHAR(10),
                        stock_status VARCHAR(100),
                        sample_available BOOLEAN NOT NULL DEFAULT 0,
                        customization_supported BOOLEAN NOT NULL DEFAULT 0,
                        notes TEXT,
                        created_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        updated_at DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP,
                        FOREIGN KEY(product_id) REFERENCES products(id) ON DELETE CASCADE
                    )
                    """
                )
            )
        db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_product_skus_product_id ON product_skus (product_id)"))
        db.session.execute(text("CREATE INDEX IF NOT EXISTS ix_product_skus_sku_code ON product_skus (sku_code)"))
        db.session.commit()
        for 企业 in Enterprise.query.filter(Enterprise.enterprise_extra_fields.isnot(None)).all():
            清理前字段 = 企业.enterprise_extra_fields or {}
            清理后字段 = 清理企业扩展字段(清理前字段)
            if 清理后字段 != 清理前字段:
                企业.enterprise_extra_fields = 清理后字段
        db.session.commit()
        inspector = inspect(db.engine)
        sku_columns = {col["name"] for col in inspector.get_columns("product_skus")}

        def 补充SKU列(列名, 定义):
            if 列名 not in sku_columns:
                db.session.execute(text(f"ALTER TABLE product_skus ADD COLUMN {列名} {定义}"))
                sku_columns.add(列名)

        补充SKU列("unit", "VARCHAR(20)")
        补充SKU列("price", "NUMERIC(18, 2)")
        补充SKU列("exw_price", "NUMERIC(18, 2)")
        补充SKU列("fob_price", "NUMERIC(18, 2)")
        补充SKU列("cif_price", "NUMERIC(18, 2)")
        补充SKU列("ddp_price", "NUMERIC(18, 2)")
        补充SKU列("gross_weight", "VARCHAR(100)")
        补充SKU列("net_weight", "VARCHAR(100)")
        补充SKU列("weight", "VARCHAR(100)")
        补充SKU列("stock_status", "VARCHAR(100)")
        补充SKU列("sample_available", "BOOLEAN NOT NULL DEFAULT 0")
        db.session.execute(text("UPDATE product_skus SET price = fob_price WHERE price IS NULL AND fob_price IS NOT NULL"))
        db.session.commit()

        价格列 = [col for col in ["fob_price", "cif_price", "ddp_price", "exw_price"] if col in product_columns]
        if "price_display" in product_columns and 价格列:
            货币表达式 = "COALESCE(currency, 'USD')" if "currency" in product_columns else "'USD'"
            价格片段映射 = {
                "fob_price": f"CASE WHEN fob_price IS NOT NULL THEN 'FOB ' || printf('%.2f', fob_price) || ' ' || {货币表达式} END",
                "cif_price": f"CASE WHEN cif_price IS NOT NULL THEN 'CIF ' || printf('%.2f', cif_price) || ' ' || {货币表达式} END",
                "ddp_price": f"CASE WHEN ddp_price IS NOT NULL THEN 'DDP ' || printf('%.2f', ddp_price) || ' ' || {货币表达式} END",
                "exw_price": f"CASE WHEN exw_price IS NOT NULL THEN '出厂价 ' || printf('%.2f', exw_price) || ' ' || {货币表达式} END",
            }
            组合片段 = []
            for idx, 列名 in enumerate(价格列):
                if idx:
                    已有列条件 = " OR ".join(f"{已存在列} IS NOT NULL" for 已存在列 in 价格列[:idx])
                    组合片段.append(
                        f"CASE WHEN ({已有列条件}) AND {列名} IS NOT NULL THEN ' / ' ELSE '' END"
                    )
                组合片段.append(f"COALESCE({价格片段映射[列名]}, '')")
            显示文案表达式 = "\n                        || ".join(组合片段)
            存在价格条件 = " OR ".join(f"{列名} IS NOT NULL" for 列名 in 价格列)
            db.session.execute(
                text(
                    f"""
                    UPDATE products
                    SET price_display = TRIM(
                        {显示文案表达式}
                    )
                    WHERE (price_display IS NULL OR price_display = '')
                      AND ({存在价格条件})
                    """
                )
            )
            db.session.commit()

        db.session.execute(
            text(
                """
                UPDATE products
                SET industry_code = (
                    SELECT enterprises.industry_code FROM enterprises WHERE enterprises.id = products.enterprise_id
                ),
                    industry_name = (
                    SELECT enterprises.industry_category FROM enterprises WHERE enterprises.id = products.enterprise_id
                )
                WHERE (industry_code IS NULL OR industry_code = '')
                  AND enterprise_id IS NOT NULL
                """
            )
        )
        db.session.commit()

        管理员 = User.query.filter_by(username="admin").first()
        if not 管理员:
            db.session.add(User(username="admin", password=generate_password_hash("admin123"), role="管理员"))
        elif not 管理员.password.startswith("pbkdf2:") and not 管理员.password.startswith("scrypt:"):
            管理员.password = generate_password_hash("admin123")
            管理员.role = "管理员"
        if not User.query.filter_by(username="user").first():
            db.session.add(User(username="user", password=generate_password_hash("user123"), role="普通用户"))
        db.session.commit()

        app.config["DATABASE_BACKUP_ROOT"].mkdir(parents=True, exist_ok=True)
        app.config["FILES_BACKUP_ROOT"].mkdir(parents=True, exist_ok=True)


app = create_app()

if __name__ == "__main__":
    init_db(app)
    app.run(host="0.0.0.0", port=5000, debug=True)
